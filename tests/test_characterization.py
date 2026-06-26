"""
聚焦的 characterization 測試（重構安全網的「人類可讀」層）。

whole-workbook 黃金測試（test_golden.py）逐格鎖住整份輸出；本檔則用具名、
可讀的案例記錄「關鍵行為」，並明確區分兩類標籤：

  ✅ LOCK-CORRECT  ：鎖定目前『正確』的行為，重構不得破壞。
  ⚠️ LOCK-DEFECT   ：鎖定目前『已知錯誤』的行為作為基準；修補時須刻意更新此處。

公式案例直接針對現況真實邏輯 BlockManager._adjust_formula_references。
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from tests.fixtures import non_table_data
from tests.golden_compare import render_quietly
from tests.render_adapter import do_render, formula_adjuster

GOLDEN_DIR = Path(__file__).resolve().parent / "golden"


@pytest.fixture(scope="module")
def adjust():
    fn = formula_adjuster()
    if fn is None:
        pytest.skip("找不到活躍的公式平移函式；whole-workbook 黃金測試仍守護真實模板公式")
    return fn


@pytest.fixture(scope="module")
def non_table_output(tmp_path_factory):
    """以目前程式渲染 non_table 模板一次，供輸出層斷言共用。"""
    out = tmp_path_factory.mktemp("char") / "non_table.xlsx"
    render_quietly(lambda: do_render(
        Path(__file__).resolve().parent.parent / "template_non_table.xlsx",
        out,
        non_table_data(),
    ))
    return load_workbook(out)


# ===================================================================
# ✅ LOCK-CORRECT — 鎖定目前正確的行為（重構不得破壞）
# ===================================================================

class TestCorrectBehavior:
    def test_simple_row_formula_shift(self, adjust):
        # 表格展開使公式下移時，列參照應同步平移
        assert adjust("=SUM(A2:A10)", 2, 5) == "=SUM(A5:A13)"

    def test_multi_reference_shift(self, adjust):
        assert adjust("=A2+B2", 2, 5) == "=A5+B5"

    def test_multi_letter_column_shift(self, adjust):
        # 超過 Z 欄（多字母欄名）的參照亦能正確平移
        assert adjust("=SUM(AA2:AB10)", 2, 5) == "=SUM(AA5:AB13)"

    def test_absolute_reference_preserved(self, adjust):
        # 絕對列參照不應被平移
        assert adjust("=$A$2", 2, 5) == "=$A$2"
        assert adjust("=A$2", 2, 5) == "=A$2"

    def test_scalar_variable_substituted(self, non_table_output):
        ws = non_table_output["工作表1"]
        assert ws["B1"].value == "統一測試操作員"
        assert ws["E1"].value == "2025年01月01日-2025年01月31日統一測試"

    def test_table_data_and_header_rendered(self, non_table_output):
        ws = non_table_output["工作表1"]
        values = {c.value for row in ws.iter_rows() for c in row}
        assert "張三" in values   # DataFrame 資料已展開
        assert "姓名" in values   # 有 header 模式下欄名已渲染


# ===================================================================
# ⚠️ LOCK-DEFECT — 鎖定目前已知錯誤的行為作為基準
#   （這些斷言記錄「現況就是錯的」；日後修補時須刻意更新）
# ===================================================================

class TestKnownDefects:
    def test_function_name_with_digits_is_corrupted(self, adjust):
        # 缺陷：正則 ([A-Z]+)(\d+) 會把函式名內的數字當成列號平移
        # LOG10 -> LOG13（函式名被破壞）
        assert adjust("=LOG10(A2)", 2, 5) == "=LOG13(A5)"
        assert adjust("=ATAN2(A2,B2)", 2, 5) == "=ATAN5(A5,B5)"

    def test_cross_sheet_reference_over_shifted(self, adjust):
        # 缺陷：未檢查是否同一工作表，跨表參照被無條件平移
        assert adjust("=Sheet2!A2", 2, 5) == "=Sheet2!A5"

    def test_variable_below_expanded_tables_not_substituted(self, non_table_output):
        # 缺陷：位於展開表格下方的 {{city}} 雖有提供資料，卻未被替換
        # （標籤渲染時的位置追蹤未涵蓋被往下推移後的純量標籤）
        ws = non_table_output["工作表10"]
        assert ws["D22"].value == "{{city}}"


# ⚠️ 圖片「雙路徑可能重複位移」的疑慮：
#    真實模板的圖片錨點列已由 whole-workbook 黃金測試（test_golden.py）
#    逐一鎖住現況；若重構改變任一圖片錨點，黃金測試會直接亮紅燈。
