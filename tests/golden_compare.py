"""
黃金檔案比對工具。

將 .xlsx 序列化為穩定的「簽章」結構（值、數字格式、關鍵樣式、合併範圍、
Excel Table 範圍、圖片錨點），並比對兩份簽章、產出可讀的差異清單。

這是重構安全網的核心：相同模板 + 相同資料 + 相同程式 → 相同簽章。
任何差異都代表渲染輸出改變，須逐項判定為「改對」或「改壞」。
"""
from __future__ import annotations

import os
from contextlib import redirect_stderr, redirect_stdout

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _color_repr(color) -> str | None:
    """安全擷取顏色的穩定表示。"""
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str):
        return rgb
    theme = getattr(color, "theme", None)
    if theme is not None:
        return f"theme:{theme}:{getattr(color, 'tint', 0)}"
    return None


def _font_sig(font) -> tuple:
    if font is None:
        return ()
    return (
        font.name,
        font.size,
        bool(font.bold),
        bool(font.italic),
        _color_repr(font.color),
    )


def _fill_sig(fill) -> tuple:
    if fill is None:
        return ()
    pattern = getattr(fill, "patternType", None)
    if not pattern:
        return ()
    return (pattern, _color_repr(getattr(fill, "fgColor", None)))


def _align_sig(alignment) -> tuple:
    if alignment is None:
        return ()
    return (
        alignment.horizontal,
        alignment.vertical,
        bool(alignment.wrap_text),
    )


def _border_sig(border) -> tuple:
    if border is None:
        return ()
    return tuple(
        getattr(getattr(border, side), "style", None)
        for side in ("left", "right", "top", "bottom")
    )


def _cell_sig(cell) -> dict:
    sig = {
        "value": cell.value,
        "data_type": cell.data_type,
        "number_format": cell.number_format,
    }
    if cell.has_style:
        sig["font"] = _font_sig(cell.font)
        sig["fill"] = _fill_sig(cell.fill)
        sig["alignment"] = _align_sig(cell.alignment)
        sig["border"] = _border_sig(cell.border)
    return sig


def _image_sigs(ws) -> list:
    sigs = []
    for img in getattr(ws, "_images", []):
        anchor = getattr(img, "anchor", None)
        entry = {"anchor_type": type(anchor).__name__ if anchor else None}
        frm = getattr(anchor, "_from", None)
        if frm is not None:
            entry["from"] = (getattr(frm, "row", None), getattr(frm, "col", None))
        to = getattr(anchor, "to", None)
        if to is not None:
            entry["to"] = (getattr(to, "row", None), getattr(to, "col", None))
        sigs.append(entry)
    # 以錨點位置排序，避免順序造成假性差異
    sigs.sort(key=lambda e: (str(e.get("from")), str(e.get("to"))))
    return sigs


def _sheet_sig(ws) -> dict:
    cells = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None or cell.has_style:
                cells[cell.coordinate] = _cell_sig(cell)

    tables = {}
    for name in ws.tables:
        table = ws.tables[name]
        af = getattr(table, "autoFilter", None)
        tables[name] = {
            "ref": getattr(table, "ref", None),
            "autofilter_ref": getattr(af, "ref", None) if af else None,
        }

    return {
        "dimensions": ws.dimensions,
        "cells": cells,
        "merged_cells": sorted(str(r) for r in ws.merged_cells.ranges),
        "tables": tables,
        "images": _image_sigs(ws),
    }


def workbook_signature(path) -> dict:
    """載入 .xlsx 並回傳穩定簽章結構。"""
    wb = load_workbook(str(path))
    return {title: _sheet_sig(wb[title]) for title in wb.sheetnames}


def _diff(a, b, path, out, limit):
    if len(out) >= limit:
        return
    if isinstance(a, dict) and isinstance(b, dict):
        for key in sorted(set(a) | set(b)):
            if key not in a:
                out.append(f"{path}.{key}: 僅出現於實際輸出 = {b[key]!r}")
            elif key not in b:
                out.append(f"{path}.{key}: 僅出現於黃金基準 = {a[key]!r}")
            else:
                _diff(a[key], b[key], f"{path}.{key}", out, limit)
            if len(out) >= limit:
                return
    elif a != b:
        out.append(f"{path}: 黃金={a!r} vs 實際={b!r}")


def compare_signatures(golden: dict, actual: dict, limit: int = 50) -> list:
    """比對黃金 vs 實際簽章，回傳人類可讀的差異字串清單（最多 limit 筆）。"""
    out: list = []
    _diff(golden, actual, "", out, limit)
    return out


def render_quietly(render_callable):
    """執行渲染並抑制現況程式的大量 print 除錯輸出（stdout/stderr）。"""
    with open(os.devnull, "w") as devnull:
        with redirect_stdout(devnull), redirect_stderr(devnull):
            return render_callable()
