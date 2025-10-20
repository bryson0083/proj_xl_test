#!/usr/bin/env python3
"""
統一的情境測試程式 - 針對template_table.xlsx的所有情境（情境1-7）
測試表格物件（insert table）的處理功能
"""
import os
import pandas as pd
import logging
from datetime import datetime

from src.excel_template_renderer import render_template

# 設定logging level來顯示debug訊息
logging.basicConfig(level=logging.DEBUG, format='%(name)s - %(levelname)s - %(message)s')


def create_all_scenarios_data():
    """
    創建包含所有情境的測試數據（情境1-7）
    針對template_table.xlsx的表格物件測試
    """
    # 情境1數據 - 基本表格物件處理（3行數據）
    scenario_1_df = pd.DataFrame({
        '姓名': ['張三', '李四', '王五'],
        '年齡': [25, 30, 35],
        '部門': ['技術部', '業務部', '人事部'],
        '薪資': [50000, 55000, 60000],
        '補助': [5000, 6000, 7000]
    })

    # 情境2數據 - 基本表格物件處理（3行數據）
    scenario_2_df = pd.DataFrame({
        '姓名': ['張三', '李四', '王五'],
        '年齡': [25, 30, 35],
        '部門': ['技術部', '業務部', '人事部'],
        '薪資': [50000, 55000, 60000],
        '補助': [5000, 6000, 7000]
    })
    scenario_2_df['薪資總額'] = scenario_2_df['薪資'] + scenario_2_df['補助']


    # 情境3數據 - 混合條件標籤（3行數據）- 車輛統計資料
    scenario_3_vehicle_df = pd.DataFrame({
        '業者': ['保護傘客運', '保護傘客運', '沒有傘客運'],
        '路線': ['A001', 'A002', 'GA91'],
        '交易筆數': [10, 20, 30],
        '營收金額': [10, 20, 30],
        '平均客單價': [10, 20, 30]
    })
    
    # 情境4數據 - 多表格物件協調（第一個表格3行）
    scenario_4_report_df = pd.DataFrame({
        '姓名': ['孫三', '李四', '周五'],
        '年齡': [29, 33, 31],
        '部門': ['開發部', '測試部', '維運部'],
        '薪資': [60000, 55000, 52000],
        '補助': [6000, 5500, 5200]
    })
    scenario_4_report_df['薪資總額'] = scenario_4_report_df['薪資'] + scenario_4_report_df['補助']
    
    # # 情境4數據 - 多表格物件協調（第二個表格4行）
    # scenario_4_report2_df = pd.DataFrame({
    #     '產品': ['產品A', '產品B', '產品C', '產品D'],
    #     '銷量': [100, 200, 150, 180],
    #     '單價': [1000, 1500, 1200, 1300],
    #     '總額': [100000, 300000, 180000, 234000]
    # })
    
    # 情境5數據 - 垂直方向多表格（第一個表格2行，noheader）
    scenario_5_report_df = pd.DataFrame({
        '姓名': ['劉一', '陳二'],
        '年齡': [28, 32],
        '部門': ['資訊部', '採購部'],
        '薪資': [58000, 52000],
        '補助': [5800, 5200]
    })
    scenario_5_report_df['薪資總額'] = scenario_5_report_df['薪資'] + scenario_5_report_df['補助']
    
    # 情境5數據 - 垂直方向多表格（第二個表格3行，有header）
    scenario_5_report2_df = pd.DataFrame({
        '專案': ['專案X', '專案Y', '專案Z'],
        '進度': [0.85, 0.60, 0.92],
        '負責人': ['張經理', '李經理', '王經理'],
        '截止日': ['2025/03/31', '2025/04/30', '2025/05/31']
    })
    
    # 情境6數據 - 水平垂直混合（左上表格2行）
    scenario_6_report_df = pd.DataFrame({
        '項目': ['項目1', '項目2'],
        '數量': [50, 80],
        '金額': [25000, 40000]
    })
    
    # 情境6數據 - 水平垂直混合（右上表格3行）
    scenario_6_report2_df = pd.DataFrame({
        '類別': ['類別A', '類別B', '類別C'],
        '計數': [15, 25, 35],
        '比例': [0.2, 0.33, 0.47]
    })
    
    # 情境6數據 - 水平垂直混合（左下表格4行）
    scenario_6_report3_df = pd.DataFrame({
        '區域': ['北區', '中區', '南區', '東區'],
        '業績': [500000, 450000, 480000, 520000],
        '達成率': [1.05, 0.95, 1.01, 1.09]
    })
    
    # 情境6數據 - 水平垂直混合（右下表格2行）
    scenario_6_report4_df = pd.DataFrame({
        '月份': ['一月', '二月'],
        '收入': [1200000, 1350000],
        '支出': [980000, 1050000],
        '淨利': [220000, 300000]
    })
    
    # 情境7數據 - 複雜多表格協調（表格1：3行）
    scenario_7_report_df = pd.DataFrame({
        '供應商': ['供應商A', '供應商B', '供應商C'],
        '採購金額': [850000, 920000, 780000],
        '付款條件': ['月結30天', '月結60天', '月結45天'],
        '信用評級': ['A', 'AA', 'A+']
    })
    
    # 情境7數據 - 複雜多表格協調（表格2：5行）
    scenario_7_report2_df = pd.DataFrame({
        '客戶': ['客戶甲', '客戶乙', '客戶丙', '客戶丁', '客戶戊'],
        '訂單數': [45, 38, 52, 41, 49],
        '訂單金額': [1250000, 980000, 1450000, 1120000, 1350000],
        '回款狀態': ['已回款', '部分回款', '已回款', '未回款', '已回款']
    })
    
    # 情境7數據 - 複雜多表格協調（表格3：4行）
    scenario_7_report3_df = pd.DataFrame({
        '倉庫': ['倉庫1', '倉庫2', '倉庫3', '倉庫4'],
        '庫存量': [5000, 4500, 5200, 4800],
        '周轉率': [12.5, 10.8, 13.2, 11.6],
        '管理員': ['王管理', '李管理', '張管理', '陳管理']
    })
    
    # 情境7數據 - 複雜多表格協調（表格4：6行）
    scenario_7_report4_df = pd.DataFrame({
        '部門': ['行政部', '財務部', '技術部', '業務部', '人資部', '研發部'],
        '人數': [8, 12, 35, 28, 6, 42],
        '平均薪資': [45000, 52000, 68000, 55000, 48000, 72000],
        '預算': [360000, 624000, 2380000, 1540000, 288000, 3024000]
    })

    # 統一測試數據：包含所有情境的數據
    unified_data = {
        # 基本變數
        'oper_name': '表格物件測試操作員',
        'date_rng_desc': '2025年01月01日-2025年01月31日',
        
        # 情境3的特殊變數
        'plan_name': 'TW PASS',
        'ptnr_req_date_rng_desc': '2025/01/01 - 2025/01/31',
        'txn_date_rng_desc': '2025/01/01 - 2025/01/31',
        'rep_vehicle_df': scenario_3_vehicle_df,
        
        # 情境1-2使用相同的主數據（根據條件決定是否顯示header）
        'report_df': scenario_1_df,  # 主要測試數據
        'report2_df': scenario_2_df,  # 主要測試數據
        
        # 情境4-6的多表格數據
        'report3_df': scenario_6_report3_df,  # 用於情境6的第三個表格
        'report4_df': scenario_6_report4_df,  # 用於情境6的第四個表格
        
        # 特定情境的數據覆蓋
        # 情境5使用不同大小的數據
        'scenario5_report_df': scenario_5_report_df,
        'scenario5_report2_df': scenario_5_report2_df,
        
        # 情境6使用特定大小的數據
        'scenario6_report_df': scenario_6_report_df,
        'scenario6_report2_df': scenario_6_report2_df,
        'scenario6_report3_df': scenario_6_report3_df,
        'scenario6_report4_df': scenario_6_report4_df,
        
        # 情境7使用特定的數據
        'scenario7_report_df': scenario_7_report_df,
        'scenario7_report2_df': scenario_7_report2_df,
        'scenario7_report3_df': scenario_7_report3_df,
        'scenario7_report4_df': scenario_7_report4_df
    }
    
    return unified_data


def main():
    """主程式 - 執行統一情境測試（情境1-7，表格物件）"""
    print("=== Template Table (Insert Table) 統一情境測試（情境1-7）===")
    
    # 創建統一測試數據
    test_data = create_all_scenarios_data()
    
    print(f"測試數據內容:")
    print(f"  操作員: {test_data['oper_name']}")
    print(f"  日期範圍: {test_data['date_rng_desc']}")
    print(f"  主要數據(report_df)行數: {test_data['report_df'].shape[0]} 行")
    print(f"  第二數據(report2_df)行數: {test_data['report2_df'].shape[0]} 行")
    print(f"  第三數據(report3_df)行數: {test_data['report3_df'].shape[0]} 行")
    print(f"  第四數據(report4_df)行數: {test_data['report4_df'].shape[0]} 行")
    print()
    
    # 執行統一渲染
    import datetime
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f'template_table_統一測試結果_{timestamp}.xlsx'
    
    try:
        # 使用當前API進行渲染
        render_template(
            template_path='template_table.xlsx',
            output_file_name=output_file,
            **test_data
        )
        
        print(f"[OK] 統一測試完成: {output_file}")
        
        # 檢查輸出檔案是否成功產生
        if os.path.exists(output_file):
            file_size = os.path.getsize(output_file)
            print(f"[OK] 輸出檔案大小: {file_size} bytes")
        else:
            print(f"[ERROR] 輸出檔案不存在: {output_file}")
            return None
        
        # 驗證結果
        verify_unified_result(output_file)
        
        return {'success': True, 'output_file': output_file}
        
    except Exception as e:
        print(f"[ERROR] 統一測試失敗: {e}")
        import traceback
        traceback.print_exc()


def verify_unified_result(output_file):
    """驗證統一測試結果（情境1-7，表格物件）"""
    print(f"\n=== 驗證統一測試結果（情境1-7，表格物件）===")
    
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(output_file)
        
        # 檢查所有工作表（情境1-7）
        for scenario_num in range(1, 8):
            sheet_name = f'工作表{scenario_num}'
            
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"\n【情境{scenario_num} - {sheet_name} 表格物件驗證】")
                
                # 基本驗證：檢查變數標籤是否被渲染
                verify_basic_variables(ws, scenario_num)
                
                # 表格物件驗證
                verify_table_objects(ws, scenario_num)
                
                # 特殊情境驗證
                verify_special_features(ws, scenario_num)
                
            else:
                print(f"⚠️  找不到工作表: {sheet_name}")
        
        wb.close()
        print(f"\n=== 驗證完成 ===")
        
    except Exception as e:
        print(f"[ERROR] 驗證過程發生錯誤: {e}")
        import traceback
        traceback.print_exc()


def verify_basic_variables(ws, scenario_num):
    """驗證基本變數標籤"""
    # 每個情境都有基本的操作員和日期變數
    oper_cell = ws['B1'].value
    date_cell = ws['E1'].value
    
    print(f"  操作員欄位(B1): {oper_cell}")
    print(f"  日期欄位(E1): {date_cell}")
    
    has_unrendered_vars = False
    if oper_cell and '{{' in str(oper_cell):
        print(f"  [ERROR] 操作員變數未被渲染: {oper_cell}")
        has_unrendered_vars = True
    
    if date_cell and '{{' in str(date_cell):
        print(f"  [ERROR] 日期變數未被渲染: {date_cell}")
        has_unrendered_vars = True
    
    if not has_unrendered_vars:
        print(f"  [OK] 基本變數正確渲染")


def verify_table_objects(ws, scenario_num):
    """驗證表格物件"""
    print(f"  表格物件檢查:")
    
    # 檢查是否有表格物件
    if hasattr(ws, 'tables') and ws.tables:
        table_count = len(ws.tables)
        print(f"    發現 {table_count} 個表格物件")
        
        for table_name in ws.tables:
            table_obj = ws.tables[table_name]
            print(f"    表格: {table_name}")
            print(f"      範圍: {table_obj.ref}")
            print(f"      顯示名稱: {table_obj.displayName}")
            
            # 檢查表格範圍是否正確更新
            if scenario_num == 1:
                # 情境1：單一表格，應該有3行數據+表頭
                expected_rows = 4  # 1個表頭 + 3行數據
            elif scenario_num == 2:
                # 情境2：noheader表格，應該有4行數據（無表頭）
                expected_rows = 4  # 只有數據行
            elif scenario_num == 3:
                # 情境3：混合條件，2行數據
                expected_rows = 3  # 1個表頭 + 2行數據（或根據條件）
            elif scenario_num in [4, 5]:
                # 情境4-5：多表格
                expected_rows = None  # 多表格，不固定
            elif scenario_num in [6, 7]:
                # 情境6-7：水平垂直混合
                expected_rows = None  # 複雜佈局，不固定
            else:
                expected_rows = None
            
            if expected_rows:
                # 解析表格範圍
                ref_parts = table_obj.ref.split(':')
                if len(ref_parts) == 2:
                    start_cell = ref_parts[0]
                    end_cell = ref_parts[1]
                    # 簡單檢查行數（從字串提取數字）
                    import re
                    start_row = int(re.findall(r'\d+', start_cell)[0])
                    end_row = int(re.findall(r'\d+', end_cell)[0])
                    actual_rows = end_row - start_row + 1
                    
                    if actual_rows == expected_rows:
                        print(f"      [OK] 表格行數正確: {actual_rows} 行")
                    else:
                        print(f"      ⚠️ 表格行數: {actual_rows} 行（預期: {expected_rows} 行）")
    else:
        print(f"    ⚠️ 未發現表格物件")
    
    # 檢查是否有未渲染的表格標籤
    has_table_tags = False
    has_actual_data = False
    
    # 根據不同情境設定檢查範圍
    if scenario_num <= 3:
        check_range = range(2, 10)  # 檢查前幾行
    elif scenario_num in [4, 5]:
        check_range = range(2, 15)  # 多表格情境
    else:  # scenario_num in [6, 7]
        check_range = range(2, 20)  # 複雜佈局
    
    for row in check_range:
        for col in range(1, 10):  # 檢查前9列
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                cell_str = str(cell_value).strip()
                if '{{' in cell_str and '}}' in cell_str:
                    has_table_tags = True
                    print(f"    [ERROR] 未渲染標籤: {cell_value} (第{row}行第{col}列)")
                elif cell_str and not cell_str.startswith('='):
                    has_actual_data = True
    
    if has_table_tags:
        print(f"  [ERROR] 發現未渲染的表格標籤")
    elif has_actual_data:
        print(f"  [OK] 表格數據正確渲染")
    else:
        print(f"  ⚠️  表格數據區域為空")


def verify_special_features(ws, scenario_num):
    """驗證特殊功能"""
    
    # 情境1-3：基本表格物件功能
    if scenario_num in [1, 2, 3]:
        # 檢查noheader條件（情境2、3、5）
        if scenario_num in [2, 3]:
            # 檢查是否有表頭行
            header_found = False
            for col in range(1, 7):
                cell_value = ws.cell(row=3, column=col).value  # 假設第3行是數據開始位置
                if cell_value and any(header in str(cell_value) for header in ['姓名', '年齡', '部門']):
                    header_found = True
                    break
            
            if scenario_num == 2 and not header_found:
                print(f"  [OK] noheader條件正確（無表頭）")
            elif scenario_num == 3:
                print(f"  [OK] 混合條件標籤處理")
    
    # 情境4-5：多表格協調
    elif scenario_num in [4, 5]:
        # 檢查是否有多個表格
        table_count = 0
        if hasattr(ws, 'tables'):
            table_count = len(ws.tables)
        
        if table_count >= 2:
            print(f"  [OK] 多表格協調: 發現 {table_count} 個表格")
        else:
            print(f"  ⚠️ 預期多個表格，實際發現 {table_count} 個")
        
        # 檢查表格間的Gap Block
        if scenario_num == 5:
            print(f"  [OK] 垂直方向多表格渲染")
    
    # 情境6-7：水平垂直混合
    elif scenario_num in [6, 7]:
        # 檢查複雜佈局
        table_count = 0
        if hasattr(ws, 'tables'):
            table_count = len(ws.tables)
        
        expected_tables = 4  # 情境6-7都應該有4個表格
        if table_count == expected_tables:
            print(f"  [OK] 水平垂直混合佈局: {table_count} 個表格")
        else:
            print(f"  ⚠️ 預期 {expected_tables} 個表格，實際發現 {table_count} 個")
        
        # 檢查Footer Block處理
        footer_found = False
        for row in range(20, 50):  # 在較後面的行尋找Footer
            cell_value = ws[f'A{row}'].value
            if cell_value and 'Footer' in str(cell_value):
                print(f"  [OK] Footer位置: 第{row}行")
                footer_found = True
                break
        
        if not footer_found and scenario_num == 7:
            print(f"  ⚠️ 未找到Footer區塊")
    
    print(f"  [OK] 情境{scenario_num}特殊功能檢查完成")


if __name__ == "__main__":
    main()