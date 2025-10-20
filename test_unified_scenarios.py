#!/usr/bin/env python3
"""
統一的情境測試程式 - 針對template_non_table的所有情境（情境1-10）
只輸出一個測試檔案，包含所有情境的測試數據
"""
import os
import pandas as pd
from datetime import datetime

from src.excel_template_renderer.api import render_template


def create_all_scenarios_data():
    """
    創建包含所有情境的測試數據（情境1-10）
    將10個情境的數據合併到一個統一的數據集中
    """
    # 情境1數據 (3行)
    scenario_1_df = pd.DataFrame({
        '姓名': ['張三', '李四', '王五'],
        '年齡': [25, 30, 35],
        '部門': ['技術部', '業務部', '人事部'],
        '薪資': [50000, 55000, 60000],
        '補助': [5000, 6000, 7000]
    })
    scenario_1_df['薪資總額'] = scenario_1_df['薪資'] + scenario_1_df['補助']
    
    # 情境2數據 (4行)
    scenario_2_df = pd.DataFrame({
        '姓名': ['陳七', '周八', '吳九', '鄭十'],
        '年齡': [28, 32, 29, 26],
        '部門': ['技術部', '財務部', '業務部', '人事部'],
        '薪資': [52000, 58000, 51000, 49000],
        '補助': [5200, 5800, 5100, 4900]
    })
    scenario_2_df['薪資總額'] = scenario_2_df['薪資'] + scenario_2_df['補助']
    
    # 情境3數據 (2行)
    scenario_3_df = pd.DataFrame({
        '姓名': ['趙一', '錢二'],
        '年齡': [31, 27],
        '部門': ['研發部', '行銷部'],
        '薪資': [65000, 48000],
        '補助': [6500, 4800]
    })
    scenario_3_df['薪資總額'] = scenario_3_df['薪資'] + scenario_3_df['補助']
    
    # 情境4數據 (3行，與情境1類似但內容不同)
    scenario_4_df = pd.DataFrame({
        '姓名': ['孫三', '李四', '周五'],
        '年齡': [29, 33, 31],
        '部門': ['開發部', '測試部', '維運部'],
        '薪資': [60000, 55000, 52000],
        '補助': [6000, 5500, 5200]
    })
    scenario_4_df['薪資總額'] = scenario_4_df['薪資'] + scenario_4_df['補助']
    
    # 情境5數據 - 統計資料DataFrame (5行)
    scenario_5_stat_df = pd.DataFrame({
        # "年齡區間": ["20~30歲", "31~40歲", "41~50歲", "51~65歲", "66歲以上"],
        "人數": [45, 38, 52, 23, 8]
    })
    
    # 情境6數據 - 複雜多標籤混合
    scenario_6_rep_agg_df = pd.DataFrame({
        "路線": ["紅線", "藍線", "綠線"],
        "浣熊市市民定期票旅次數": [1200, 980, 1150],
        "浣熊市市民定期票補助金額": [24000, 19600, 23000],
        "浣熊市非市民定期票旅次數": [300, 250, 320],
        "浣熊市非市民定期票補助金額": [6000, 5000, 6400],
        "貓熊市民定期票旅次數": [150, 120, 180],
        "貓熊市民定期票補助金額": [3000, 2400, 3600],
        "貓熊非市民定期票旅次數": [80, 60, 90],
        "貓熊非市民定期票補助金額": [1600, 1200, 1800],
        "總補助金額": [34600, 28200, 34800]
    })
    
    # 情境7數據 - 複雜報表模板
    scenario_7_rep_vehicle_df = pd.DataFrame({
        "業者": ["A公司", "B公司", "C公司"],
        "行政區劃": ["中央區", "東區", "西區"],
        "營收款_比例": [0.3, 0.25, 0.45],
        "營收款_筆數": [1200, 980, 1350],
        "營收款_金額": [1500000, 1200000, 1800000],
        "分配票收金額": [1260000, 1050000, 1890000],
        "差額補貼款金額": [240000, 150000, -90000],
        "中央差額補貼款比例": [0.6, 0.6, 0.0],
        "中央差額補貼款金額": [144000, 90000, 0],
        "地方差額補貼款金額": [0, 0, 0],  # 新增欄位，預設零值
    })
    
    # 情境8數據 - 學生卡價差補貼申請表
    scenario_8_tb4_routeid_df = pd.DataFrame({
        "路線IC代碼": ["R001", "R002", "R003"],
        "路線編號": ["101", "102", "103"],
        "路線名稱": ["市中心線", "工業區線", "住宅區線"],
        "學生刷卡旅次量\nA=IC學生票刷卡量": [2500, 1800, 2200],
        "市民卡(學生卡)\n應收總金額(元)\nB=市民卡營收金額": [125000, 90000, 110000],
        "市民卡(學生卡)\n實際營收金額(元)\nC=學生票營收金額": [100000, 72000, 88000],
        "市民卡(學生卡)基本里程優惠金額(元)D": [12500, 9000, 11000],
        "票價上限補貼金額(元)E": [7500, 5400, 6600],
        "學生卡25折補助金額(元)F\nF=B-C-D-E-G-H-I-J": [5000, 3600, 4400],
        "捷運轉乘優惠補貼金額(元)G": [3000, 2160, 2640],
        "A22-A23區間優惠補貼金額(元)H": [0, 0, 0],
        "國道及台鐵轉乘市區客運優惠金額(元)I": [0, 0, 0],
        "其它補貼(元)J": [0, 0, 0]
    })
    
    # 情境9數據 - 多月份愛心卡統計（4個區域的數據）
    scenario_9_rep_area_disable_01_df = pd.DataFrame({
        "年齡": ["65-74", "75-84", "85+"],
        "公車次數": [1200, 800, 300],
        "公車點數": [24000, 16000, 6000],
        "愛心計程車次數": [150, 100, 50],
        "愛心計程車點數": [7500, 5000, 2500],
        "桃園捷運次數": [300, 200, 80],
        "桃園捷運點數": [6000, 4000, 1600],
        "台鐵次數": [80, 50, 20],
        "台鐵點數": [1600, 1000, 400],
        "運動中心次數": [40, 25, 10],
        "運動中心點數": [0, 0, 0],  # 補零值
        "活動中心次數": [0, 0, 0],   # 補零值
        "活動中心點數": [0, 0, 0]    # 補零值
    })
    
    scenario_9_rep_area_disable_02_df = pd.DataFrame({
        "年齡": ["65-74", "75-84", "85+"],
        "公車次數": [1150, 750, 280],
        "公車點數": [23000, 15000, 5600],
        "愛心計程車次數": [140, 95, 45],
        "愛心計程車點數": [7000, 4750, 2250],
        "桃園捷運次數": [280, 190, 75],
        "桃園捷運點數": [5600, 3800, 1500],
        "台鐵次數": [75, 45, 18],
        "台鐵點數": [1500, 900, 360],
        "運動中心次數": [38, 22, 8],
        "運動中心點數": [0, 0, 0],  # 補零值
        "活動中心次數": [0, 0, 0],   # 補零值
        "活動中心點數": [0, 0, 0]    # 補零值
    })
    
    scenario_9_rep_area_disable_03_df = pd.DataFrame({
        "年齡": ["65-74", "75-84", "85+"],
        "公車次數": [1300, 850, 320],
        "公車點數": [26000, 17000, 6400],
        "愛心計程車次數": [160, 110, 55],
        "愛心計程車點數": [8000, 5500, 2750],
        "桃園捷運次數": [320, 210, 85],
        "桃園捷運點數": [6400, 4200, 1700],
        "台鐵次數": [85, 55, 22],
        "台鐵點數": [1700, 1100, 440],
        "運動中心次數": [42, 28, 12],
        "運動中心點數": [0, 0, 0],  # 補零值
        "活動中心次數": [0, 0, 0],   # 補零值
        "活動中心點數": [0, 0, 0]    # 補零值
    })
    
    scenario_9_rep_area_disable_04_df = pd.DataFrame({
        "年齡": ["65-74", "75-84", "85+"],
        "公車次數": [1250, 820, 310],
        "公車點數": [25000, 16400, 6200],
        "愛心計程車次數": [155, 105, 52],
        "愛心計程車點數": [7750, 5250, 2600],
        "桃園捷運次數": [310, 200, 82],
        "桃園捷運點數": [6200, 4000, 1640],
        "台鐵次數": [82, 52, 21],
        "台鐵點數": [1640, 1040, 420],
        "運動中心次數": [41, 26, 11],
        "運動中心點數": [0, 0, 0],  # 補零值
        "活動中心次數": [0, 0, 0],   # 補零值
        "活動中心點數": [0, 0, 0]    # 補零值
    })
    
    # 情境10數據 - 複雜公共運輸定期票營收月報表（3個表格）
    scenario_10_partner_ticket_df = pd.DataFrame({
        "業者": ["甲客運", "乙客運", "丙客運"],
        "行政區劃": ["信義區", "大安區", "中山區"],
        "營收款": [2000000, 1800000, 2200000],
        "比例": [0.3, 0.27, 0.33],
        "筆數": [1500, 1350, 1650],
        "分配票收金額": [2040000, 1836000, 2244000],
        "差額補貼款金額": [-40000, -36000, -44000]
    })
    
    scenario_10_partner_center_df = pd.DataFrame({
        "業者": ["甲客運", "乙客運", "丙客運"],
        "行政區劃": ["信義區", "大安區", "中山區"],
        "中央差額補貼款比例": [0.0, 0.0, 0.0],
        "中央差額補貼款金額": [0, 0, 0]
    })
    
    scenario_10_partner_local_df = pd.DataFrame({
        "業者": ["甲客運", "乙客運", "丙客運"],
        "行政區劃": ["信義區", "大安區", "中山區"],
        "地方差額補貼款金額": [-40000, -36000, -44000]
    })

    # 統一測試數據：包含所有情境的數據
    unified_data = {
        # 基本變數（情境1-5）
        'oper_name': '統一測試操作員',
        'date_rng_desc': '2025年01月01日-2025年01月31日統一測試',
        
        # 主要數據使用情境1的數據
        'report_df': scenario_1_df,
        
        # 情境5數據 - 統計資料
        'stat_df': scenario_5_stat_df,
        
        # 情境6數據 - 複雜多標籤混合
        'ptnr_req_date_rng_desc': '2024年5月1日-2024年5月31日',
        'txn_date_rng_desc': '2024年5月1日-2024年5月31日',
        'rep_agg_df': scenario_6_rep_agg_df,
        
        # 情境7數據 - 複雜報表模板
        'plan_name': '城市通行',
        'rep_partnerreq_date': '2024年6月30日',
        'purchase_amt': 5000000,
        'all_cost': 800000,
        'rep_vehicle_df': scenario_7_rep_vehicle_df,
        
        # 情境8數據 - 學生卡價差補貼申請表
        'date_rng_desc1': '2024年7月',
        'tb4_routeid_df': scenario_8_tb4_routeid_df,
        
        # 情境9數據 - 多月份愛心卡統計
        'rep_year': '2024',
        'rep_area_disable_01_df': scenario_9_rep_area_disable_01_df,
        'rep_area_disable_02_df': scenario_9_rep_area_disable_02_df,
        'rep_area_disable_03_df': scenario_9_rep_area_disable_03_df,
        'rep_area_disable_04_df': scenario_9_rep_area_disable_04_df,
        
        # 情境10數據 - 複雜公共運輸定期票營收月報表
        'tpass_purchase_amt': 8000000,
        'tpass_all_cost': 1200000,
        'city': '台北市',
        'partner_ticket_df': scenario_10_partner_ticket_df,
        'partner_center_df': scenario_10_partner_center_df,
        'partner_local_df': scenario_10_partner_local_df
    }
    
    return unified_data


def main():
    """主程式 - 執行統一情境測試（情境1-10）"""
    print("=== Template Non-Table 統一情境測試（情境1-10）===")
    
    # 創建統一測試數據
    test_data = create_all_scenarios_data()
    
    print(f"測試數據內容:")
    print(f"  操作員: {test_data['oper_name']}")
    print(f"  日期範圍: {test_data['date_rng_desc']}")
    print(f"  主要數據行數: {test_data['report_df'].shape[0]} 行")
    print(f"  主要數據列數: {test_data['report_df'].shape[1]} 列")
    print(f"  統計數據行數: {test_data['stat_df'].shape[0]} 行")
    print(f"  統計數據列數: {test_data['stat_df'].shape[1]} 列")
    print(f"  情境6聚合數據行數: {test_data['rep_agg_df'].shape[0]} 行")
    print(f"  情境7車輛數據行數: {test_data['rep_vehicle_df'].shape[0]} 行")
    print(f"  情境8路線數據行數: {test_data['tb4_routeid_df'].shape[0]} 行")
    print(f"  情境9愛心卡數據(區域1)行數: {test_data['rep_area_disable_01_df'].shape[0]} 行")
    print(f"  情境10合作夥伴票券數據行數: {test_data['partner_ticket_df'].shape[0]} 行")
    print()
    
    # 執行統一渲染
    import datetime
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f'template_non_table_統一測試結果_{timestamp}.xlsx'
    
    try:
        # 使用新版API並保留註冊表檔案
        result = render_template(
            template_path='template_non_table.xlsx',
            output_file_name=output_file,
            validate_result=True,  # 啟用驗證
            **test_data
        )
        
        print(f"[OK] 統一測試完成: {output_file}")
        print(f"[OK] 註冊表檔案: {result['registry_file']}")
        print(f"[OK] 渲染狀態: {result['success']}")
        
        if 'validation_result' in result and result['validation_result']:
            print(f"[OK] 驗證結果: 通過")
        
        # 驗證結果
        verify_unified_result(output_file)
        
        # 顯示註冊表檔案位置（確保不被刪除）
        print(f"\n=== 註冊表檔案保留 ===")
        registry_file = result['registry_file']
        if os.path.exists(registry_file):
            print(f"註冊表檔案已保留: {registry_file}")
            # 顯示檔案大小
            file_size = os.path.getsize(registry_file)
            print(f"檔案大小: {file_size} bytes")
        else:
            print(f"[ERROR] 註冊表檔案不存在: {registry_file}")
        
        return result
        
    except Exception as e:
        print(f"[ERROR] 統一測試失敗: {e}")
        import traceback
        traceback.print_exc()


def verify_unified_result(output_file):
    """驗證統一測試結果（情境1-10）"""
    print(f"\n=== 驗證統一測試結果（情境1-10）===")
    
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(output_file)
        
        # 檢查所有工作表（情境1-10）
        for scenario_num in range(1, 11):
            sheet_name = f'工作表{scenario_num}'
            
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"\n【情境{scenario_num} - {sheet_name} 驗證】")
                
                # 基本驗證：檢查變數標籤是否被渲染
                verify_basic_variables(ws, scenario_num)
                
                # 表格數據驗證
                verify_table_data(ws, scenario_num)
                
                # 特殊情境驗證
                verify_special_features(ws, scenario_num)
                
            else:
                print(f"⚠️  找不到工作表: {sheet_name}")
        
        wb.close()
        print(f"\n=== 驗證完成 ===")
        
    except Exception as e:
        print(f"[ERROR] 驗證過程發生錯誤: {e}")
        import traceback
        traceback.print_exc()


def verify_basic_variables(ws, scenario_num):
    """驗證基本變數標籤"""
    # 每個情境都有基本的操作員和日期變數
    
    # 根據情境檢查不同的欄位位置
    if scenario_num == 8:
        # 情境8: A1位置包含了date_rng_desc1和oper_name兩個標籤
        cell_value = ws['A1'].value
        print(f"  A1標題欄位: {cell_value}")
        
        has_unrendered_vars = False
        if cell_value and '{{' in str(cell_value):
            print(f"  [ERROR] A1變數未被渲染: {cell_value}")
            has_unrendered_vars = True
        
        if not has_unrendered_vars:
            print(f"  [OK] 基本變數正確渲染")
        return
    
    # 其他情境的標準檢查
    oper_cell = ws['B1'].value
    
    # 根據情境檢查不同的日期欄位
    if scenario_num == 6:
        date_cell = ws['E1'].value  # ptnr_req_date_rng_desc
    elif scenario_num == 7:
        date_cell = ws['E1'].value  # rep_partnerreq_date
    elif scenario_num == 9:
        date_cell = ws['B2'].value  # rep_year
    elif scenario_num == 10:
        date_cell = ws['E1'].value  # rep_partnerreq_date
    else:
        date_cell = ws['E1'].value  # date_rng_desc
    
    print(f"  操作員欄位: {oper_cell}")
    print(f"  日期欄位: {date_cell}")
    
    has_unrendered_vars = False
    if oper_cell and '{{' in str(oper_cell):
        print(f"  [ERROR] 操作員變數未被渲染: {oper_cell}")
        has_unrendered_vars = True
    
    if date_cell and '{{' in str(date_cell):
        print(f"  [ERROR] 日期變數未被渲染: {date_cell}")
        has_unrendered_vars = True
    
    if not has_unrendered_vars:
        print(f"  [OK] 基本變數正確渲染")


def verify_table_data(ws, scenario_num):
    """驗證表格數據"""
    print(f"  表格數據檢查:")
    
    # 檢查是否有未渲染的表格標籤
    has_table_tags = False
    has_actual_data = False
    
    # 根據不同情境設定檢查範圍
    if scenario_num <= 5:
        check_range = range(2, 10)  # 檢查前幾行
    elif scenario_num == 6:
        check_range = range(2, 8)   # 情境6的表格較小
    elif scenario_num in [7, 8]:
        check_range = range(2, 12)  # 情境7,8有較多數據
    elif scenario_num == 9:
        check_range = range(2, 20)  # 情境9有多個表格
    else:  # scenario_num == 10
        check_range = range(2, 25)  # 情境10最複雜
    
    for row in check_range:
        for col in range(1, 15):  # 檢查前14列
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                cell_str = str(cell_value).strip()
                if '{{' in cell_str and '}}' in cell_str:
                    has_table_tags = True
                    print(f"    [ERROR] 未渲染標籤: {cell_value} (第{row}行第{col}列)")
                elif cell_str and not cell_str.startswith('=') and not cell_str.startswith('Report'):
                    has_actual_data = True
    
    if has_table_tags:
        print(f"  [ERROR] 發現未渲染的表格標籤")
    elif has_actual_data:
        print(f"  [OK] 表格數據正確渲染")
    else:
        print(f"  ⚠️  表格數據區域為空")


def verify_special_features(ws, scenario_num):
    """驗證特殊功能"""
    # 檢查Footer位置（情境3,4有Footer）
    if scenario_num in [3, 4]:
        footer_found = False
        for row in range(1, 30):
            cell_value = ws[f'A{row}'].value
            if cell_value and 'Report Footer' in str(cell_value):
                print(f"  [OK] Footer位置: 第{row}行")
                footer_found = True
                break
        
        if not footer_found:
            print(f"  ⚠️  找不到Footer")
    
    # 檢查公式（情境2,4,7,8,10有公式）
    if scenario_num in [2, 4, 7, 8, 10]:
        formula_found = False
        for row in range(1, 20):
            for col in range(1, 10):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and str(cell_value).startswith('='):
                    formula_found = True
                    break
            if formula_found:
                break
        
        if formula_found:
            print(f"  [OK] 發現公式計算")
        else:
            print(f"  ⚠️  未發現公式")
    
    # 檢查多表格情境（情境9,10）
    if scenario_num in [9, 10]:
        table_count = 0
        for row in range(1, 50):  # 擴大搜尋範圍
            for col in range(1, 15):
                cell_value = ws.cell(row=row, column=col).value
                # 簡單計算：如果發現連續的非空數據行，視為一個表格
                if cell_value and str(cell_value).strip() and not str(cell_value).startswith('='):
                    # 檢查這一行是否像表格的開始
                    next_cell = ws.cell(row=row, column=col+1).value if col < 14 else None
                    if next_cell and str(next_cell).strip():
                        table_count += 1
                        break
        
        if table_count >= 3:  # 情境9,10都應該有多個表格
            print(f"  [OK] 發現多個表格結構 (約{table_count}個)")
        else:
            print(f"  ⚠️  表格數量可能不足")
    
    print(f"  [OK] 情境{scenario_num}特殊功能檢查完成")


if __name__ == "__main__":
    main()
