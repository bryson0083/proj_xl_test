"""
容器管理器
"""
import uuid
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from ..models.base import BlockType, ObjectType, CellPosition, DataShape, RangePosition
from ..models.container import Container
from ..models.objects import Block, ObjectInfo, TableObject, ImageObject
from ..models.tag import Tag
from .template_scanner import TemplateScanner


class ContainerManager:
    """
    容器管理器類別
    
    負責建立和管理Container物件，包含物件註冊表功能
    現在使用TemplateScanner進行統一的掃描與註冊
    """
    
    def __init__(self):
        """初始化容器管理器"""
        self.template_scanner = TemplateScanner()
    
    def create_containers(self, workbook: Workbook, tags: Optional[List[Tag]] = None) -> List[Container]:
        """
        建立容器物件清單，每個Sheet對應一個Container
        
        Args:
            workbook: Excel工作簿物件
            tags: 解析出的標籤清單（可選，如果未提供將自動掃描）
            
        Returns:
            List[Container]: 容器物件清單
        """
        # 使用TemplateScanner進行統一掃描與註冊
        return self.template_scanner.scan_and_register_template(workbook)
    
    def _group_tags_by_sheet(self, workbook: Workbook, tags: List[Tag]) -> Dict[str, List[Tag]]:
        """
        按工作表分組標籤
        
        Args:
            workbook: Excel工作簿
            tags: 標籤清單
            
        Returns:
            Dict[str, List[Tag]]: 按工作表分組的標籤
        """
        sheet_tags = {}
        
        # 初始化每個工作表的標籤清單
        for sheet_name in workbook.sheetnames:
            sheet_tags[sheet_name] = []
        
        # 根據標籤的sheet_name資訊分組
        for tag in tags:
            if tag.sheet_name in sheet_tags:
                sheet_tags[tag.sheet_name].append(tag)
            
        return sheet_tags
    
    def _create_single_container(self, workbook: Workbook, sheet_name: str, tags: List[Tag]) -> Container:
        """
        建立單一容器物件
        
        Args:
            workbook: Excel工作簿
            sheet_name: 工作表名稱
            tags: 該工作表的標籤清單
            
        Returns:
            Container: 容器物件
        """
        container_id = f"container_{sheet_name}_{uuid.uuid4().hex[:8]}"
        worksheet = workbook[sheet_name]
        
        # 掃描並建立物件清單
        objects = []
        
        # 從標籤建立物件
        for tag in tags:
            obj_info = self._create_object_from_tag(tag, sheet_name)
            objects.append(obj_info)
        
        # 掃描表格物件
        table_objects = self._scan_table_objects(worksheet, sheet_name)
        objects.extend(table_objects)
        
        # 掃描圖片物件
        image_objects = self._scan_image_objects(worksheet, sheet_name)
        objects.extend(image_objects)
        
        # 掃描非標籤的文字內容（用於Footer等區域）
        text_objects = self._scan_text_content(worksheet, sheet_name, tags)
        objects.extend(text_objects)
        
        # 初始化空的區塊清單（稍後由classify_blocks填入）
        blocks = []
        
        return Container(
            container_id=container_id,
            sheet_name=sheet_name,
            blocks=blocks,
            objects=objects
        )
    
    def _create_object_from_tag(self, tag: Tag, sheet_name: str) -> ObjectInfo:
        """
        從標籤建立物件資訊
        
        Args:
            tag: 標籤物件
            sheet_name: 工作表名稱
            
        Returns:
            ObjectInfo: 物件資訊
        """
        obj_id = f"tag_{tag.tag_name}_{uuid.uuid4().hex[:8]}"
        
        # 根據標籤類型決定物件類型
        if tag.tag_type.value == "table":
            obj_type = ObjectType.TABLE
            is_multi_rows = True
        else:
            obj_type = ObjectType.SIMPLE
            is_multi_rows = False
        
        return ObjectInfo(
            obj_id=obj_id,
            display_name=tag.tag_name,
            obj_type=obj_type,
            block_id="",  # 稍後由classify_blocks設定
            sheet_name=sheet_name,
            is_multi_rows=is_multi_rows,
            having_header=not (tag.has_condition and tag.condition == "noheader"),
            cell_position=tag.cell_position,
            data_shape=DataShape(rows=1, cols=1)  # 初始值，渲染時會更新
        )
    
    def _scan_table_objects(self, worksheet, sheet_name: str) -> List[ObjectInfo]:
        """
        掃描工作表中的表格物件
        
        Args:
            worksheet: 工作表物件
            sheet_name: 工作表名稱
            
        Returns:
            List[ObjectInfo]: 表格物件清單
        """
        objects = []
        
        # 掃描Excel表格物件
        for table in worksheet.tables.values():
            obj_id = f"table_obj_{table.name}_{uuid.uuid4().hex[:8]}"
            
            # 解析表格範圍
            table_range = table.ref
            start_cell, end_cell = table_range.split(':')
            
            # 簡單解析座標（A1格式）
            start_row, start_col = self._parse_cell_reference(start_cell)
            
            obj_info = ObjectInfo(
                obj_id=obj_id,
                display_name=table.name,
                obj_type=ObjectType.TABLE_OBJ,
                block_id="",  # 稍後設定
                sheet_name=sheet_name,
                is_multi_rows=True,
                having_header=True,  # 表格物件通常有表頭
                cell_position=CellPosition(row=start_row, col=start_col),
                data_shape=DataShape(rows=1, cols=1)  # 稍後計算
            )
            objects.append(obj_info)
            
        return objects
    
    def _scan_image_objects(self, worksheet, sheet_name: str) -> List[ObjectInfo]:
        """
        掃描工作表中的圖片物件
        
        Args:
            worksheet: 工作表物件
            sheet_name: 工作表名稱
            
        Returns:
            List[ObjectInfo]: 圖片物件清單
        """
        objects = []
        
        # 掃描圖片物件
        for image in worksheet._images:
            obj_id = f"image_obj_{uuid.uuid4().hex[:8]}"
            
            # 取得圖片錨點位置
            anchor = image.anchor
            if hasattr(anchor, '_from'):
                from_pos = anchor._from
                row = from_pos.row if hasattr(from_pos, 'row') else 1
                col = from_pos.col if hasattr(from_pos, 'col') else 1
            else:
                row, col = 1, 1
            
            obj_info = ObjectInfo(
                obj_id=obj_id,
                display_name=f"Image_{len(objects) + 1}",
                obj_type=ObjectType.IMAGE_OBJ,
                block_id="",  # 稍後設定
                sheet_name=sheet_name,
                is_multi_rows=False,
                having_header=False,
                cell_position=CellPosition(row=row, col=col),
                data_shape=DataShape(rows=1, cols=1)
            )
            objects.append(obj_info)
            
        return objects
    
    def _parse_cell_reference(self, cell_ref: str) -> tuple:
        """
        解析儲存格參考（如A1）為行列座標
        
        Args:
            cell_ref: 儲存格參考字串
            
        Returns:
            tuple: (row, col) 座標
        """
        # 簡單實現，後續可改進
        from openpyxl.utils import coordinate_to_tuple
        return coordinate_to_tuple(cell_ref)
    
    def _scan_text_content(self, worksheet, sheet_name: str, tags: List[Tag]) -> List[ObjectInfo]:
        """
        掃描工作表中的非標籤文字內容
        
        Args:
            worksheet: 工作表物件
            sheet_name: 工作表名稱
            tags: 已解析的標籤列表（用於排除標籤位置）
            
        Returns:
            List[ObjectInfo]: 文字物件清單
        """
        objects = []
        
        # 建立標籤位置集合，用於排除
        tag_positions = set()
        for tag in tags:
            if tag.sheet_name == sheet_name:
                tag_positions.add((tag.cell_position.row, tag.cell_position.col))
        
        # 掃描工作表中的所有文字內容（限制在有實際內容的範圍內）
        max_row = min(worksheet.max_row or 1, 50)  # 限制掃描範圍避免效能問題
        max_col = min(worksheet.max_column or 1, 20)
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                if (cell.value is not None and 
                    isinstance(cell.value, str) and 
                    cell.value.strip() and
                    not cell.value.startswith('{{') and
                    not cell.value.startswith('#{{') and
                    (cell.row, cell.column) not in tag_positions):
                    
                    obj_id = f"text_obj_{uuid.uuid4().hex[:8]}"
                    
                    obj_info = ObjectInfo(
                        obj_id=obj_id,
                        display_name=f"Text_{cell.value[:20]}",  # 使用內容的前20字符作為顯示名稱
                        obj_type=ObjectType.TEXT,
                        block_id="",  # 稍後設定
                        sheet_name=sheet_name,
                        is_multi_rows=False,
                        having_header=False,
                        cell_position=CellPosition(row=cell.row, col=cell.column),
                        data_shape=DataShape(rows=1, cols=1)
                    )
                    objects.append(obj_info)
        
        return objects
    
    def classify_blocks(self, container: Container) -> Container:
        """
        將物件分類到Header/Gap/Footer區塊
        
        Args:
            container: 容器物件
            
        Returns:
            Container: 更新後的容器物件
        """
        # 找出所有表格類型的標籤位置（#前綴）
        table_positions = []
        for obj in container.objects:
            if obj.obj_type == ObjectType.TABLE:
                table_positions.append(obj.cell_position.row)
        
        table_positions.sort()
        
        # 根據表格位置分類區塊
        if not table_positions:
            # 沒有表格標籤，所有物件歸類為Header Block
            header_block = Block(
                block_id=f"header_{container.container_id}",
                block_type=BlockType.HEADER,
                rng_from=RangePosition(row=1, col=1, row_off=0, col_off=0),
                rng_to=RangePosition(row=1048576, col=16384, row_off=0, col_off=0)  # Excel最大範圍
            )
            container.blocks = [header_block]
            
            # 設定所有物件的block_id
            for obj in container.objects:
                obj.block_id = header_block.block_id
                
        else:
            # 有表格標籤，進行區塊分類
            blocks = []
            
            # Header Block: 第1行到第一個表格標籤前
            first_table_row = table_positions[0]
            if first_table_row > 1:
                header_block = Block(
                    block_id=f"header_{container.container_id}",
                    block_type=BlockType.HEADER,
                    rng_from=RangePosition(row=1, col=1, row_off=0, col_off=0),
                    rng_to=RangePosition(row=first_table_row - 1, col=16384, row_off=0, col_off=0)
                )
                blocks.append(header_block)
            
            # Gap Blocks: 表格標籤之間的區域
            for i, table_row in enumerate(table_positions):
                gap_block = Block(
                    block_id=f"gap_{i}_{container.container_id}",
                    block_type=BlockType.GAP,
                    rng_from=RangePosition(row=table_row, col=1, row_off=0, col_off=0),
                    rng_to=RangePosition(row=table_row, col=16384, row_off=0, col_off=0)  # 初始範圍
                )
                blocks.append(gap_block)
            
            # Footer Block: 最後一個表格標籤後到Sheet末尾  
            last_table_row = table_positions[-1]
            footer_block = Block(
                block_id=f"footer_{container.container_id}",
                block_type=BlockType.FOOTER,
                rng_from=RangePosition(row=last_table_row + 1, col=1, row_off=0, col_off=0),
                rng_to=RangePosition(row=1048576, col=16384, row_off=0, col_off=0)
            )
            blocks.append(footer_block)
            
            container.blocks = blocks
            
            # 分配物件到對應的區塊
            self._assign_objects_to_blocks(container)
        
        return container
    
    def _assign_objects_to_blocks(self, container: Container) -> None:
        """
        將物件分配到對應的區塊
        
        Args:
            container: 容器物件
        """
        for obj in container.objects:
            obj_row = obj.cell_position.row
            
            # 找到物件所屬的區塊
            for block in container.blocks:
                if block.rng_from.row <= obj_row <= block.rng_to.row:
                    obj.block_id = block.block_id
                    break
    
    def update_object_properties(self, container: Container, obj_id: str, properties: Dict) -> None:
        """
        更新物件屬性
        
        Args:
            container: 容器物件
            obj_id: 物件ID
            properties: 要更新的屬性字典
            
        Raises:
            ValueError: 找不到指定的物件
        """
        obj = container.get_object_by_id(obj_id)
        
        # 更新支援的屬性
        if 'data_shape' in properties:
            shape_data = properties['data_shape']
            obj.data_shape = DataShape(rows=shape_data['rows'], cols=shape_data['cols'])
        
        if 'cell_position' in properties:
            pos_data = properties['cell_position']
            obj.cell_position = CellPosition(row=pos_data['row'], col=pos_data['col'])
    
    def calculate_range_coordinates(self, container: Container) -> None:
        """
        計算所有物件的範圍座標
        
        Args:
            container: 容器物件
        """
        # 重新計算每個區塊的範圍
        for block in container.blocks:
            block_objects = container.get_objects_by_block_id(block.block_id)
            if block_objects:
                # 找出區塊中物件的最小和最大位置
                min_row = min(obj.cell_position.row for obj in block_objects)
                max_row = max(obj.cell_position.row + obj.data_shape.rows - 1 for obj in block_objects)
                min_col = min(obj.cell_position.col for obj in block_objects)
                max_col = max(obj.cell_position.col + obj.data_shape.cols - 1 for obj in block_objects)
                
                # 更新區塊範圍
                block.rng_from = RangePosition(row=min_row, col=min_col, row_off=0, col_off=0)
                block.rng_to = RangePosition(row=max_row, col=max_col, row_off=0, col_off=0)
    
    def update_container(self, container: Container, data: dict) -> None:
        """
        更新容器物件的資料形狀資訊
        
        當渲染過程中，table/table_obj 標籤完成渲染後，需要更新對應物件的 data_shape 資訊，
        以便後續的位移計算能正確進行。
        
        Args:
            container: 要更新的容器物件
            data: 包含標籤名稱和對應 DataFrame 的資料字典
        """
        import pandas as pd
        
        for obj_info in container.objects:
            # 只更新 table 和 table_obj 類型的物件
            if obj_info.obj_type in [ObjectType.TABLE, ObjectType.TABLE_OBJ]:
                # 尋找對應的資料
                tag_name = obj_info.display_name
                
                # 移除條件後綴（例如：tb4_routeid_df | noheader -> tb4_routeid_df）
                clean_tag_name = tag_name.split('|')[0].strip() if '|' in tag_name else tag_name
                
                if clean_tag_name in data:
                    dataframe = data[clean_tag_name]
                    
                    if isinstance(dataframe, pd.DataFrame):
                        # 更新資料形狀
                        obj_info.data_shape.rows = len(dataframe)
                        obj_info.data_shape.cols = len(dataframe.columns)
                        
                        # 如果有 header，總行數需要加1
                        if obj_info.having_header:
                            obj_info.data_shape.rows += 1
                            
                        print(f"DEBUG: 更新物件 {obj_info.display_name} 的資料形狀: "
                              f"rows={obj_info.data_shape.rows}, cols={obj_info.data_shape.cols}")
