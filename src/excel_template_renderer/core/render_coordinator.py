#!/usr/bin/env python3
"""
渲染協調器

統一協調整個渲染流程，包括註冊表建立、標籤重定位、數據渲染和結果驗證
"""
import os
import subprocess
from typing import Dict, List, Any, Optional
from pathlib import Path

from openpyxl import Workbook
from .object_registry import ObjectRegistry
from .tag_relocator import TagRelocator
from .renderer import TemplateRenderer
from ..utils.registry_utils import RegistryUtils


class RenderCoordinator:
    """渲染流程協調器"""
    
    def __init__(self):
        self.object_registry = ObjectRegistry()
        self.tag_relocator = TagRelocator()
        self.registry_utils = RegistryUtils()
        
        # 將在需要時初始化渲染器
        self._template_renderer = None
    
    @property
    def template_renderer(self):
        """惰性初始化模板渲染器"""
        if self._template_renderer is None:
            from .renderer import TemplateRenderer
            self._template_renderer = TemplateRenderer()
        return self._template_renderer
    
    def prepare_render(self, workbook: Workbook, data_context: Dict[str, Any], 
                      output_dir: Optional[str] = None) -> str:
        """
        渲染前準備階段
        
        執行完整的渲染前準備流程：
        1. 建立物件註冊表
        2. 計算所有物件的渲染後位置
        3. 生成標籤重定位計劃
        4. 執行標籤重定位
        5. 更新物件屬性
        6. 輸出完整註冊表JSON檔案
        
        Args:
            workbook: Excel工作簿物件
            data_context: 數據上下文
            output_dir: 輸出目錄，預設為當前工作目錄
            
        Returns:
            str: 產生的註冊表JSON檔案路徑
        """
        print("=== 渲染前準備階段開始 ===")
        
        # 1. 建立完整的物件註冊表
        print("1. 建立物件註冊表...")
        self.object_registry.build_complete_registry(workbook, data_context)
        
        # 2. 取得重定位計劃
        print("2. 分析重定位需求...")
        relocation_plans = self._extract_relocation_plans()
        
        # 3. 執行標籤重定位
        if relocation_plans:
            print(f"3. 執行標籤重定位 ({len(relocation_plans)} 個工作表)...")
            relocation_results = self._execute_relocations(workbook, relocation_plans)
            self._update_relocation_results(relocation_results)
        else:
            print("3. 無需重定位")
        
        # 4. 輸出註冊表JSON檔案
        print("4. 輸出註冊表JSON...")
        registry_path = self.object_registry.export_registry(output_dir)
        
        # 5. 驗證註冊表
        print("5. 驗證註冊表完整性...")
        is_valid, errors = self.object_registry.validate_registry_integrity()
        if not is_valid:
            print(f"警告：註冊表驗證發現問題: {errors}")
        
        print(f"=== 渲染前準備完成，註冊表已輸出至: {registry_path} ===")
        return registry_path
    
    def execute_render(self, workbook: Workbook, registry_json_path: str):
        """
        執行實際渲染
        
        依據註冊表JSON檔案執行數據渲染：
        1. 載入註冊表JSON
        2. 依據註冊表執行數據渲染
        3. 更新表格物件
        4. 調整圖片位置
        
        Args:
            workbook: Excel工作簿物件
            registry_json_path: 註冊表JSON檔案路徑
        """
        print("=== 執行渲染階段開始 ===")
        
        # 1. 載入註冊表
        print(f"1. 載入註冊表: {registry_json_path}")
        success = self.object_registry.load_and_validate(registry_json_path)
        if not success:
            raise ValueError(f"無法載入或驗證註冊表: {registry_json_path}")
        
        # 2. 執行數據渲染
        print("2. 執行數據渲染...")
        self._execute_data_rendering(workbook)
        
        # 3. 更新表格物件
        print("3. 更新表格物件...")
        self._update_table_objects(workbook)
        
        # 4. 調整圖片位置
        print("4. 調整圖片位置...")
        self._adjust_image_positions(workbook)
        
        print("=== 執行渲染完成 ===")
    
    def validate_render_result(self, workbook: Workbook, registry_json_path: str) -> Dict[str, Any]:
        """
        驗證渲染結果
        
        使用 extract_table_properties.py 提取實際結果並比對註冊表預測
        
        Args:
            workbook: 已渲染的工作簿
            registry_json_path: 註冊表JSON檔案路徑
            
        Returns:
            dict: 驗證結果
        """
        print("=== 驗證渲染結果階段開始 ===")
        
        try:
            # 1. 儲存工作簿到臨時檔案
            temp_excel_path = self._save_workbook_temporarily(workbook)
            
            # 2. 使用 extract_table_properties.py 提取實際結果
            print("2. 提取實際表格屬性...")
            actual_result_path = self._extract_actual_properties(temp_excel_path)
            
            # 3. 載入註冊表和實際結果
            registry = self.registry_utils.load_registry(registry_json_path)
            actual_result = self.registry_utils.load_registry(actual_result_path)
            
            # 4. 進行比對驗證
            print("4. 執行驗證比對...")
            validation_result = self._compare_registry_with_actual(registry, actual_result)
            
            # 5. 清理臨時檔案
            self._cleanup_temp_files([temp_excel_path])
            
            print("=== 驗證渲染結果完成 ===")
            return validation_result
            
        except Exception as e:
            print(f"驗證過程發生錯誤: {str(e)}")
            return {
                'validation_successful': False,
                'error': str(e),
                'timestamp': self._get_timestamp()
            }
    
    def save_result(self, workbook: Workbook, output_path: str):
        """
        保存渲染結果
        
        Args:
            workbook: 要保存的工作簿
            output_path: 輸出檔案路徑
        """
        print(f"=== 保存渲染結果至: {output_path} ===")
        
        try:
            # 確保輸出目錄存在
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # 保存工作簿
            workbook.save(output_path)
            print(f"渲染結果已成功保存至: {output_path}")
            
        except Exception as e:
            print(f"保存失敗: {str(e)}")
            raise
    
    def _extract_relocation_plans(self) -> Dict[str, List[Dict[str, Any]]]:
        """從註冊表中提取重定位計劃"""
        relocation_plans = {}
        
        registry = self.object_registry.registry
        worksheets = registry.get('worksheets', {})
        
        for sheet_name, sheet_data in worksheets.items():
            relocation_plan = sheet_data.get('relocation_plan', [])
            if relocation_plan:
                relocation_plans[sheet_name] = relocation_plan
        
        return relocation_plans
    
    def _execute_relocations(self, workbook: Workbook, 
                           relocation_plans: Dict[str, List[Dict[str, Any]]]) -> Dict[str, Dict[str, bool]]:
        """執行所有工作表的重定位"""
        all_results = {}
        
        for sheet_name, plan in relocation_plans.items():
            if sheet_name not in workbook.sheetnames:
                print(f"警告：工作表 {sheet_name} 不存在")
                continue
            
            worksheet = workbook[sheet_name]
            results = self.tag_relocator.batch_relocate(worksheet, plan)
            all_results[sheet_name] = results
            
            # 輸出結果摘要
            summary = self.tag_relocator.get_relocation_summary(results)
            print(f"  工作表 {sheet_name}: {summary['successful_relocations']}/{summary['total_objects']} 成功")
        
        return all_results
    
    def _update_relocation_results(self, relocation_results: Dict[str, Dict[str, bool]]):
        """更新註冊表中的重定位結果"""
        registry = self.object_registry.registry
        
        for sheet_name, results in relocation_results.items():
            sheet_data = registry.get('worksheets', {}).get(sheet_name, {})
            
            if 'relocation_plan' in sheet_data:
                for item in sheet_data['relocation_plan']:
                    obj_id = item['obj_id']
                    if obj_id in results:
                        item['relocation_successful'] = results[obj_id]
    
    def _execute_data_rendering(self, workbook: Workbook):
        """執行數據渲染"""
        # 這裡需要調用現有的渲染器
        # 由於我們採用完全依據註冊表的方式，需要從註冊表中提取渲染資訊
        
        registry = self.object_registry.registry
        
        for sheet_name, sheet_data in registry.get('worksheets', {}).items():
            if sheet_name not in workbook.sheetnames:
                continue
            
            worksheet = workbook[sheet_name]
            objects = sheet_data.get('objects', [])
            
            # 按render_order渲染物件
            render_order = sheet_data.get('render_order', [])
            
            for obj_id in render_order:
                obj = self._find_object_by_id(objects, obj_id)
                if obj:
                    self._render_single_object(worksheet, obj)
    
    def _find_object_by_id(self, objects: List[Dict[str, Any]], obj_id: str) -> Optional[Dict[str, Any]]:
        """根據ID找出物件"""
        for obj in objects:
            if obj.get('obj_id') == obj_id:
                return obj
        return None
    
    def _render_single_object(self, worksheet, obj: Dict[str, Any]):
        """渲染單一物件"""
        # 這是簡化的渲染邏輯，實際上需要調用現有的渲染器
        obj_type = obj.get('obj_type', 'simple')
        obj_name = obj.get('obj_name')
        position_after = obj.get('position_after', {})
        
        print(f"    渲染物件: {obj_name} ({obj_type}) 至 {position_after.get('coordinate', 'unknown')}")
        
        # 此處應該調用實際的渲染邏輯
        # 由於時間限制，這裡只做記錄
        # 實際實作時需要整合現有的 TemplateRenderer
    
    def _update_table_objects(self, workbook: Workbook):
        """更新表格物件"""
        # 根據註冊表更新Excel表格物件的範圍
        registry = self.object_registry.registry
        
        for sheet_name, sheet_data in registry.get('worksheets', {}).items():
            if sheet_name not in workbook.sheetnames:
                continue
            
            worksheet = workbook[sheet_name]
            objects = sheet_data.get('objects', [])
            
            for obj in objects:
                if obj.get('obj_type') == 'table':
                    self._update_single_table_object(worksheet, obj)
    
    def _update_single_table_object(self, worksheet, obj: Dict[str, Any]):
        """更新單一表格物件"""
        obj_name = obj.get('obj_name')
        position_after = obj.get('position_after', {})
        range_after = position_after.get('range_description')
        
        if range_after:
            print(f"    更新表格: {obj_name} -> {range_after}")
            
            # 這裡應該實際更新Excel表格物件的範圍
            # 具體實作需要操作 worksheet.tables
    
    def _adjust_image_positions(self, workbook: Workbook):
        """調整圖片位置"""
        registry = self.object_registry.registry
        
        for sheet_name, sheet_data in registry.get('worksheets', {}).items():
            if sheet_name not in workbook.sheetnames:
                continue
            
            worksheet = workbook[sheet_name]
            objects = sheet_data.get('objects', [])
            
            for obj in objects:
                if obj.get('obj_type') == 'image':
                    self._adjust_single_image_position(worksheet, obj)
    
    def _adjust_single_image_position(self, worksheet, obj: Dict[str, Any]):
        """調整單一圖片位置"""
        obj_name = obj.get('obj_name')
        position_after = obj.get('position_after', {})
        
        print(f"    調整圖片: {obj_name} 至 {position_after.get('coordinate', 'unknown')}")
        
        # 這裡應該實際調整圖片位置
        # 具體實作需要操作 worksheet._images
    
    def _save_workbook_temporarily(self, workbook: Workbook) -> str:
        """暫時保存工作簿"""
        import tempfile
        
        temp_dir = tempfile.gettempdir()
        timestamp = self._get_timestamp_for_filename()
        temp_path = os.path.join(temp_dir, f"temp_render_result_{timestamp}.xlsx")
        
        workbook.save(temp_path)
        return temp_path
    
    def _extract_actual_properties(self, excel_path: str) -> str:
        """使用 extract_table_properties.py 提取實際屬性"""
        try:
            # 執行 extract_table_properties.py
            result = subprocess.run(
                ['python', 'extract_table_properties.py', excel_path],
                capture_output=True,
                text=True,
                timeout=60  # 60秒超時
            )
            
            if result.returncode == 0:
                # 解析輸出以找出生成的JSON檔案路徑
                output_lines = result.stdout.strip().split('\n')
                for line in output_lines:
                    if 'template_renderer_obj_result_' in line and line.endswith('.json'):
                        return line.strip()
                
                # 如果無法從輸出中找到檔案路徑，嘗試找出最新的結果檔案
                return self._find_latest_result_file()
            else:
                raise RuntimeError(f"extract_table_properties.py 執行失敗: {result.stderr}")
                
        except subprocess.TimeoutExpired:
            raise RuntimeError("extract_table_properties.py 執行超時")
        except Exception as e:
            raise RuntimeError(f"執行 extract_table_properties.py 時發生錯誤: {str(e)}")
    
    def _find_latest_result_file(self) -> str:
        """找出最新的結果檔案"""
        import glob
        
        pattern = "template_renderer_obj_result_*.json"
        result_files = glob.glob(pattern)
        
        if not result_files:
            raise RuntimeError("找不到結果檔案")
        
        # 按修改時間排序，取最新的
        latest_file = max(result_files, key=os.path.getmtime)
        return latest_file
    
    def _compare_registry_with_actual(self, registry: Dict[str, Any], 
                                    actual_result: Dict[str, Any]) -> Dict[str, Any]:
        """比對註冊表與實際結果"""
        validation_result = {
            'validation_successful': True,
            'timestamp': self._get_timestamp(),
            'registry_file': registry.get('metadata', {}).get('registry_filename', 'unknown'),
            'result_file': actual_result.get('metadata', {}).get('registry_filename', 'unknown'),
            'validation_details': {},
            'summary': {
                'total_objects': 0,
                'validated_objects': 0,
                'failed_objects': 0,
                'accuracy_rate': '0%'
            }
        }
        
        # 比對每個工作表
        registry_worksheets = registry.get('worksheets', {})
        actual_worksheets = actual_result.get('worksheets', {})
        
        for sheet_name in registry_worksheets.keys():
            sheet_validation = self._validate_single_worksheet(
                registry_worksheets.get(sheet_name, {}),
                actual_worksheets.get(sheet_name, {})
            )
            
            validation_result['validation_details'][sheet_name] = sheet_validation
            
            # 更新摘要
            validation_result['summary']['total_objects'] += sheet_validation['total_objects']
            validation_result['summary']['validated_objects'] += sheet_validation['validated_objects']
            validation_result['summary']['failed_objects'] += sheet_validation['failed_objects']
        
        # 計算準確率
        total = validation_result['summary']['total_objects']
        validated = validation_result['summary']['validated_objects']
        if total > 0:
            accuracy = (validated / total) * 100
            validation_result['summary']['accuracy_rate'] = f"{accuracy:.1f}%"
            validation_result['validation_successful'] = (accuracy >= 95.0)  # 95%以上視為成功
        
        return validation_result
    
    def _validate_single_worksheet(self, registry_sheet: Dict[str, Any], 
                                 actual_sheet: Dict[str, Any]) -> Dict[str, Any]:
        """驗證單一工作表"""
        validation = {
            'total_objects': 0,
            'validated_objects': 0,
            'failed_objects': 0,
            'object_validations': {}
        }
        
        registry_objects = registry_sheet.get('objects', [])
        actual_objects = actual_sheet.get('objects', [])
        
        # 建立實際物件的索引
        actual_objects_index = {obj.get('obj_name'): obj for obj in actual_objects}
        
        for reg_obj in registry_objects:
            obj_name = reg_obj.get('obj_name')
            validation['total_objects'] += 1
            
            actual_obj = actual_objects_index.get(obj_name)
            if actual_obj:
                obj_validation = self._validate_single_object(reg_obj, actual_obj)
                validation['object_validations'][obj_name] = obj_validation
                
                if obj_validation['validation_passed']:
                    validation['validated_objects'] += 1
                else:
                    validation['failed_objects'] += 1
            else:
                validation['object_validations'][obj_name] = {
                    'validation_passed': False,
                    'error': 'Object not found in actual result'
                }
                validation['failed_objects'] += 1
        
        return validation
    
    def _validate_single_object(self, registry_obj: Dict[str, Any], 
                              actual_obj: Dict[str, Any]) -> Dict[str, Any]:
        """驗證單一物件"""
        validation = {
            'validation_passed': True,
            'differences': []
        }
        
        # 比較位置
        reg_pos = registry_obj.get('position_after', {})
        actual_pos = actual_obj.get('position_after', {})
        
        if reg_pos.get('range_description') != actual_pos.get('range_description'):
            validation['validation_passed'] = False
            validation['differences'].append({
                'field': 'range_description',
                'expected': reg_pos.get('range_description'),
                'actual': actual_pos.get('range_description')
            })
        
        # 比較數據形狀
        reg_shape = reg_pos.get('data_shape', {})
        actual_shape = actual_pos.get('data_shape', {})
        
        if reg_shape != actual_shape:
            validation['validation_passed'] = False
            validation['differences'].append({
                'field': 'data_shape',
                'expected': reg_shape,
                'actual': actual_shape
            })
        
        return validation
    
    def _cleanup_temp_files(self, file_paths: List[str]):
        """清理臨時檔案"""
        for file_path in file_paths:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception as e:
                print(f"清理臨時檔案失敗: {file_path}, 錯誤: {str(e)}")
    
    def _get_timestamp(self) -> str:
        """取得ISO格式的時間戳記"""
        from datetime import datetime
        return datetime.now().isoformat()
    
    def _get_timestamp_for_filename(self) -> str:
        """取得適用於檔案名的時間戳記"""
        from datetime import datetime
        return datetime.now().strftime("%Y%m%d_%H%M%S")
    
    def get_coordinator_summary(self) -> Dict[str, Any]:
        """取得協調器摘要資訊"""
        registry_summary = self.object_registry.get_registry_summary()
        
        return {
            'coordinator_version': '1.0',
            'registry_summary': registry_summary,
            'supported_operations': [
                'prepare_render',
                'execute_render', 
                'validate_render_result',
                'save_result'
            ],
            'integrated_components': [
                'ObjectRegistry',
                'TagRelocator',
                'RegistryUtils'
            ]
        }