#!/usr/bin/env python3
"""
標籤重定位器

實現標籤在Excel工作表中的移動功能，包括儲存格內容、格式、公式等的完整複製
"""
from typing import Dict, List, Any, Tuple, Optional
import copy

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.merge import MergedCellRange

from ..utils.position_utils import PositionUtils


class TagRelocator:
    """標籤重定位器"""
    
    def __init__(self):
        self.position_utils = PositionUtils()
        self._backup_storage = {}  # 備份儲存
    
    def relocate_tag(self, worksheet: Worksheet, from_pos: str, to_pos: str) -> bool:
        """
        移動單一標籤
        
        完整處理儲存格的移動：
        1. 複製儲存格內容（值、格式、公式）
        2. 處理合併儲存格
        3. 清除原始位置
        4. 寫入新位置
        
        Args:
            worksheet: Excel工作表
            from_pos: 原始位置 (如 'B5')
            to_pos: 目標位置 (如 'C7')
            
        Returns:
            bool: 移動是否成功
        """
        try:
            # 解析位置
            from_row, from_col = self.position_utils.excel_to_position(from_pos)
            to_row, to_col = self.position_utils.excel_to_position(to_pos)
            
            # 取得原始儲存格
            from_cell = worksheet.cell(row=from_row, column=from_col)
            
            # 檢查原始儲存格是否為空
            if self._is_cell_empty(from_cell):
                return True  # 空儲存格，視為成功
            
            # 檢查目標位置是否可用
            if not self._is_position_available(worksheet, to_pos):
                print(f"警告：目標位置 {to_pos} 不可用")
                return False
            
            # 備份原始儲存格
            backup_data = self._backup_cell(worksheet, from_pos)
            
            # 複製到新位置
            success = self._copy_cell_to_position(worksheet, from_row, from_col, to_row, to_col)
            
            if success:
                # 清除原始位置
                self._clear_cell(worksheet, from_row, from_col)
                return True
            else:
                # 復原失敗，回復備份
                self._restore_cell(worksheet, from_pos, backup_data)
                return False
                
        except Exception as e:
            print(f"標籤重定位失敗: {from_pos} -> {to_pos}, 錯誤: {str(e)}")
            return False
    
    def batch_relocate(self, worksheet: Worksheet, relocation_plan: List[Dict[str, Any]]) -> Dict[str, bool]:
        """
        批次移動多個標籤
        
        確保移動順序避免衝突
        
        Args:
            worksheet: Excel工作表
            relocation_plan: 重定位計劃清單
            
        Returns:
            dict: 每個物件的移動結果 {obj_id: success}
        """
        results = {}
        
        # 驗證重定位計劃
        is_valid, conflicts = self.validate_relocation(worksheet, relocation_plan)
        if not is_valid:
            print(f"重定位計劃驗證失敗: {conflicts}")
            return {item['obj_id']: False for item in relocation_plan}
        
        # 按優先級排序（如果有的話）
        sorted_plan = sorted(relocation_plan, key=lambda x: x.get('priority', 100))
        
        # 建立批次備份
        batch_backup = self._create_batch_backup(worksheet, sorted_plan)
        
        try:
            # 執行移動
            for item in sorted_plan:
                obj_id = item['obj_id']
                from_pos = item['from_coordinate']
                to_pos = item['to_coordinate']
                
                success = self.relocate_tag(worksheet, from_pos, to_pos)
                results[obj_id] = success
                
                if not success:
                    print(f"物件 {obj_id} 移動失敗: {from_pos} -> {to_pos}")
                    # 可以選擇繼續或中斷
                    # break  # 如果要中斷整個批次操作
                
        except Exception as e:
            print(f"批次重定位過程中發生錯誤: {str(e)}")
            # 如果需要，可以進行整體回滾
            # self._restore_batch_backup(worksheet, batch_backup)
        
        return results
    
    def validate_relocation(self, worksheet: Worksheet, 
                          relocation_plan: List[Dict[str, Any]]) -> Tuple[bool, List[str]]:
        """
        驗證重定位計劃
        
        檢查目標位置是否可用，是否有衝突
        
        Args:
            worksheet: Excel工作表
            relocation_plan: 重定位計劃
            
        Returns:
            tuple: (is_valid, conflicts)
        """
        conflicts = []
        target_positions = set()
        
        for item in relocation_plan:
            from_pos = item.get('from_coordinate')
            to_pos = item.get('to_coordinate')
            obj_id = item.get('obj_id', 'unknown')
            
            if not from_pos or not to_pos:
                conflicts.append(f"物件 {obj_id} 缺少座標資訊")
                continue
            
            # 檢查座標格式
            try:
                self.position_utils.excel_to_position(from_pos)
                self.position_utils.excel_to_position(to_pos)
            except ValueError as e:
                conflicts.append(f"物件 {obj_id} 座標格式錯誤: {str(e)}")
                continue
            
            # 檢查目標位置衝突
            if to_pos in target_positions:
                conflicts.append(f"多個物件試圖移動到相同位置: {to_pos}")
            else:
                target_positions.add(to_pos)
            
            # 檢查目標位置是否被其他內容佔用（除了參與移動的位置）
            if not self._is_position_available_for_relocation(worksheet, to_pos, relocation_plan):
                conflicts.append(f"目標位置 {to_pos} 被其他內容佔用")
        
        return len(conflicts) == 0, conflicts
    
    def backup_cells(self, worksheet: Worksheet, positions: List[str]) -> Dict[str, Dict[str, Any]]:
        """
        備份儲存格內容以便復原
        
        Args:
            worksheet: Excel工作表
            positions: 位置清單
            
        Returns:
            dict: 備份數據
        """
        backup = {}
        
        for pos in positions:
            backup_data = self._backup_cell(worksheet, pos)
            if backup_data:
                backup[pos] = backup_data
        
        return backup
    
    def restore_cells(self, worksheet: Worksheet, backup: Dict[str, Dict[str, Any]]):
        """
        從備份復原儲存格
        
        Args:
            worksheet: Excel工作表
            backup: 備份數據
        """
        for pos, backup_data in backup.items():
            self._restore_cell(worksheet, pos, backup_data)
    
    def _is_cell_empty(self, cell: Cell) -> bool:
        """檢查儲存格是否為空"""
        return (cell.value is None and 
                not cell.has_style and
                cell.comment is None and
                cell.hyperlink is None)
    
    def _is_position_available(self, worksheet: Worksheet, position: str) -> bool:
        """檢查位置是否可用（完全空白）"""
        try:
            row, col = self.position_utils.excel_to_position(position)
            cell = worksheet.cell(row=row, column=col)
            return self._is_cell_empty(cell)
        except Exception:
            return False
    
    def _is_position_available_for_relocation(self, worksheet: Worksheet, position: str,
                                           relocation_plan: List[Dict[str, Any]]) -> bool:
        """檢查位置是否可用於重定位（考慮重定位計劃中的位置）"""
        # 如果位置在重定位計劃的源位置中，視為可用
        source_positions = {item['from_coordinate'] for item in relocation_plan}
        if position in source_positions:
            return True
        
        # 否則檢查是否為空
        return self._is_position_available(worksheet, position)
    
    def _backup_cell(self, worksheet: Worksheet, position: str) -> Optional[Dict[str, Any]]:
        """備份單一儲存格"""
        try:
            row, col = self.position_utils.excel_to_position(position)
            cell = worksheet.cell(row=row, column=col)
            
            if self._is_cell_empty(cell):
                return None
            
            backup_data = {
                'position': position,
                'row': row,
                'col': col,
                'value': cell.value,
                'data_type': cell.data_type,
                'style': self._backup_cell_style(cell),
                'comment': cell.comment.text if cell.comment else None,
                'hyperlink': str(cell.hyperlink.target) if cell.hyperlink else None,
                'formula': cell.formula if hasattr(cell, 'formula') else None,
                'is_merged': self._is_cell_merged(worksheet, row, col)
            }
            
            return backup_data
            
        except Exception as e:
            print(f"備份儲存格失敗: {position}, 錯誤: {str(e)}")
            return None
    
    def _backup_cell_style(self, cell: Cell) -> Dict[str, Any]:
        """備份儲存格樣式"""
        try:
            return {
                'font': copy.deepcopy(cell.font.__dict__ if cell.font else {}),
                'fill': copy.deepcopy(cell.fill.__dict__ if cell.fill else {}),
                'border': copy.deepcopy(cell.border.__dict__ if cell.border else {}),
                'alignment': copy.deepcopy(cell.alignment.__dict__ if cell.alignment else {}),
                'number_format': cell.number_format,
                'protection': copy.deepcopy(cell.protection.__dict__ if cell.protection else {})
            }
        except Exception as e:
            print(f"備份樣式失敗: {str(e)}")
            return {}
    
    def _is_cell_merged(self, worksheet: Worksheet, row: int, col: int) -> Optional[str]:
        """檢查儲存格是否為合併儲存格的一部分"""
        try:
            for merge_range in worksheet.merged_cells.ranges:
                if (merge_range.min_row <= row <= merge_range.max_row and
                    merge_range.min_col <= col <= merge_range.max_col):
                    return str(merge_range)
            return None
        except Exception:
            return None
    
    def _copy_cell_to_position(self, worksheet: Worksheet, from_row: int, from_col: int,
                             to_row: int, to_col: int) -> bool:
        """複製儲存格到新位置"""
        try:
            from_cell = worksheet.cell(row=from_row, column=from_col)
            to_cell = worksheet.cell(row=to_row, column=to_col)
            
            # 複製值和數據類型
            to_cell.value = from_cell.value
            to_cell.data_type = from_cell.data_type
            
            # 複製樣式
            self._copy_cell_style(from_cell, to_cell)
            
            # 複製註解
            if from_cell.comment:
                to_cell.comment = copy.deepcopy(from_cell.comment)
            
            # 複製超連結
            if from_cell.hyperlink:
                to_cell.hyperlink = from_cell.hyperlink
            
            # 處理公式（如果需要調整引用）
            if hasattr(from_cell, 'formula') and from_cell.formula:
                adjusted_formula = self._adjust_formula_references(
                    from_cell.formula, from_row, from_col, to_row, to_col
                )
                to_cell.formula = adjusted_formula
            
            return True
            
        except Exception as e:
            print(f"複製儲存格失敗: ({from_row},{from_col}) -> ({to_row},{to_col}), 錯誤: {str(e)}")
            return False
    
    def _copy_cell_style(self, from_cell: Cell, to_cell: Cell):
        """複製儲存格樣式"""
        try:
            if from_cell.has_style:
                to_cell.font = copy.deepcopy(from_cell.font)
                to_cell.fill = copy.deepcopy(from_cell.fill)
                to_cell.border = copy.deepcopy(from_cell.border)
                to_cell.alignment = copy.deepcopy(from_cell.alignment)
                to_cell.number_format = from_cell.number_format
                to_cell.protection = copy.deepcopy(from_cell.protection)
        except Exception as e:
            print(f"複製樣式失敗: {str(e)}")
    
    def _adjust_formula_references(self, formula: str, from_row: int, from_col: int,
                                 to_row: int, to_col: int) -> str:
        """調整公式中的引用"""
        # 簡單實現：對於複雜的公式調整，可能需要更sophisticated的解析
        # 目前只處理基本情況
        try:
            # 計算位移
            row_offset = to_row - from_row
            col_offset = to_col - from_col
            
            # 這裡需要實現公式引用的調整邏輯
            # 暫時返回原公式
            # TODO: 實現完整的公式引用調整
            return formula
            
        except Exception as e:
            print(f"調整公式失敗: {formula}, 錯誤: {str(e)}")
            return formula
    
    def _clear_cell(self, worksheet: Worksheet, row: int, col: int):
        """清除儲存格內容"""
        try:
            cell = worksheet.cell(row=row, column=col)
            
            # 清除值和數據類型
            cell.value = None
            cell.data_type = 's'  # 重設為字符串類型
            
            # 清除註解
            if cell.comment:
                cell.comment = None
            
            # 清除超連結
            if cell.hyperlink:
                cell.hyperlink = None
            
            # 重設樣式為預設
            from openpyxl.styles import Font, Fill, Border, Alignment, Protection
            cell.font = Font()
            cell.fill = Fill()
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = 'General'
            cell.protection = Protection()
            
        except Exception as e:
            print(f"清除儲存格失敗: ({row},{col}), 錯誤: {str(e)}")
    
    def _restore_cell(self, worksheet: Worksheet, position: str, backup_data: Dict[str, Any]):
        """復原單一儲存格"""
        if not backup_data:
            return
        
        try:
            row, col = backup_data['row'], backup_data['col']
            cell = worksheet.cell(row=row, column=col)
            
            # 復原值和數據類型
            cell.value = backup_data['value']
            cell.data_type = backup_data['data_type']
            
            # 復原樣式
            style_data = backup_data.get('style', {})
            self._restore_cell_style(cell, style_data)
            
            # 復原註解
            if backup_data.get('comment'):
                from openpyxl.comments import Comment
                cell.comment = Comment(backup_data['comment'], 'Restored')
            
            # 復原超連結
            if backup_data.get('hyperlink'):
                from openpyxl.worksheet.hyperlink import Hyperlink
                cell.hyperlink = Hyperlink(backup_data['hyperlink'])
            
            # 復原公式
            if backup_data.get('formula'):
                cell.formula = backup_data['formula']
            
        except Exception as e:
            print(f"復原儲存格失敗: {position}, 錯誤: {str(e)}")
    
    def _restore_cell_style(self, cell: Cell, style_data: Dict[str, Any]):
        """復原儲存格樣式"""
        try:
            from openpyxl.styles import Font, Fill, Border, Alignment, Protection
            
            if 'font' in style_data and style_data['font']:
                cell.font = Font(**style_data['font'])
            
            if 'fill' in style_data and style_data['fill']:
                cell.fill = Fill(**style_data['fill'])
            
            if 'border' in style_data and style_data['border']:
                cell.border = Border(**style_data['border'])
            
            if 'alignment' in style_data and style_data['alignment']:
                cell.alignment = Alignment(**style_data['alignment'])
            
            if 'number_format' in style_data:
                cell.number_format = style_data['number_format']
            
            if 'protection' in style_data and style_data['protection']:
                cell.protection = Protection(**style_data['protection'])
                
        except Exception as e:
            print(f"復原樣式失敗: {str(e)}")
    
    def _create_batch_backup(self, worksheet: Worksheet, 
                           relocation_plan: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
        """建立批次備份"""
        batch_backup = {}
        
        # 備份所有源位置
        for item in relocation_plan:
            from_pos = item.get('from_coordinate')
            if from_pos:
                backup_data = self._backup_cell(worksheet, from_pos)
                if backup_data:
                    batch_backup[from_pos] = backup_data
        
        return batch_backup
    
    def _restore_batch_backup(self, worksheet: Worksheet, batch_backup: Dict[str, Dict[str, Any]]):
        """復原批次備份"""
        for position, backup_data in batch_backup.items():
            self._restore_cell(worksheet, position, backup_data)
    
    def get_relocation_summary(self, results: Dict[str, bool]) -> Dict[str, Any]:
        """取得重定位摘要"""
        total = len(results)
        successful = sum(1 for success in results.values() if success)
        failed = total - successful
        
        return {
            'total_objects': total,
            'successful_relocations': successful,
            'failed_relocations': failed,
            'success_rate': f"{(successful/total*100):.1f}%" if total > 0 else "0%",
            'failed_objects': [obj_id for obj_id, success in results.items() if not success]
        }