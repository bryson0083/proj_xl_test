"""
公式處理器
"""
import re
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils import column_index_from_string, get_column_letter

from ..models.base import CellPosition


@dataclass
class FormulaShiftInfo:
    """公式位移資訊"""
    sheet_name: str
    start_row: int
    shift_amount: int
    affected_columns: List[int]


@dataclass
class FormulaReference:
    """公式參照"""
    original_text: str
    sheet_name: Optional[str]
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    is_absolute_row: bool
    is_absolute_col: bool


class FormulaProcessor:
    """
    公式處理器類別
    
    負責處理Excel公式中的儲存格參照更新
    """
    
    # 儲存格參照的正則表達式模式
    CELL_REF_PATTERN = re.compile(
        r'(?:([a-zA-Z_][a-zA-Z0-9_\.]*!))?'  # 工作表名稱（可選）
        r'(\$?)([A-Z]+)'                     # 欄位（可能有$前綴）
        r'(\$?)(\d+)'                        # 行號（可能有$前綴）
        r'(?::(\$?)([A-Z]+)(\$?)(\d+))?',    # 範圍結束（可選）
        re.IGNORECASE
    )
    
    def update_formulas_after_shift(
        self, 
        worksheet: Worksheet, 
        shift_info: FormulaShiftInfo
    ) -> int:
        """
        在表格位移後更新公式
        
        Args:
            worksheet: 工作表物件
            shift_info: 位移資訊
            
        Returns:
            int: 更新的公式數量
        """
        updated_count = 0
        
        # 掃描工作表中所有有公式的儲存格
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value and isinstance(cell.value, str):
                    original_formula = cell.value
                    new_formula = self._update_formula_references(original_formula, shift_info)
                    
                    if new_formula != original_formula:
                        try:
                            cell.value = new_formula
                            updated_count += 1
                        except Exception:
                            # 忽略無法更新的儲存格（如合併儲存格）
                            pass
        
        return updated_count
    
    def _update_formula_references(
        self, 
        formula: str, 
        shift_info: FormulaShiftInfo
    ) -> str:
        """
        更新公式中的儲存格參照
        
        Args:
            formula: 原始公式
            shift_info: 位移資訊
            
        Returns:
            str: 更新後的公式
        """
        def replace_reference(match):
            ref = self._parse_cell_reference(match)
            updated_ref = self._shift_reference(ref, shift_info)
            return self._format_reference(updated_ref)
        
        return self.CELL_REF_PATTERN.sub(replace_reference, formula)
    
    def _parse_cell_reference(self, match) -> FormulaReference:
        """
        解析儲存格參照
        
        Args:
            match: 正則表達式匹配結果
            
        Returns:
            FormulaReference: 解析後的參照資訊
        """
        groups = match.groups()
        
        # 解析工作表名稱
        sheet_name = groups[0].rstrip('!') if groups[0] else None
        
        # 解析起始位置
        is_abs_col_start = bool(groups[1])
        start_col_letter = groups[2]
        is_abs_row_start = bool(groups[3])
        start_row = int(groups[4])
        
        start_col = column_index_from_string(start_col_letter)
        
        # 解析結束位置（如果是範圍）
        if groups[5] is not None:  # 有範圍結束
            is_abs_col_end = bool(groups[5])
            end_col_letter = groups[6]
            is_abs_row_end = bool(groups[7])
            end_row = int(groups[8])
            
            end_col = column_index_from_string(end_col_letter)
            
            # 對於範圍，使用統一的絕對/相對設定
            is_absolute_col = is_abs_col_start and is_abs_col_end
            is_absolute_row = is_abs_row_start and is_abs_row_end
        else:
            end_row = start_row
            end_col = start_col
            is_absolute_col = is_abs_col_start
            is_absolute_row = is_abs_row_start
        
        return FormulaReference(
            original_text=match.group(0),
            sheet_name=sheet_name,
            start_row=start_row,
            end_row=end_row,
            start_col=start_col,
            end_col=end_col,
            is_absolute_row=is_absolute_row,
            is_absolute_col=is_absolute_col
        )
    
    def _shift_reference(
        self, 
        ref: FormulaReference, 
        shift_info: FormulaShiftInfo
    ) -> FormulaReference:
        """
        位移儲存格參照
        
        Args:
            ref: 原始參照
            shift_info: 位移資訊
            
        Returns:
            FormulaReference: 位移後的參照
        """
        # 如果是絕對行參照，不需要位移
        if ref.is_absolute_row:
            return ref
        
        # 如果參照不在位移範圍內，不需要位移
        if not self._should_shift_reference(ref, shift_info):
            return ref
        
        # 計算新的行號
        new_start_row = ref.start_row + shift_info.shift_amount
        new_end_row = ref.end_row + shift_info.shift_amount
        
        # 確保行號不會變成負數
        if new_start_row < 1:
            new_start_row = 1
        if new_end_row < 1:
            new_end_row = 1
        
        return FormulaReference(
            original_text=ref.original_text,
            sheet_name=ref.sheet_name,
            start_row=new_start_row,
            end_row=new_end_row,
            start_col=ref.start_col,
            end_col=ref.end_col,
            is_absolute_row=ref.is_absolute_row,
            is_absolute_col=ref.is_absolute_col
        )
    
    def _should_shift_reference(
        self, 
        ref: FormulaReference, 
        shift_info: FormulaShiftInfo
    ) -> bool:
        """
        判斷參照是否需要位移
        
        Args:
            ref: 儲存格參照
            shift_info: 位移資訊
            
        Returns:
            bool: 是否需要位移
        """
        # 檢查是否為相同工作表（如果有指定工作表名稱）
        if ref.sheet_name and ref.sheet_name != shift_info.sheet_name:
            return False
        
        # 檢查是否在位移的行範圍之後
        return ref.start_row >= shift_info.start_row
    
    def _format_reference(self, ref: FormulaReference) -> str:
        """
        格式化儲存格參照
        
        Args:
            ref: 儲存格參照
            
        Returns:
            str: 格式化後的參照字串
        """
        result = ""
        
        # 添加工作表名稱
        if ref.sheet_name:
            result += f"{ref.sheet_name}!"
        
        # 添加絕對/相對符號和欄位
        if ref.is_absolute_col:
            result += "$"
        result += get_column_letter(ref.start_col)
        
        # 添加絕對/相對符號和行號
        if ref.is_absolute_row:
            result += "$"
        result += str(ref.start_row)
        
        # 如果是範圍，添加結束位置
        if ref.end_row != ref.start_row or ref.end_col != ref.start_col:
            result += ":"
            if ref.is_absolute_col:
                result += "$"
            result += get_column_letter(ref.end_col)
            if ref.is_absolute_row:
                result += "$"
            result += str(ref.end_row)
        
        return result
    
    def analyze_formula_dependencies(
        self, 
        worksheet: Worksheet
    ) -> Dict[str, List[str]]:
        """
        分析公式相依性
        
        Args:
            worksheet: 工作表物件
            
        Returns:
            Dict[str, List[str]]: 儲存格到其相依儲存格的映射
        """
        dependencies = {}
        
        for row in worksheet.iter_rows():
            for cell in row:
                if (cell.data_type == 'f' and cell.value and 
                    isinstance(cell.value, str) and cell.column is not None):
                    cell_address = f"{get_column_letter(cell.column)}{cell.row}"
                    refs = self._extract_cell_references(cell.value)
                    dependencies[cell_address] = refs
        
        return dependencies
    
    def _extract_cell_references(self, formula: str) -> List[str]:
        """
        提取公式中的儲存格參照
        
        Args:
            formula: 公式字串
            
        Returns:
            List[str]: 儲存格參照清單
        """
        references = []
        
        for match in self.CELL_REF_PATTERN.finditer(formula):
            ref = self._parse_cell_reference(match)
            ref_str = self._format_reference_simple(ref)
            references.append(ref_str)
        
        return references
    
    def _format_reference_simple(self, ref: FormulaReference) -> str:
        """
        簡單格式化儲存格參照（不包含絕對符號）
        
        Args:
            ref: 儲存格參照
            
        Returns:
            str: 簡化的參照字串
        """
        result = ""
        
        if ref.sheet_name:
            result += f"{ref.sheet_name}!"
        
        result += f"{get_column_letter(ref.start_col)}{ref.start_row}"
        
        if ref.end_row != ref.start_row or ref.end_col != ref.start_col:
            result += f":{get_column_letter(ref.end_col)}{ref.end_row}"
        
        return result
    
    def validate_formula_syntax(self, formula: str) -> Tuple[bool, Optional[str]]:
        """
        驗證公式語法
        
        Args:
            formula: 公式字串
            
        Returns:
            Tuple[bool, Optional[str]]: (是否有效, 錯誤訊息)
        """
        try:
            # 基本檢查：公式應該以等號開始
            if not formula.startswith('='):
                return False, "公式必須以等號 (=) 開始"
            
            # 檢查括號配對
            if not self._check_parentheses_balance(formula):
                return False, "括號不配對"
            
            # 檢查儲存格參照格式
            invalid_refs = self._check_invalid_references(formula)
            if invalid_refs:
                return False, f"無效的儲存格參照: {', '.join(invalid_refs)}"
            
            return True, None
            
        except Exception as e:
            return False, f"公式驗證錯誤: {str(e)}"
    
    def _check_parentheses_balance(self, formula: str) -> bool:
        """
        檢查括號是否配對
        
        Args:
            formula: 公式字串
            
        Returns:
            bool: 括號是否配對
        """
        count = 0
        for char in formula:
            if char == '(':
                count += 1
            elif char == ')':
                count -= 1
                if count < 0:
                    return False
        return count == 0
    
    def _check_invalid_references(self, formula: str) -> List[str]:
        """
        檢查無效的儲存格參照
        
        Args:
            formula: 公式字串
            
        Returns:
            List[str]: 無效參照清單
        """
        invalid_refs = []
        
        # 使用更嚴格的模式檢查參照
        strict_pattern = re.compile(
            r'(?:[a-zA-Z_][a-zA-Z0-9_\.]*!)?'  # 工作表名稱
            r'\$?[A-Z]{1,3}\$?\d{1,7}'        # 儲存格參照
            r'(?::\$?[A-Z]{1,3}\$?\d{1,7})?', # 範圍（可選）
            re.IGNORECASE
        )
        
        matches = self.CELL_REF_PATTERN.finditer(formula)
        for match in matches:
            ref_text = match.group(0)
            if not strict_pattern.fullmatch(ref_text):
                invalid_refs.append(ref_text)
        
        return invalid_refs
    
    def backup_formulas(self, worksheet: Worksheet) -> Dict[str, str]:
        """
        備份工作表中的所有公式
        
        Args:
            worksheet: 工作表物件
            
        Returns:
            Dict[str, str]: 儲存格位置到公式的映射
        """
        formulas = {}
        
        for row in worksheet.iter_rows():
            for cell in row:
                if (cell.data_type == 'f' and cell.value and 
                    isinstance(cell.value, str) and cell.column is not None):
                    cell_address = f"{get_column_letter(cell.column)}{cell.row}"
                    formulas[cell_address] = cell.value
        
        return formulas
    
    def restore_formulas(
        self, 
        worksheet: Worksheet, 
        formulas: Dict[str, str]
    ) -> None:
        """
        還原工作表中的公式
        
        Args:
            worksheet: 工作表物件
            formulas: 儲存格位置到公式的映射
        """
        for cell_address, formula in formulas.items():
            try:
                cell = worksheet[cell_address]
                cell.value = formula
            except Exception:
                # 忽略還原失敗的情況
                pass
