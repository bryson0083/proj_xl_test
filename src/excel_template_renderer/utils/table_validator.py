#!/usr/bin/env python3
"""
表格物件驗證器

整合 extract_table_properties.py 的邏輯，用於驗證渲染結果與註冊表預測的一致性
"""
import json
import subprocess
import os
from typing import Dict, List, Any, Tuple, Optional
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from .registry_utils import RegistryUtils


class TableValidator:
    """表格物件驗證器"""
    
    def __init__(self):
        self.registry_utils = RegistryUtils()
    
    def extract_actual_tables(self, excel_path: str) -> Dict[str, Any]:
        """
        使用 extract_table_properties.py 邏輯提取實際表格屬性
        
        Args:
            excel_path: Excel檔案路徑
            
        Returns:
            dict: 實際表格屬性數據
        """
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel檔案不存在: {excel_path}")
        
        try:
            # 載入工作簿
            workbook = load_workbook(excel_path, data_only=True)
            
            # 建立結果結構
            result = self._create_empty_result_structure()
            result['source_file'] = os.path.basename(excel_path)
            
            # 掃描所有工作表
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                sheet_data = self._extract_sheet_table_properties(worksheet, sheet_name)
                
                if sheet_data['objects']:  # 只記錄有物件的工作表
                    result['worksheets'][sheet_name] = sheet_data
            
            # 更新摘要
            result['summary'] = self._calculate_summary(result['worksheets'])
            
            return result
            
        except Exception as e:
            raise RuntimeError(f"提取表格屬性失敗: {str(e)}")
    
    def compare_with_registry(self, actual: Dict[str, Any], 
                            registry: Dict[str, Any]) -> Dict[str, Any]:
        """
        比對實際結果與註冊表預測
        
        Args:
            actual: 實際結果數據
            registry: 註冊表預測數據
            
        Returns:
            dict: 詳細比對結果
        """
        comparison = {
            'comparison_timestamp': datetime.now().isoformat(),
            'registry_file': registry.get('metadata', {}).get('registry_filename', 'unknown'),
            'actual_file': actual.get('source_file', 'unknown'),
            'validation_results': {},
            'summary': {
                'total_worksheets': 0,
                'validated_worksheets': 0,
                'total_objects': 0,
                'validated_objects': 0,
                'failed_objects': 0,
                'accuracy_rate': '0%'
            }
        }
        
        # 比對每個工作表
        registry_worksheets = registry.get('worksheets', {})
        actual_worksheets = actual.get('worksheets', {})
        
        all_sheet_names = set(registry_worksheets.keys()) | set(actual_worksheets.keys())
        
        for sheet_name in all_sheet_names:
            registry_sheet = registry_worksheets.get(sheet_name, {})
            actual_sheet = actual_worksheets.get(sheet_name, {})
            
            sheet_comparison = self._compare_single_worksheet(
                sheet_name, registry_sheet, actual_sheet
            )
            
            comparison['validation_results'][sheet_name] = sheet_comparison
            
            # 更新摘要統計
            comparison['summary']['total_worksheets'] += 1
            if sheet_comparison['worksheet_valid']:
                comparison['summary']['validated_worksheets'] += 1
            
            comparison['summary']['total_objects'] += sheet_comparison['total_objects']
            comparison['summary']['validated_objects'] += sheet_comparison['validated_objects']
            comparison['summary']['failed_objects'] += sheet_comparison['failed_objects']
        
        # 計算準確率
        total_objects = comparison['summary']['total_objects']
        validated_objects = comparison['summary']['validated_objects']
        
        if total_objects > 0:
            accuracy = (validated_objects / total_objects) * 100
            comparison['summary']['accuracy_rate'] = f"{accuracy:.1f}%"
        
        return comparison
    
    def generate_validation_report(self, comparison: Dict[str, Any]) -> str:
        """
        生成驗證報告
        
        Args:
            comparison: 比對結果
            
        Returns:
            str: 生成的報告檔案路徑
        """
        # 產生報告檔案名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_filename = f"table_validation_report_{timestamp}.json"
        
        try:
            # 寫入報告檔案
            with open(report_filename, 'w', encoding='utf-8') as f:
                json.dump(comparison, f, ensure_ascii=False, indent=2)
            
            # 產生摘要資訊
            self._print_validation_summary(comparison)
            
            return report_filename
            
        except Exception as e:
            raise RuntimeError(f"生成驗證報告失敗: {str(e)}")
    
    def validate_with_extract_tool(self, excel_path: str) -> str:
        """
        使用 extract_table_properties.py 工具執行驗證
        
        Args:
            excel_path: Excel檔案路徑
            
        Returns:
            str: 生成的結果JSON檔案路徑
        """
        try:
            # 執行 extract_table_properties.py
            result = subprocess.run(
                ['python', 'extract_table_properties.py', excel_path],
                capture_output=True,
                text=True,
                timeout=120,  # 2分鐘超時
                cwd=os.getcwd()
            )
            
            if result.returncode != 0:
                raise RuntimeError(f"extract_table_properties.py 執行失敗: {result.stderr}")
            
            # 解析輸出找出生成的檔案
            output_lines = result.stdout.strip().split('\n')
            for line in output_lines:
                if 'template_renderer_obj_result_' in line and line.endswith('.json'):
                    json_path = line.strip()
                    if os.path.exists(json_path):
                        return json_path
            
            # 如果無法從輸出解析，嘗試找最新的結果檔案
            return self._find_latest_result_file()
            
        except subprocess.TimeoutExpired:
            raise RuntimeError("extract_table_properties.py 執行超時")
        except Exception as e:
            raise RuntimeError(f"執行驗證工具失敗: {str(e)}")
    
    def auto_validate_render_result(self, excel_path: str, registry_path: str) -> Dict[str, Any]:
        """
        自動化驗證渲染結果
        
        Args:
            excel_path: 渲染後的Excel檔案路徑
            registry_path: 註冊表JSON檔案路徑
            
        Returns:
            dict: 完整驗證結果
        """
        try:
            print(f"開始自動驗證: {excel_path}")
            
            # 1. 提取實際結果
            print("1. 提取實際表格屬性...")
            actual_result = self.extract_actual_tables(excel_path)
            
            # 2. 載入註冊表
            print("2. 載入註冊表...")
            registry = self.registry_utils.load_registry(registry_path)
            
            # 3. 執行比對
            print("3. 執行比對驗證...")
            comparison = self.compare_with_registry(actual_result, registry)
            
            # 4. 生成報告
            print("4. 生成驗證報告...")
            report_path = self.generate_validation_report(comparison)
            
            # 5. 返回完整結果
            validation_result = {
                'validation_successful': comparison['summary']['accuracy_rate'] != '0%',
                'accuracy_rate': comparison['summary']['accuracy_rate'],
                'report_file': report_path,
                'summary': comparison['summary'],
                'detailed_results': comparison
            }
            
            print(f"驗證完成，準確率: {validation_result['accuracy_rate']}")
            return validation_result
            
        except Exception as e:
            print(f"自動驗證失敗: {str(e)}")
            return {
                'validation_successful': False,
                'error': str(e),
                'timestamp': datetime.now().isoformat()
            }
    
    def _create_empty_result_structure(self) -> Dict[str, Any]:
        """建立空的結果結構"""
        return {
            'extraction_timestamp': datetime.now().isoformat(),
            'source_file': '',
            'worksheets': {},
            'summary': {
                'total_worksheets': 0,
                'total_objects': 0,
                'table_objects': 0,
                'simple_objects': 0
            }
        }
    
    def _extract_sheet_table_properties(self, worksheet, sheet_name: str) -> Dict[str, Any]:
        """提取單一工作表的表格屬性"""
        sheet_data = {
            'sheet_index': worksheet.parent.worksheets.index(worksheet),
            'objects': [],
            'tables_found': 0,
            'simple_objects_found': 0
        }
        
        # 1. 掃描Excel表格物件
        for table in worksheet.tables.values():
            table_info = self._extract_table_object_info(table, sheet_name)
            sheet_data['objects'].append(table_info)
            sheet_data['tables_found'] += 1
        
        # 2. 掃描可能的簡單物件（含有數據的區域）
        simple_objects = self._scan_simple_objects(worksheet, sheet_name)
        sheet_data['objects'].extend(simple_objects)
        sheet_data['simple_objects_found'] = len(simple_objects)
        
        return sheet_data
    
    def _extract_table_object_info(self, table, sheet_name: str) -> Dict[str, Any]:
        """提取表格物件資訊"""
        table_range = table.ref
        
        # 解析範圍
        range_parts = table_range.split(':')
        start_cell = range_parts[0]
        end_cell = range_parts[-1] if len(range_parts) > 1 else start_cell
        
        # 計算起始位置
        from openpyxl.utils import coordinate_from_string, column_index_from_string
        start_coord = coordinate_from_string(start_cell)
        end_coord = coordinate_from_string(end_cell)
        
        start_row = start_coord[1]
        start_col = column_index_from_string(start_coord[0])
        end_row = end_coord[1]
        end_col = column_index_from_string(end_coord[0])
        
        # 計算數據形狀
        rows = end_row - start_row + 1
        cols = end_col - start_col + 1
        
        return {
            'obj_name': table.name,
            'display_name': f"#{{{{{table.name}}}}}",
            'obj_type': 'table',
            'position_after': {
                'row': start_row,
                'col': start_col,
                'coordinate': start_cell,
                'range_start': start_cell,
                'range_end': end_cell,
                'range_description': table_range,
                'data_shape': {
                    'rows': rows,
                    'cols': cols
                }
            },
            'table_properties': {
                'name': table.name,
                'display_name': table.displayName,
                'range': table_range,
                'has_headers': table.headerRowCount > 0,
                'total_ref': table.ref
            }
        }
    
    def _scan_simple_objects(self, worksheet, sheet_name: str) -> List[Dict[str, Any]]:
        """掃描簡單物件"""
        simple_objects = []
        
        # 簡單掃描：找出包含模板標籤格式的儲存格
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_value = str(cell.value).strip()
                    
                    # 檢查是否為模板標籤格式
                    if (cell_value.startswith('{{') and cell_value.endswith('}}')) or \
                       (cell_value.startswith('#{{') and cell_value.endswith('}}')):
                        
                        obj_info = {
                            'obj_name': cell_value.replace('{', '').replace('}', '').replace('#', ''),
                            'display_name': cell_value,
                            'obj_type': 'table' if cell_value.startswith('#{{') else 'simple',
                            'position_after': {
                                'row': cell.row,
                                'col': cell.column,
                                'coordinate': cell.coordinate,
                                'range_start': cell.coordinate,
                                'range_end': cell.coordinate,
                                'range_description': cell.coordinate,
                                'data_shape': {
                                    'rows': 1,
                                    'cols': 1
                                }
                            }
                        }
                        
                        simple_objects.append(obj_info)
        
        return simple_objects
    
    def _calculate_summary(self, worksheets: Dict[str, Any]) -> Dict[str, Any]:
        """計算摘要統計"""
        summary = {
            'total_worksheets': len(worksheets),
            'total_objects': 0,
            'table_objects': 0,
            'simple_objects': 0
        }
        
        for sheet_data in worksheets.values():
            objects = sheet_data.get('objects', [])
            summary['total_objects'] += len(objects)
            
            for obj in objects:
                if obj.get('obj_type') == 'table':
                    summary['table_objects'] += 1
                else:
                    summary['simple_objects'] += 1
        
        return summary
    
    def _compare_single_worksheet(self, sheet_name: str, registry_sheet: Dict[str, Any], 
                                actual_sheet: Dict[str, Any]) -> Dict[str, Any]:
        """比對單一工作表"""
        comparison = {
            'worksheet_name': sheet_name,
            'worksheet_valid': True,
            'total_objects': 0,
            'validated_objects': 0,
            'failed_objects': 0,
            'object_validations': {},
            'missing_in_actual': [],
            'extra_in_actual': []
        }
        
        registry_objects = registry_sheet.get('objects', [])
        actual_objects = actual_sheet.get('objects', [])
        
        # 建立物件名稱索引
        registry_index = {obj.get('obj_name'): obj for obj in registry_objects}
        actual_index = {obj.get('obj_name'): obj for obj in actual_objects}
        
        # 比對註冊表中的物件
        for reg_obj in registry_objects:
            obj_name = reg_obj.get('obj_name')
            comparison['total_objects'] += 1
            
            actual_obj = actual_index.get(obj_name)
            if actual_obj:
                obj_validation = self._validate_object_match(reg_obj, actual_obj)
                comparison['object_validations'][obj_name] = obj_validation
                
                if obj_validation['match_successful']:
                    comparison['validated_objects'] += 1
                else:
                    comparison['failed_objects'] += 1
                    comparison['worksheet_valid'] = False
            else:
                comparison['missing_in_actual'].append(obj_name)
                comparison['failed_objects'] += 1
                comparison['worksheet_valid'] = False
        
        # 找出實際結果中多出的物件
        for actual_obj in actual_objects:
            obj_name = actual_obj.get('obj_name')
            if obj_name not in registry_index:
                comparison['extra_in_actual'].append(obj_name)
        
        return comparison
    
    def _validate_object_match(self, registry_obj: Dict[str, Any], 
                             actual_obj: Dict[str, Any]) -> Dict[str, Any]:
        """驗證單一物件匹配"""
        validation = {
            'match_successful': True,
            'differences': []
        }
        
        # 比較物件類型
        reg_type = registry_obj.get('obj_type')
        actual_type = actual_obj.get('obj_type')
        if reg_type != actual_type:
            validation['match_successful'] = False
            validation['differences'].append({
                'field': 'obj_type',
                'expected': reg_type,
                'actual': actual_type
            })
        
        # 比較位置資訊
        reg_pos = registry_obj.get('position_after', {})
        actual_pos = actual_obj.get('position_after', {})
        
        # 範圍比較
        reg_range = reg_pos.get('range_description')
        actual_range = actual_pos.get('range_description')
        if reg_range != actual_range:
            validation['match_successful'] = False
            validation['differences'].append({
                'field': 'range_description',
                'expected': reg_range,
                'actual': actual_range
            })
        
        # 數據形狀比較
        reg_shape = reg_pos.get('data_shape', {})
        actual_shape = actual_pos.get('data_shape', {})
        if reg_shape != actual_shape:
            validation['match_successful'] = False
            validation['differences'].append({
                'field': 'data_shape',
                'expected': reg_shape,
                'actual': actual_shape
            })
        
        return validation
    
    def _print_validation_summary(self, comparison: Dict[str, Any]):
        """列印驗證摘要"""
        summary = comparison['summary']
        
        print("\n=== 表格驗證報告摘要 ===")
        print(f"總工作表數: {summary['total_worksheets']}")
        print(f"總物件數: {summary['total_objects']}")
        print(f"驗證通過: {summary['validated_objects']}")
        print(f"驗證失敗: {summary['failed_objects']}")
        print(f"準確率: {summary['accuracy_rate']}")
        
        # 列出失敗的物件
        for sheet_name, sheet_result in comparison['validation_results'].items():
            if not sheet_result['worksheet_valid']:
                print(f"\n工作表 '{sheet_name}' 驗證失敗:")
                
                for obj_name, obj_result in sheet_result['object_validations'].items():
                    if not obj_result['match_successful']:
                        print(f"  - 物件 '{obj_name}': {len(obj_result['differences'])} 個差異")
                        for diff in obj_result['differences']:
                            print(f"    * {diff['field']}: 預期 '{diff['expected']}' 實際 '{diff['actual']}'")
                
                if sheet_result['missing_in_actual']:
                    print(f"  - 缺少物件: {sheet_result['missing_in_actual']}")
                
                if sheet_result['extra_in_actual']:
                    print(f"  - 多出物件: {sheet_result['extra_in_actual']}")
        
        print("========================\n")
    
    def _find_latest_result_file(self) -> str:
        """找出最新的結果檔案"""
        import glob
        
        pattern = "template_renderer_obj_result_*.json"
        result_files = glob.glob(pattern)
        
        if not result_files:
            raise RuntimeError("找不到結果檔案")
        
        # 按修改時間排序，取最新的
        latest_file = max(result_files, key=os.path.getmtime)
        return latest_file