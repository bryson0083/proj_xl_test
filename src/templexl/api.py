"""
主要 API 介面。
"""
import logging
import os
import re
import uuid
from typing import Any, Dict, Optional

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

from .core.parser import TemplateParser
from .core.renderer import TemplateRenderer
from .core.container import ContainerManager
from .core.block_manager import BlockManager
from .exceptions import TemplateError, TemplateNotFoundError, FileFormatError, RenderError
from .context import RenderContext
from .models.container import Container
from .result import RenderResult
from .report import RenderReport


def render(
    template: str,
    output: str,
    data: Optional[Dict[str, Any]] = None,
    *,
    with_report: bool = False,
) -> RenderResult:
    """以資料渲染 Excel 模板並寫出報表。

    Args:
        template: 模板檔路徑（``.xlsx`` 或 ``.xlsm``）。
        output: 輸出檔路徑。
        data: 渲染資料；鍵為標籤名稱，值為純量或 pandas DataFrame。
        with_report: 是否產生渲染報告（除錯/維護用），預設關閉。
            啟用時報告以記憶體物件掛在 ``RenderResult.report``，不會自動寫入磁碟。

    Returns:
        RenderResult: 含 ``output_path``、``warnings``，以及（啟用時）``report``。

    Raises:
        TemplateNotFoundError: 模板檔不存在。
        FileFormatError: 不支援的檔案格式。
        RenderError: 渲染過程發生錯誤。

    Examples:
        >>> import pandas as pd
        >>> from templexl import render
        >>> result = render(
        ...     "template.xlsx",
        ...     "output.xlsx",
        ...     data={
        ...         "oper_name": "OPER_NAME",
        ...         "report_df": pd.DataFrame({"姓名": ["Alice", "Bob"]}),
        ...     },
        ... )
        >>> result.output_path
        'output.xlsx'
    """
    template = str(template)
    output = str(output)
    data = dict(data) if data else {}

    try:
        if not os.path.exists(template):
            raise TemplateNotFoundError(template)
        if not _is_supported_format(template):
            raise FileFormatError(template)

        workbook = load_workbook(template)
        render_context = RenderContext(
            process_id=str(uuid.uuid4()),
            template_path=template,
            output_path=output,
            data=data,
        )

        containers = _execute_render_pipeline(workbook, render_context)
        _final_table_autofilter_sync(workbook)
        workbook.save(output)

        warnings = _collect_unresolved_tag_warnings(workbook)
        report = (
            _build_render_report(containers, render_context, template, output)
            if with_report
            else None
        )
        return RenderResult(output_path=output, report=report, warnings=warnings)

    except TemplateError:
        # 本套件的已知例外（含 TemplateNotFoundError/FileFormatError/RenderError）原樣往上拋
        raise
    except Exception as e:
        raise RenderError(f"渲染過程發生未預期錯誤: {str(e)}")


def _collect_unresolved_tag_warnings(workbook: Workbook) -> list:
    """掃描輸出工作簿，回報任何仍未被解析的標籤（非致命警告）。"""
    warnings: list = []
    pattern = re.compile(r"#?\{\{.*?\}\}")
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and pattern.search(cell.value):
                    warnings.append(
                        f"未解析的標籤 '{cell.value}'"
                        f"（工作表 '{worksheet.title}' 儲存格 {cell.coordinate}）"
                    )
    return warnings


def _is_supported_format(file_path: str) -> bool:
    """
    檢查檔案格式是否支援
    
    Args:
        file_path: 檔案路徑
        
    Returns:
        bool: 是否支援
    """
    supported_extensions = ['.xlsx', '.xlsm']
    file_ext = os.path.splitext(file_path)[1].lower()
    return file_ext in supported_extensions


def _execute_render_pipeline(workbook: Workbook, render_context: 'RenderContext') -> list:
    """
    執行渲染管道

    Args:
        workbook: Excel工作簿
        render_context: 渲染上下文

    Returns:
        list: 容器清單
    """
    # 使用新的架構：統一的模板掃描與註冊表建立
    container_manager = ContainerManager()
    
    # 重要修正：先掃描原始標籤，保留所有標籤的完整資訊
    from .core.parser import TemplateParser
    parser = TemplateParser()
    original_tags = parser.parse_template(workbook)
    
    # 建立容器
    containers = container_manager.create_containers(workbook)
    
    # 3. 區塊分類
    for container in containers:
        container_manager.classify_blocks(container)
    
    # 4. 建立標籤到物件的映射關係
    # 使用原始標籤資訊建立映射，確保相同名稱的多個標籤都能正確處理
    _build_tag_object_mapping_with_original_tags(containers, original_tags, render_context)
    
    # 5. 創建渲染器和區塊管理器
    renderer = TemplateRenderer()
    block_manager = BlockManager()
    
    # 6. 按區塊順序執行渲染
    for container in containers:
        _render_container(container, workbook, render_context, renderer, block_manager)

    return containers


def _render_container(
    container: Container,
    workbook: Workbook,
    render_context: RenderContext,
    renderer: 'TemplateRenderer', 
    block_manager: 'BlockManager'
) -> None:
    """
    渲染單一容器
    
    Args:
        container: 容器物件
        workbook: Excel工作簿
        render_context: 渲染上下文
        renderer: 渲染器
        block_manager: 區塊管理器
    """
    logger.debug(f"DEBUG: 使用新的block分區渲染搬移機制處理容器: {container.sheet_name}")
    logger.debug(f"DEBUG: Calling block_manager.process_container_with_block_moving")
    
    try:
        # 使用新的block搬移機制
        block_manager.process_container_with_block_moving(container, workbook, render_context, renderer)
        logger.debug(f"DEBUG: Successfully called process_container_with_block_moving")
    except Exception as e:
        logger.debug(f"DEBUG: Error in process_container_with_block_moving: {e}")
        import traceback
        logger.debug("例外堆疊", exc_info=True)
        raise
    
    logger.debug(f"DEBUG: Finished processing container: {container.sheet_name}")


def _build_tag_object_mapping_with_original_tags(containers, original_tags, render_context):
    """
    使用原始標籤資訊建立標籤到物件的映射關係
    
    確保相同名稱的多個標籤都能正確對應到各自的物件
    
    Args:
        containers: 容器清單
        original_tags: 原始標籤清單（來自TemplateParser的掃描結果）
        render_context: 渲染上下文
    """
    from .models.base import ObjectType
    
    logger.debug(f"DEBUG_MAPPING: Building tag-object mapping with {len(original_tags)} original tags")
    
    # 建立位置到標籤清單的映射（支援同一位置多個標籤）
    position_to_tags = {}
    for tag in original_tags:
        key = (tag.sheet_name, tag.cell_position.row, tag.cell_position.col)
        if key not in position_to_tags:
            position_to_tags[key] = []
        position_to_tags[key].append(tag)
        logger.debug(f"DEBUG_MAPPING: Tag {tag.tag_name} at position {key}")
    
    # 為每個物件找到對應的標籤
    for container in containers:
        logger.debug(f"DEBUG_MAPPING: Processing container {container.sheet_name} with {len(container.objects)} objects")
        for obj in container.objects:
            # 處理標籤相關的物件（SIMPLE、TABLE和TABLE_OBJ類型）
            # TABLE_OBJ是標籤與表格物件綁定後的類型
            if obj.obj_type in [ObjectType.SIMPLE, ObjectType.TABLE, ObjectType.TABLE_OBJ]:
                key = (obj.sheet_name, obj.cell_position.row, obj.cell_position.col)
                logger.debug(f"DEBUG_MAPPING: Object {obj.obj_id} type={obj.obj_type} display_name={obj.display_name} at position {key}")
                # 對於TABLE_OBJ類型，需要特殊處理（因為綁定改變了位置）
                if obj.obj_type == ObjectType.TABLE_OBJ:
                    # TABLE_OBJ的display_name保留了原始標籤名
                    # 需要在所有標籤中查找匹配的標籤名
                    matched_tag = None
                    for pos_key, tags in position_to_tags.items():
                        if pos_key[0] == obj.sheet_name:  # 同一工作表
                            for tag in tags:
                                if tag.tag_name == obj.display_name:
                                    matched_tag = tag
                                    logger.debug(f"DEBUG_MAPPING: Found tag {tag.tag_name} for TABLE_OBJ {obj.obj_id}")
                                    break
                            if matched_tag:
                                break
                elif key in position_to_tags:
                    # 獲取該位置的所有標籤
                    candidate_tags = position_to_tags[key]
                    
                    # 嘗試根據物件名稱匹配標籤
                    matched_tag = None
                    for tag in candidate_tags:
                        # 檢查物件ID是否包含標籤名稱
                        if tag.tag_name in obj.obj_id:
                            matched_tag = tag
                            break
                    
                    # 如果沒有匹配到，使用第一個標籤作為備選
                    if not matched_tag and candidate_tags:
                        matched_tag = candidate_tags[0]
                else:
                    matched_tag = None
                    
                if matched_tag:
                    render_context.add_tag_mapping(obj.obj_id, matched_tag)
                    logger.debug(f"DEBUG: 映射物件 {obj.obj_id} 到標籤 {matched_tag.tag_name} 在位置 ({matched_tag.cell_position.row}, {matched_tag.cell_position.col})")
                else:
                    logger.debug(f"DEBUG: 警告 - 無法找到物件 {obj.obj_id} 的匹配標籤 at position {key}")


def _final_table_autofilter_sync(workbook: Workbook) -> None:
    """
    在儲存前進行最終的表格範圍同步檢查
    確保所有表格的 autoFilter.ref 與 table.ref 保持一致
    
    Args:
        workbook: Excel工作簿
    """
    logger.debug("DEBUG_FINAL_SYNC: 開始進行最終的表格範圍同步檢查")
    sync_count = 0
    
    for worksheet in workbook.worksheets:
        logger.debug(f"DEBUG_FINAL_SYNC: 檢查工作表 {worksheet.title}")
        
        for table_name in worksheet.tables:
            table = worksheet.tables[table_name]
            table_ref = getattr(table, 'ref', '')
            
            if hasattr(table, 'autoFilter') and table.autoFilter:
                autofilter_ref = table.autoFilter.ref
                
                if table_ref != autofilter_ref:
                    logger.debug(f"DEBUG_FINAL_SYNC: 發現不同步 - 表格 {table_name}")
                    logger.debug(f"  table.ref: {table_ref}")
                    logger.debug(f"  autoFilter.ref: {autofilter_ref}")
                    logger.debug(f"  正在同步...")
                    
                    # 同步 autoFilter.ref
                    table.autoFilter.ref = table_ref
                    sync_count += 1
                    
                    logger.debug(f"  已同步為: {table.autoFilter.ref}")
                else:
                    logger.debug(f"DEBUG_FINAL_SYNC: 表格 {table_name} 範圍已同步: {table_ref}")
            else:
                logger.debug(f"DEBUG_FINAL_SYNC: 表格 {table_name} 沒有 autoFilter")
    
    logger.debug(f"DEBUG_FINAL_SYNC: 同步檢查完成，共修正 {sync_count} 個表格")


def _build_render_report(
    containers: list,
    render_context: 'RenderContext',
    template_path: str,
    output_path: str,
) -> RenderReport:
    """建立渲染報告（記憶體物件，不寫入磁碟）。

    報告內容直接取自渲染後的容器狀態（單一真相來源），不另行重算物件落點，
    避免「報告與實際輸出漂移」的維護陷阱。欄名以 ``get_column_letter`` 轉換，
    正確支援超過 Z 欄的多字母欄名。

    Args:
        containers: 渲染後的容器清單。
        render_context: 渲染上下文（用於對應原始模板標籤）。
        template_path: 模板文件路徑。
        output_path: 輸出文件路徑。

    Returns:
        RenderReport: 物件清單與摘要。
    """
    worksheets: dict = {}
    total_objects = 0

    for container in containers:
        objects = []
        for obj in container.objects:
            coordinate = f"{get_column_letter(obj.cell_position.col)}{obj.cell_position.row}"
            tag = render_context.get_tag_for_object(obj.obj_id)
            objects.append({
                'obj_id': obj.obj_id,
                'display_name': obj.display_name,
                'obj_type': obj.obj_type.value if hasattr(obj.obj_type, 'value') else str(obj.obj_type),
                'sheet_name': obj.sheet_name,
                'having_header': obj.having_header,
                'is_multi_rows': obj.is_multi_rows,
                'block_id': obj.block_id,
                'template_tag': tag.tag_name if tag else None,
                # 渲染後實際狀態（單一真相來源：渲染流程已更新的容器物件）
                'position': {
                    'row': obj.cell_position.row,
                    'col': obj.cell_position.col,
                    'coordinate': coordinate,
                },
                'data_shape': {
                    'rows': obj.data_shape.rows,
                    'cols': obj.data_shape.cols,
                },
            })
            total_objects += 1

        worksheets[container.sheet_name] = {
            'container_id': container.container_id,
            'sheet_name': container.sheet_name,
            'total_objects': len(container.objects),
            'total_blocks': len(container.blocks),
            'objects': objects,
        }

    summary = {
        'total_worksheets': len(containers),
        'total_objects': total_objects,
    }

    return RenderReport(
        template_path=template_path,
        output_path=output_path,
        worksheets=worksheets,
        summary=summary,
    )
