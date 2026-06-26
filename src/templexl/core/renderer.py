"""
數據渲染器
"""
import logging

logger = logging.getLogger(__name__)

from typing import Any, Union, Optional
import pandas as pd
from datetime import datetime, date

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import NamedStyle

from ..models.base import DataType, ObjectType
from ..models.tag import Tag
from ..models.container import Container
from ..models.objects import ObjectInfo
from ..exceptions import InvalidDataTypeError, RenderError


class TemplateRenderer:
    """
    模板渲染器類別
    
    負責將數據渲染到Excel模板中
    """
    
    def _update_table_range_sync(self, table, new_range: str, debug_name: str = "", dataframe=None, include_header: bool = True):
        """
        統一更新表格範圍，確保 table.ref 和 autoFilter.ref 保持同步
        並正確設定所有必要的表格屬性以避免Excel錯誤

        Args:
            table: Excel表格物件
            new_range: 新的範圍字串 (例如: "A2:F5")
            debug_name: 用於DEBUG訊息的物件名稱
            dataframe: DataFrame物件（用於更新tableColumns）
            include_header: 是否包含表頭
        """
        old_table_ref = table.ref if hasattr(table, 'ref') else 'Unknown'
        old_autofilter_ref = table.autoFilter.ref if (hasattr(table, 'autoFilter') and table.autoFilter) else 'None'

        # *** 保留原始表格樣式設定 ***
        original_style_info = self._preserve_table_style(table, debug_name)

        # 更新表格範圍
        table.ref = new_range
        logger.debug(f"DEBUG_SYNC: 表格 {debug_name} - table.ref 更新: {old_table_ref} -> {new_range}")

        # *** 關鍵修正：設定表格屬性避免Excel錯誤 ***

        # 1. 處理headerRowCount（修正：noheader模式仍有表頭，只是保留模板表頭）
        if hasattr(table, 'headerRowCount'):
            old_headerRowCount = table.headerRowCount
            # noheader模式實際上仍有表頭（保留的模板表頭），所以headerRowCount應該保持為1
            table.headerRowCount = 1  # 所有表格都有表頭（noheader模式保留模板表頭）
            logger.debug(f"DEBUG_SYNC: 表格 {debug_name} - headerRowCount設為: {table.headerRowCount} (include_header={include_header}, 說明: {'動態表頭' if include_header else '保留模板表頭'})")

        # 1.5. 處理autoFilter（同步更新範圍，不清除）
        if hasattr(table, 'autoFilter') and table.autoFilter:
            table.autoFilter.ref = new_range
            logger.debug(f"DEBUG_SYNC: 表格 {debug_name} - autoFilter.ref 同步更新: {old_autofilter_ref} -> {new_range}")
        else:
            logger.debug(f"DEBUG_SYNC: 表格 {debug_name} - 沒有 autoFilter，跳過同步")

        # 2. 設定totalsRowCount（關鍵修正：對於一般數據表格應設為None）
        if hasattr(table, 'totalsRowCount'):
            table.totalsRowCount = None  # 不需要合計行
            logger.debug(f"DEBUG_SYNC: 表格 {debug_name} - totalsRowCount 設為: {table.totalsRowCount}")

        # 3. 設定totalsRowShown（關鍵修正：應設為False）
        if hasattr(table, 'totalsRowShown'):
            table.totalsRowShown = False
            logger.debug(f"DEBUG_SYNC: 表格 {debug_name} - totalsRowShown 設為: {table.totalsRowShown}")

        # 4. 更新tableColumns（重要修正：根據include_header決定是否更新列定義）
        if (dataframe is not None and hasattr(dataframe, 'columns') and
            hasattr(table, 'tableColumns')):
            try:
                from openpyxl.worksheet.table import TableColumn

                if include_header:
                    # 有header模式：正常更新列定義
                    table.tableColumns = []
                    for i, column_name in enumerate(dataframe.columns):
                        col = TableColumn(id=i+1, name=str(column_name))
                        table.tableColumns.append(col)
                    logger.debug(f"DEBUG_SYNC: 表格 {debug_name} - 更新tableColumns: {len(table.tableColumns)} 列")
                    for i, col in enumerate(table.tableColumns):
                        logger.debug(f"  列{i+1}: {col.name}")
                else:
                    # noheader模式：完全保留原始tableColumns，不做任何修改
                    original_columns_count = len(table.tableColumns) if table.tableColumns else 0
                    logger.debug(f"DEBUG_SYNC: 表格 {debug_name} - noheader模式，完全保留原始tableColumns")
                    logger.debug(f"  原始列數: {original_columns_count}")

                    # 顯示原始的tableColumns
                    if table.tableColumns:
                        for i, col in enumerate(table.tableColumns):
                            logger.debug(f"  保留列{i+1}: {col.name}")
                    else:
                        logger.debug(f"  警告：原始tableColumns為空")

            except Exception as e:
                logger.debug(f"DEBUG_SYNC: 表格 {debug_name} - tableColumns處理失敗: {str(e)}")

        # 5. 恢復原始表格樣式設定
        self._restore_table_style(table, original_style_info, debug_name)

        # 6. 最終驗證表格屬性的一致性
        self._validate_table_properties(table, debug_name)

        logger.debug(f"DEBUG_SYNC: 表格 {debug_name} - 所有屬性更新完成")

    def _validate_table_properties(self, table, debug_name: str = ""):
        """
        驗證表格屬性的一致性，確保不會導致Excel錯誤

        Args:
            table: Excel表格物件
            debug_name: 用於DEBUG訊息的物件名稱
        """
        try:
            # 驗證基本屬性
            if not hasattr(table, 'ref') or not table.ref:
                logger.debug(f"DEBUG_VALIDATE: 警告 - 表格 {debug_name} 缺少有效的 ref 屬性")
                return

            # 驗證 autoFilter 與 table.ref 的一致性
            if hasattr(table, 'autoFilter') and table.autoFilter:
                if hasattr(table.autoFilter, 'ref'):
                    if table.autoFilter.ref != table.ref:
                        logger.debug(f"DEBUG_VALIDATE: 警告 - 表格 {debug_name} autoFilter.ref 與 table.ref 不一致:")
                        logger.debug(f"  table.ref: {table.ref}")
                        logger.debug(f"  autoFilter.ref: {table.autoFilter.ref}")
                        # 強制同步
                        table.autoFilter.ref = table.ref
                        logger.debug(f"DEBUG_VALIDATE: 已強制同步 autoFilter.ref 到: {table.ref}")

            # 驗證 totalsRowCount 和 totalsRowShown 的一致性
            if hasattr(table, 'totalsRowCount') and hasattr(table, 'totalsRowShown'):
                totals_count = table.totalsRowCount
                totals_shown = table.totalsRowShown

                if totals_count is None or totals_count == 0:
                    if totals_shown is True:
                        logger.debug(f"DEBUG_VALIDATE: 修正 - 表格 {debug_name} totalsRowCount 為 {totals_count} 但 totalsRowShown 為 True")
                        table.totalsRowShown = False
                        logger.debug(f"DEBUG_VALIDATE: 已修正 totalsRowShown 為: False")

            # 驗證 headerRowCount 的合理性
            if hasattr(table, 'headerRowCount'):
                header_count = table.headerRowCount
                if header_count is not None and header_count not in [0, 1]:
                    logger.debug(f"DEBUG_VALIDATE: 警告 - 表格 {debug_name} headerRowCount 值異常: {header_count}")

            logger.debug(f"DEBUG_VALIDATE: 表格 {debug_name} 屬性驗證完成")

        except Exception as e:
            logger.debug(f"DEBUG_VALIDATE: 表格 {debug_name} 屬性驗證失敗: {str(e)}")
    
    def render_simple_tag(
        self, 
        tag: Tag, 
        data: Any, 
        workbook: Workbook, 
        worksheet: Worksheet
    ) -> None:
        """
        渲染簡單變數標籤
        
        Args:
            tag: 標籤物件
            data: 要渲染的數據
            workbook: Excel工作簿
            worksheet: 工作表
            
        Raises:
            InvalidDataTypeError: 不支援的數據類型
            RenderError: 渲染過程錯誤
        """
        try:
            # 檢查是否為DataFrame
            if hasattr(data, 'shape') and hasattr(data, 'columns'):
                # DataFrame處理
                self._render_simple_dataframe(tag, data, worksheet)
            else:
                # 一般數據處理
                self._render_simple_data(tag, data, worksheet)
            
        except Exception as e:
            raise RenderError(
                f"渲染簡單標籤失敗: {str(e)}", 
                tag_name=tag.tag_name, 
                sheet_name=worksheet.title
            )
    
    def _render_simple_data(self, tag: Tag, data: Any, worksheet: Worksheet) -> None:
        """渲染簡單數據到單一儲存格"""
        # 取得目標儲存格
        target_cell = worksheet.cell(row=tag.cell_position.row, column=tag.cell_position.col)
        
        # 檢查是否為合併儲存格，如果是則跳過
        from openpyxl.cell.cell import MergedCell
        if isinstance(target_cell, MergedCell):
            return
        
        # 獲取原始值
        original_value = target_cell.value
        
        # 處理標籤替換
        if original_value and isinstance(original_value, str):
            # 使用正則表達式替換，忽略標籤中的空白
            import re
            # 構建支援任意空白的標籤模式：{{ tag_name }} 或 {{tag_name}}
            tag_pattern = rf"\{{\{{\s*{re.escape(tag.tag_name)}\s*\}}\}}"
            
            # 準備替換的數據值
            replacement_value = self._prepare_data_for_replacement(data, tag)
            
            # 執行標籤替換（使用正則表達式）
            new_value = re.sub(tag_pattern, str(replacement_value), original_value)
            target_cell.value = new_value
        else:
            # 如果儲存格沒有文本，直接設定數據值
            replacement_value = self._prepare_data_for_replacement(data, tag)
            target_cell.value = replacement_value
    
    def _render_simple_dataframe(self, tag: Tag, dataframe, worksheet: Worksheet) -> None:
        """渲染簡單DataFrame標籤到多個儲存格"""
        import pandas as pd
        
        start_row = tag.cell_position.row
        start_col = tag.cell_position.col
        
        # Simple標籤邏輯：
        # 1. 有noheader條件：只渲染數據，不渲染header
        # 2. 無noheader條件：渲染header + 數據
        # 注意：simple模板標籤不需要考慮往下推移位置與cell風格樣式，直接替換數據即可
        
        has_noheader = tag.has_condition and tag.condition == "noheader"
        
        # 先清除原始標籤文字
        original_cell = worksheet.cell(row=start_row, column=start_col)
        if original_cell.value and isinstance(original_cell.value, str):
            # 清除包含標籤的文字
            import re
            tag_pattern = rf"\{{\{{\s*{re.escape(tag.tag_name)}\s*(\|\s*[^}}]*?)?\s*\}}\}}"
            if re.search(tag_pattern, original_cell.value):
                original_cell.value = None
        
        # 檢查是否為合併儲存格的一部分，如果是則需要處理合併範圍內的渲染
        merge_range = self._find_merged_range_for_cell(worksheet, start_row, start_col)
        if merge_range:
            # 如果標籤在合併儲存格中，使用合併儲存格的左上角作為開始位置
            start_row = merge_range.min_row
            start_col = merge_range.min_col
            logger.debug(f"DEBUG: 標籤 {tag.tag_name} 在合併儲存格中，調整起始位置為 ({start_row}, {start_col})")
        
        # 標準Simple標籤處理：根據是否有noheader決定是否渲染header
        logger.debug(f"DEBUG_RENDER: Calling _render_simple_full_dataframe for {tag.tag_name}, has_noheader={has_noheader}")
        self._render_simple_full_dataframe(dataframe, worksheet, start_row, start_col, has_noheader)
        
        # 設定標籤數據類型
        tag.data_type = DataType.DATAFRAME
    
    def _render_simple_full_dataframe(
        self,
        dataframe: pd.DataFrame,
        worksheet: Worksheet,
        start_row: int,
        start_col: int,
        has_noheader: bool = False
    ) -> None:
        """
        渲染簡單標籤的完整DataFrame（包括表頭和數據）
        
        Args:
            dataframe: DataFrame物件
            worksheet: 工作表
            start_row: 開始行
            start_col: 開始列
            has_noheader: 是否跳過表頭（noheader條件）
        """
        import pandas as pd
        from datetime import datetime, date
        
        logger.debug(f"DEBUG_SIMPLE_FULL: Entered _render_simple_full_dataframe, start=({start_row},{start_col}), noheader={has_noheader}")
        logger.debug(f"DEBUG_SIMPLE_FULL: DataFrame shape={dataframe.shape}, columns={list(dataframe.columns)}")
        
        current_row = start_row
        
        # 渲染表頭（如果需要）
        if not has_noheader:
            for col_idx, col_name in enumerate(dataframe.columns):
                cell = worksheet.cell(row=current_row, column=start_col + col_idx)
                cell.value = str(col_name)
            current_row += 1
        
        # 渲染數據行
        for row_idx, (_, row_data) in enumerate(dataframe.iterrows()):
            logger.debug(f"DEBUG_UNIVERSAL_MERGE: 處理第{row_idx}行數據，行號{current_row + row_idx}")
            
            # 追蹤累積的列偏移量
            cumulative_offset = 0
            
            for col_idx, value in enumerate(row_data):
                # 計算目標列位置（考慮累積偏移）
                target_row = current_row + row_idx
                target_col = start_col + col_idx + cumulative_offset
                
                # 檢查是否是合併儲存格的主要儲存格（左上角）
                merge_range = self._find_merged_range_for_cell(worksheet, target_row, target_col)
                if merge_range and merge_range.min_row == target_row and merge_range.min_col == target_col:
                    # 這是合併範圍的主儲存格，計算跨度
                    merge_span = merge_range.max_col - merge_range.min_col
                    if merge_span > 0:
                        cumulative_offset += merge_span
                        logger.debug(f"DEBUG_UNIVERSAL_MERGE: col_idx={col_idx} 是合併主儲存格，跨度={merge_span}，累積偏移={cumulative_offset}")
                
                # 檢查目標儲存格是否可用，如果被合併占用則找下一個可用列
                elif self._is_cell_merged_and_not_top_left(worksheet, target_row, target_col):
                    old_col = target_col
                    target_col = self._find_next_available_column(worksheet, target_row, target_col)
                    additional_offset = target_col - old_col
                    cumulative_offset += additional_offset
                    logger.debug(f"DEBUG_UNIVERSAL_MERGE: col_idx={col_idx}被合併占用，從{old_col}跳到{target_col}，額外偏移={additional_offset}")
                
                cell = worksheet.cell(row=target_row, column=target_col)
                
                # 處理不同的數據類型
                if pd.isna(value):
                    cell.value = None
                elif isinstance(value, (datetime, date)):
                    cell.value = value
                elif isinstance(value, (int, float, bool)):
                    cell.value = value
                else:
                    cell.value = str(value)
    
    def _prepare_data_for_replacement(self, data: Any, tag: Tag) -> Any:
        """
        準備用於替換的數據值
        
        Args:
            data: 原始數據
            tag: 標籤物件
            
        Returns:
            Any: 準備好的數據值
        """
        if isinstance(data, str):
            tag.data_type = DataType.STRING
            return data
            
        elif isinstance(data, (int, float)):
            tag.data_type = DataType.NUMBER
            return data
            
        elif isinstance(data, (datetime, date)):
            tag.data_type = DataType.DATE
            return data
            
        else:
            # 嘗試轉換為字串
            tag.data_type = DataType.STRING
            return str(data)
    
    def render_table_tag(
        self, 
        tag: Tag, 
        dataframe: pd.DataFrame, 
        workbook: Workbook, 
        worksheet: Worksheet,
        obj_info: ObjectInfo
    ) -> int:
        """
        渲染表格數據標籤
        
        Args:
            tag: 標籤物件
            dataframe: 要渲染的DataFrame
            workbook: Excel工作簿
            worksheet: 工作表
            obj_info: 物件資訊
            
        Returns:
            int: 實際渲染的行數（用於計算位移）
            
        Raises:
            InvalidDataTypeError: 數據類型錯誤
            RenderError: 渲染過程錯誤
        """
        if not isinstance(dataframe, pd.DataFrame):
            raise InvalidDataTypeError(type(dataframe).__name__, tag.tag_name)
        
        try:
            start_row = tag.cell_position.row
            start_col = tag.cell_position.col
            template_position = (start_row, start_col)
            
            # 檢查是否有noheader條件
            skip_header = tag.has_condition and tag.condition == "noheader"
            
            # *** 移除重複的template rows複製邏輯 ***
            # 新的流程對齊機制中，template rows複製已在BlockManager._process_template_rows_by_render_order中處理
            # 這裡不再重複插入行，避免雙重插入問題
            data_rows_needed = len(dataframe)
            header_rows = 1 if (not skip_header and obj_info.having_header) else 0
            total_rows_needed = data_rows_needed + header_rows
            
            logger.debug(f"DEBUG_TABLE_TAG: render_table_tag for {tag.tag_name}, skip_header={skip_header}, data_rows={data_rows_needed}")
            logger.debug(f"DEBUG_TABLE_TAG: DataFrame columns={list(dataframe.columns)}")
            # 移除：self._copy_template_rows(worksheet, start_row, total_rows_needed - 1)
            
            # 渲染表頭（如果需要）
            current_row = start_row
            if not skip_header and obj_info.having_header:
                self._render_dataframe_header(dataframe, worksheet, current_row, start_col, template_position)
                current_row += 1
            
            # 渲染數據行
            rows_rendered = self._render_dataframe_data(dataframe, worksheet, current_row, start_col, template_position)
            
            # 更新物件資訊
            total_rows = rows_rendered
            if not skip_header and obj_info.having_header:
                total_rows += 1
            
            obj_info.data_shape.rows = total_rows
            obj_info.data_shape.cols = len(dataframe.columns)
            tag.data_type = DataType.DATAFRAME
            
            # 如果這是一個TABLE_OBJ類型，需要更新表格物件的範圍
            logger.debug(f"DEBUG_TABLE_CHECK: obj_info.obj_type = {obj_info.obj_type}, ObjectType.TABLE_OBJ = {ObjectType.TABLE_OBJ}")
            logger.debug(f"DEBUG_TABLE_CHECK: obj_info.obj_type == ObjectType.TABLE_OBJ: {obj_info.obj_type == ObjectType.TABLE_OBJ}")
            logger.debug(f"DEBUG_TABLE_CHECK: obj_info.display_name = {obj_info.display_name}")
            
            if obj_info.obj_type == ObjectType.TABLE_OBJ:
                logger.debug(f"DEBUG_TABLE_OBJ: render_table_tag handling TABLE_OBJ for {obj_info.display_name}")
                self._update_table_object_range(worksheet, obj_info, dataframe, tag, start_row, start_col, total_rows)
            else:
                logger.debug(f"DEBUG_TABLE_CHECK: 不是 TABLE_OBJ 類型，跳過範圍更新")
            
            return total_rows
            
        except Exception as e:
            raise RenderError(
                f"渲染表格標籤失敗: {str(e)}", 
                tag_name=tag.tag_name, 
                sheet_name=worksheet.title
            )
    
    def render_table_object(
        self,
        tag: Tag,
        dataframe: pd.DataFrame,
        workbook: Workbook,
        worksheet: Worksheet,
        obj_info: ObjectInfo
    ) -> int:
        """
        渲染表格物件 - 支援多表格gap計算和位置動態調整

        Args:
            tag: 標籤物件
            dataframe: 要渲染的DataFrame
            workbook: Excel工作簿
            worksheet: 工作表
            obj_info: 物件資訊

        Returns:
            int: 實際渲染的行數

        Raises:
            RenderError: 渲染過程錯誤
        """
        if not isinstance(dataframe, pd.DataFrame):
            raise InvalidDataTypeError(type(dataframe).__name__, tag.tag_name)

        try:
            # 找到對應的表格物件
            table_name = obj_info.display_name
            logger.debug(f"DEBUG_TABLE_OBJ: render_table_object called for {table_name}")
            if table_name not in worksheet.tables:
                raise RenderError(f"找不到表格物件: {table_name}")

            table = worksheet.tables[table_name]

            # 取得表格範圍
            table_range = table.ref
            start_cell, end_cell = table_range.split(':')
            original_start_row, start_col = self._parse_cell_reference(start_cell)
            original_end_row, end_col = self._parse_cell_reference(end_cell)

            # *** 新增：計算動態起始位置（gap區域計算）***
            actual_start_row = self._calculate_dynamic_table_start_row(
                worksheet, table_name, original_start_row, dataframe
            )

            logger.debug(f"DEBUG_TABLE_OBJ: 表格 {table_name} 原始起始行: {original_start_row}, 動態調整後起始行: {actual_start_row}")

            # 計算需要插入的行數
            data_rows_needed = len(dataframe)
            header_rows = 1 if obj_info.having_header else 0
            total_rows_needed = data_rows_needed + header_rows

            # *** 修正：如果起始位置有變化，需要調整template row複製位置 ***
            template_row = actual_start_row
            if total_rows_needed > 1:  # 只有在需要多行時才複製
                self._copy_template_rows(worksheet, template_row, total_rows_needed - 1)

            # 渲染數據到表格中
            self._render_dataframe_to_table(dataframe, worksheet, table, actual_start_row, start_col)

            # 更新表格範圍 - 使用DataFrame欄位數和原始表格欄位數中的較大值（包含公式欄）
            from openpyxl.utils import range_boundaries
            orig_min_col, orig_min_row, orig_max_col, orig_max_row = range_boundaries(table.ref)
            original_table_cols = orig_max_col - orig_min_col + 1
            dataframe_cols = len(dataframe.columns)

            # 使用DataFrame欄位數和原始表格欄位數中的較大值
            actual_cols = max(dataframe_cols, original_table_cols)

            # *** 修正：使用動態計算的起始位置 ***
            new_end_row = actual_start_row + len(dataframe) + (1 if obj_info.having_header else 0) - 1
            new_end_col = start_col + actual_cols - 1  # 使用實際需要的欄數
            new_start_cell = self._format_cell_reference(actual_start_row, start_col)
            new_range = f"{new_start_cell}:{self._format_cell_reference(new_end_row, new_end_col)}"

            # 使用統一方法更新表格範圍，確保 table.ref 和 autoFilter.ref 同步
            self._update_table_range_sync(table, new_range, table_name, dataframe, obj_info.having_header)

            # 更新物件資訊 - 使用實際欄位數和調整後的位置
            obj_info.data_shape.rows = len(dataframe) + (1 if obj_info.having_header else 0)
            obj_info.data_shape.cols = actual_cols  # 使用實際需要的欄數

            # *** 新增：計算位置偏移量並更新圖片物件 ***
            row_offset = actual_start_row - original_start_row
            if row_offset != 0:
                logger.debug(f"DEBUG_TABLE_OBJ: 檢測到位置偏移 {row_offset} 行，更新圖片物件位置")
                self._update_images_after_table_shift(worksheet, table_name, row_offset,
                                                     original_end_row, new_end_row)

            return obj_info.data_shape.rows

        except Exception as e:
            raise RenderError(
                f"渲染表格物件失敗: {str(e)}",
                tag_name=tag.tag_name,
                sheet_name=worksheet.title
            )
    
    def _render_dataframe_header(
        self, 
        dataframe: pd.DataFrame, 
        worksheet: Worksheet, 
        start_row: int, 
        start_col: int,
        template_cell_position: Optional[tuple] = None
    ) -> None:
        """
        渲染DataFrame表頭
        
        Args:
            dataframe: DataFrame物件
            worksheet: 工作表
            start_row: 開始行
            start_col: 開始列
            template_cell_position: 模板標籤所在位置 (row, col)，用於複製樣式
        """
        for col_idx, column_name in enumerate(dataframe.columns):
            cell = worksheet.cell(row=start_row, column=start_col + col_idx)
            
            # 檢查是否為合併儲存格，如果是則跳過設定值
            from openpyxl.cell.cell import MergedCell
            if not isinstance(cell, MergedCell):
                cell.value = column_name
            
            # 複製模板樣式到表頭
            if template_cell_position:
                template_cell = worksheet.cell(row=template_cell_position[0], column=template_cell_position[1])
                self._copy_cell_style(template_cell, cell)
    
    def _render_dataframe_data(
        self, 
        dataframe: pd.DataFrame, 
        worksheet: Worksheet, 
        start_row: int, 
        start_col: int,
        template_cell_position: Optional[tuple] = None
    ) -> int:
        """
        渲染DataFrame數據
        
        Args:
            dataframe: DataFrame物件
            worksheet: 工作表
            start_row: 開始行
            start_col: 開始列
            template_cell_position: 模板標籤所在位置 (row, col)，用於複製樣式
            
        Returns:
            int: 渲染的行數
        """
        # 首先檢查模板行（如果有）中是否有公式
        formula_columns = {}
        if template_cell_position:
            template_row = template_cell_position[0]
            # 檢查整個模板行的所有相關列是否有公式
            # 需要檢查比DataFrame更多的列，因為可能有計算欄位
            max_check_cols = max(len(dataframe.columns) + 3, 10)  # 至少檢查10列
            for col_idx in range(max_check_cols):
                actual_col = start_col + col_idx
                check_cell = worksheet.cell(row=template_row, column=actual_col)
                if check_cell.value and isinstance(check_cell.value, str) and check_cell.value.startswith('='):
                    formula_columns[col_idx] = check_cell.value
        
        # 先偵測模板的合併模式
        logger.debug(f"DEBUG_MERGE_DETECT: Starting merge detection for worksheet '{worksheet.title}'")
        logger.debug(f"DEBUG_MERGE_DETECT: Template row: {template_cell_position[0] if template_cell_position else 'None'}")
        logger.debug(f"DEBUG_MERGE_DETECT: Start col: {start_col}, DataFrame columns: {len(dataframe.columns)}")
        
        merge_pattern = self._detect_template_merge_pattern(
            worksheet, 
            template_cell_position[0] if template_cell_position else start_row,
            start_col,
            len(dataframe.columns)
        ) if template_cell_position else {}
        
        logger.debug(f"DEBUG_MERGE_DETECT: Detected merge pattern: {merge_pattern}")
        
        for row_idx, (_, row_data) in enumerate(dataframe.iterrows()):
            current_row = start_row + row_idx
            
            # 根據偵測到的模式創建合併儲存格
            if merge_pattern:
                for col_offset, span in merge_pattern.items():
                    merge_start_col = start_col + col_offset
                    merge_end_col = merge_start_col + span - 1
                    merge_range = f"{self._get_column_letter(merge_start_col)}{current_row}:{self._get_column_letter(merge_end_col)}{current_row}"
                    try:
                        worksheet.merge_cells(merge_range)
                        logger.debug(f"DEBUG_MERGE_CREATE: Created merge range {merge_range}")
                    except Exception as e:
                        logger.debug(f"DEBUG_MERGE_CREATE: Failed to create merge {merge_range}: {e}")
            
            # 處理DataFrame中的數據列 - 使用智能列偏移
            cumulative_offset = 0
            
            logger.debug(f"DEBUG_DATA_PLACE: Row {current_row} - Starting data placement")
            logger.debug(f"DEBUG_DATA_PLACE: Columns: {list(dataframe.columns)}")
            logger.debug(f"DEBUG_DATA_PLACE: Values: {list(row_data)}")
            
            for col_idx, value in enumerate(row_data):
                # 計算基礎目標列位置
                base_target_col = start_col + col_idx + cumulative_offset
                target_col = base_target_col
                
                logger.debug(f"DEBUG_DATA_PLACE: Col {col_idx} ('{dataframe.columns[col_idx]}'): value='{value}', base_target={base_target_col}, cumulative_offset={cumulative_offset}")
                
                # 檢查當前位置是否有合併儲存格
                merge_range = self._find_merged_range_for_cell(worksheet, current_row, target_col)
                
                if merge_range:
                    logger.debug(f"DEBUG_DATA_PLACE: Found merge at {target_col}: {merge_range}")
                    # 如果是合併範圍的起始儲存格
                    if merge_range.min_row == current_row and merge_range.min_col == target_col:
                        # 計算合併跨度並更新偏移
                        merge_span = merge_range.max_col - merge_range.min_col
                        if merge_span > 0:
                            logger.debug(f"DEBUG_DATA_PLACE: Is merge start cell, span={merge_span}")
                            cumulative_offset += merge_span
                    # 如果不是起始儲存格，需要跳過
                    elif self._is_cell_merged_and_not_top_left(worksheet, current_row, target_col):
                        # 找到下一個可用的列
                        old_col = target_col
                        target_col = self._find_next_available_column(worksheet, current_row, target_col)
                        additional_offset = target_col - old_col
                        cumulative_offset += additional_offset
                        logger.debug(f"DEBUG_DATA_PLACE: Not merge start, jumping from {old_col} to {target_col}")
                else:
                    logger.debug(f"DEBUG_DATA_PLACE: No merge at {target_col}")
                
                cell = worksheet.cell(row=current_row, column=target_col)
                
                # 檢查是否為合併儲存格，如果是則跳過
                from openpyxl.cell.cell import MergedCell
                if isinstance(cell, MergedCell):
                    logger.debug(f"DEBUG_DATA_PLACE: Cell at {target_col} is MergedCell, skipping")
                    continue
                
                logger.debug(f"DEBUG_DATA_PLACE: PLACING '{value}' at column {target_col} (cell {self._get_column_letter(target_col)}{current_row})")
                
                # 檢查該欄位是否有公式
                if col_idx in formula_columns:
                    # 對於有公式的欄位，複製公式並調整引用
                    original_formula = formula_columns[col_idx]
                    if template_cell_position:
                        adjusted_formula = self._adjust_formula_references(
                            original_formula,
                            template_cell_position[0],  # 模板行
                            start_col + col_idx,        # 公式列
                            current_row,                # 目標行
                            start_col + col_idx         # 目標列
                        )
                        cell.value = adjusted_formula
                        logger.debug(f"DEBUG_DATA_PLACE: Set formula at {self._get_column_letter(target_col)}{current_row}")
                else:
                    # 處理不同的數據類型
                    if pd.isna(value):
                        cell.value = None
                    elif isinstance(value, (datetime, date)):
                        cell.value = value
                    elif isinstance(value, (int, float, bool)):
                        cell.value = value
                    else:
                        cell.value = str(value)
                    logger.debug(f"DEBUG_DATA_PLACE: Successfully set value '{cell.value}' at {self._get_column_letter(target_col)}{current_row}")
                
                # 複製模板樣式到數據單元格
                if template_cell_position:
                    # 找到對應列的模板儲存格來複製樣式
                    template_col = cell.column  # 使用實際儲存格的列位置
                    template_cell = worksheet.cell(row=template_cell_position[0], column=template_col)
                    self._copy_cell_style(template_cell, cell)
            
            # 處理超出DataFrame範圍的公式列（如計算欄位）
            for col_idx in formula_columns:
                if col_idx >= len(dataframe.columns):  # 超出DataFrame範圍的列
                    cell = worksheet.cell(row=current_row, column=start_col + col_idx)
                    
                    # 檢查是否為合併儲存格，如果是則跳過
                    from openpyxl.cell.cell import MergedCell
                    if isinstance(cell, MergedCell):
                        continue
                    
                    # 複製公式並調整引用
                    original_formula = formula_columns[col_idx]
                    if template_cell_position:
                        adjusted_formula = self._adjust_formula_references(
                            original_formula,
                            template_cell_position[0],  # 模板行
                            start_col + col_idx,        # 公式列
                            current_row,                # 目標行
                            start_col + col_idx         # 目標列
                        )
                        cell.value = adjusted_formula
                    
                    # 複製模板樣式到數據單元格
                    if template_cell_position:
                        # 找到對應列的模板儲存格來複製樣式
                        template_col = start_col + col_idx
                        template_cell = worksheet.cell(row=template_cell_position[0], column=template_col)
                        self._copy_cell_style(template_cell, cell)
        
            # 為整個行範圍複製樣式，包括沒有數據或公式的列
            if template_cell_position:
                # 檢查模板行有多少列需要複製樣式（檢查到最多15列）
                max_template_cols = 15
                for col_idx in range(max_template_cols):
                    template_col = start_col + col_idx
                    template_cell = worksheet.cell(row=template_cell_position[0], column=template_col)
                    
                    # 如果模板儲存格有樣式（邊框、填充等），則複製到對應的數據行
                    if (template_cell.border and (template_cell.border.left.style or 
                                                 template_cell.border.right.style or 
                                                 template_cell.border.top.style or 
                                                 template_cell.border.bottom.style)) or \
                       (template_cell.fill and template_cell.fill.fill_type) or \
                       (template_cell.font and template_cell.font.name):
                        
                        target_cell = worksheet.cell(row=current_row, column=template_col)
                        
                        # 檢查是否為合併儲存格，如果是則跳過
                        from openpyxl.cell.cell import MergedCell
                        if not isinstance(target_cell, MergedCell):
                            self._copy_cell_style(template_cell, target_cell)
        
        return len(dataframe)
    
    def _render_dataframe_to_table(
        self, 
        dataframe: pd.DataFrame, 
        worksheet: Worksheet, 
        table, 
        start_row: int, 
        start_col: int
    ) -> None:
        """
        將DataFrame渲染到Excel表格物件中
        
        Args:
            dataframe: DataFrame物件
            worksheet: 工作表
            table: Excel表格物件
            start_row: 開始行
            start_col: 開始列
        """
        # 渲染表頭（如果表格有表頭）
        logger.debug(f"DEBUG_RENDER_TO_TABLE: 表格 {getattr(table, 'name', 'Unknown')} headerRowCount={table.headerRowCount}")
        if table.headerRowCount > 0:
            current_row = start_row
            for col_idx, column_name in enumerate(dataframe.columns):
                cell = worksheet.cell(row=current_row, column=start_col + col_idx)
                
                # 檢查是否為合併儲存格，如果是則跳過設定值
                from openpyxl.cell.cell import MergedCell
                if not isinstance(cell, MergedCell):
                    cell.value = column_name
            start_row += 1
        
        # 渲染數據
        self._render_dataframe_data(dataframe, worksheet, start_row, start_col)
        
        # 數據渲染完成後，檢查並更新表格範圍（確保 autoFilter.ref 同步）
        self._ensure_table_range_sync_after_data_render(table, dataframe, start_row, start_col)
    
    def _ensure_table_range_sync_after_data_render(self, table, dataframe, data_start_row: int, data_start_col: int):
        """
        在數據渲染完成後，確保表格範圍正確更新，並同步 autoFilter.ref

        Args:
            table: Excel表格物件
            dataframe: 已渲染的DataFrame
            data_start_row: 數據開始行（不包含表頭）
            data_start_col: 數據開始列
        """
        try:
            # 計算新的表格範圍
            header_rows = table.headerRowCount if hasattr(table, 'headerRowCount') else 0
            table_start_row = data_start_row - header_rows if header_rows > 0 else data_start_row
            table_end_row = data_start_row + len(dataframe) - 1
            table_end_col = data_start_col + len(dataframe.columns) - 1

            # 生成新範圍字串
            new_range = f"{self._format_cell_reference(table_start_row, data_start_col)}:{self._format_cell_reference(table_end_row, table_end_col)}"

            # 檢查是否需要更新
            current_ref = getattr(table, 'ref', '')
            if current_ref != new_range:
                logger.debug(f"DEBUG_DATA_RENDER: 偵測到表格範圍需要更新 - 當前: {current_ref} -> 應為: {new_range}")
                # 使用統一方法更新表格範圍
                table_name = getattr(table, 'name', 'Unknown')
                # *** 修正：使用表格本身的 headerRowCount 來判斷 include_header ***
                include_header = header_rows > 0
                logger.debug(f"DEBUG_DATA_RENDER: 使用表格 headerRowCount={header_rows} 決定 include_header={include_header}")
                self._update_table_range_sync(table, new_range, f"[數據渲染後]{table_name}", dataframe, include_header)
            else:
                logger.debug(f"DEBUG_DATA_RENDER: 表格範圍已經正確，無需更新: {current_ref}")

        except Exception as e:
            logger.debug(f"DEBUG_DATA_RENDER: 檢查表格範圍時發生錯誤: {str(e)}")
    
    def _copy_template_rows(self, worksheet: Worksheet, template_row: int, num_copies: int) -> None:
        """
        複製模板行（實現Excel的選取整行→插入複製行的操作）
        
        這個方法模擬Excel中的操作流程：
        1. 選取模板行的entire row（包含風格樣式、公式、合併儲存格）
        2. 在模板行下方插入指定數量的新行
        3. 將模板行的所有內容（樣式、公式、合併）複製到新插入的行
        4. 調整公式引用，保持相對引用正確
        5. 讓原有的其他內容往下推移
        
        Args:
            worksheet: 工作表
            template_row: 模板行號（包含標籤的行）
            num_copies: 需要複製的行數（通常是 DataFrame.shape[0] - 1）
        """
        if num_copies <= 0:
            return
            
        try:
            logger.debug(f"DEBUG: 開始複製模板行 {template_row}，需要複製 {num_copies} 行")
            
            # Step 1: 記錄模板行的所有資訊（在插入行之前）
            template_data = self._capture_template_row_data(worksheet, template_row)
            
            # Step 2: 在模板行下方插入新行（一次性插入所有需要的行）
            # 這會讓原有的其他內容往下推移
            worksheet.insert_rows(template_row + 1, num_copies)
            logger.debug(f"DEBUG: 已插入 {num_copies} 行在第 {template_row + 1} 行")
            
            # Step 3: 將模板行的內容複製到新插入的每一行
            for copy_index in range(num_copies):
                target_row = template_row + 1 + copy_index
                self._copy_template_row_to_target(worksheet, template_data, template_row, target_row)
                logger.debug(f"DEBUG: 已複製模板行到第 {target_row} 行")
                
        except Exception as e:
            logger.error(f"ERROR: 複製模板行失敗: {str(e)}")
            raise RenderError(f"複製模板行失敗: {str(e)}")
    
    def _capture_template_row_data(self, worksheet: Worksheet, template_row: int) -> dict:
        """
        捕獲模板行的所有資訊（樣式、值、公式、合併儲存格等）
        
        Args:
            worksheet: 工作表
            template_row: 模板行號
            
        Returns:
            dict: 包含模板行所有資訊的字典
        """
        template_data = {
            'cells': {},
            'merged_ranges': [],
            'max_col': worksheet.max_column
        }
        
        # 捕獲每個儲存格的資訊
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=template_row, column=col)
            
            # 記錄儲存格資訊
            template_data['cells'][col] = {
                'value': cell.value,
                'data_type': cell.data_type,
                'style': self._copy_cell_style_dict(cell),
                'is_merged': False,
                'merged_range': None
            }
        
        # 捕獲合併儲存格範圍
        for merged_range in worksheet.merged_cells.ranges:
            if merged_range.min_row <= template_row <= merged_range.max_row:
                template_data['merged_ranges'].append({
                    'min_row': merged_range.min_row,
                    'max_row': merged_range.max_row,
                    'min_col': merged_range.min_col,
                    'max_col': merged_range.max_col,
                    'range_string': str(merged_range)
                })
                
                # 標記合併範圍內的儲存格
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if col in template_data['cells']:
                        template_data['cells'][col]['is_merged'] = True
                        template_data['cells'][col]['merged_range'] = merged_range
        
        return template_data
    
    def _copy_template_row_to_target(self, worksheet: Worksheet, template_data: dict, template_row: int, target_row: int) -> None:
        """
        將模板行資料複製到目標行
        
        Args:
            worksheet: 工作表
            template_data: 模板行資料
            template_row: 模板行號
            target_row: 目標行號
        """
        from openpyxl.cell.cell import MergedCell
        
        # 複製每個儲存格
        for col, cell_data in template_data['cells'].items():
            target_cell = worksheet.cell(row=target_row, column=col)
            
            # 跳過合併儲存格（這些會在後續處理合併範圍時處理）
            if isinstance(target_cell, MergedCell):
                continue
            
            # 複製樣式
            self._apply_cell_style_dict(target_cell, cell_data['style'])
            
            # 複製值和公式
            if cell_data['value'] is not None:
                # 檢查是否為標籤字串（包含{{和}}）
                is_tag_cell = (isinstance(cell_data['value'], str) and 
                             '{{' in cell_data['value'] and '}}' in cell_data['value'])
                
                if not is_tag_cell:
                    # 處理公式
                    if cell_data['data_type'] == 'f' and isinstance(cell_data['value'], str):
                        # 調整公式中的引用
                        adjusted_formula = self._adjust_formula_references(
                            cell_data['value'],
                            template_row,  # 原始行
                            col,           # 原始列
                            target_row,    # 目標行
                            col            # 目標列
                        )
                        target_cell.value = adjusted_formula
                    else:
                        # 普通值（但不是標籤）
                        target_cell.value = cell_data['value']
                # 如果是標籤儲存格，只複製樣式，不複製值
        
        # 處理合併儲存格
        self._copy_merged_ranges_to_target_row(worksheet, template_data['merged_ranges'], template_row, target_row)
    
    def _copy_merged_ranges_to_target_row(self, worksheet: Worksheet, merged_ranges: list, template_row: int, target_row: int) -> None:
        """
        將模板行的合併範圍複製到目標行
        只複製確實屬於模板行本身的合併儲存格，不複製其他行的合併儲存格
        
        Args:
            worksheet: 工作表
            merged_ranges: 合併範圍列表
            template_row: 模板行號
            target_row: 目標行號
        """
        # 檢查模板行是否包含標籤
        template_has_tag = False
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=template_row, column=col)
            if cell.value and isinstance(cell.value, str):
                if '{{' in cell.value and '}}' in cell.value:
                    template_has_tag = True
                    break
        
        if template_has_tag:
            logger.debug(f"DEBUG: renderer - 模板行 {template_row} 包含標籤，不複製合併儲存格到數據行")
            # 標籤行不應該複製合併儲存格給數據行
            return
        
        for merge_info in merged_ranges:
            # 只處理單行合併範圍（確實屬於模板行本身的）
            if merge_info['min_row'] == merge_info['max_row'] == template_row:
                # 計算行偏移
                row_offset = target_row - template_row
                
                # 創建新的合併範圍
                new_min_row = merge_info['min_row'] + row_offset
                new_max_row = merge_info['max_row'] + row_offset
                new_min_col = merge_info['min_col']
                new_max_col = merge_info['max_col']
                
                # 確保新範圍有效且不與現有合併範圍衝突
                if new_min_row > 0 and new_max_row > 0:
                    new_range_string = f"{self._get_column_letter(new_min_col)}{new_min_row}:{self._get_column_letter(new_max_col)}{new_max_row}"
                    
                    try:
                        # 檢查是否已存在重疊的合併範圍
                        existing_merge = self._find_merged_range_for_cell(worksheet, new_min_row, new_min_col)
                        if not existing_merge:
                            worksheet.merge_cells(new_range_string)
                            logger.debug(f"DEBUG: renderer創建合併儲存格範圍: {new_range_string}")
                    except Exception as e:
                        logger.debug(f"DEBUG: 合併儲存格失敗 {new_range_string}: {e}")
                        # 繼續處理，不中斷整個流程
    
    def _copy_cell_style_dict(self, cell) -> dict:
        """
        將儲存格樣式複製到字典中
        
        Args:
            cell: 來源儲存格
            
        Returns:
            dict: 樣式資訊字典
        """
        return {
            'font': cell.font,
            'fill': cell.fill,
            'border': cell.border,
            'alignment': cell.alignment,
            'number_format': cell.number_format,
            'protection': cell.protection
        }
    
    def _apply_cell_style_dict(self, target_cell, style_dict: dict) -> None:
        """
        將樣式字典套用到目標儲存格
        
        Args:
            target_cell: 目標儲存格
            style_dict: 樣式資訊字典
        """
        try:
            target_cell.font = style_dict['font']
            target_cell.fill = style_dict['fill']
            target_cell.border = style_dict['border']
            target_cell.alignment = style_dict['alignment']
            target_cell.number_format = style_dict['number_format']
            target_cell.protection = style_dict['protection']
        except Exception as e:
            logger.debug(f"DEBUG: 套用樣式失敗: {e}")
            # 繼續處理，不中斷流程

    def copy_styles_and_formulas(self, source_range: str, target_range: str, worksheet: Worksheet) -> None:
        """
        複製樣式和公式
        
        Args:
            source_range: 來源範圍
            target_range: 目標範圍  
            worksheet: 工作表
            
        Raises:
            RenderError: 複製過程錯誤
        """
        try:
            # 解析範圍
            source_cells = worksheet[source_range]
            target_cells = worksheet[target_range]
            
            # 如果是單一儲存格
            if not isinstance(source_cells, tuple):
                source_cells = ((source_cells,),)
                target_cells = ((target_cells,),)
            
            # 複製樣式和公式
            for src_row, tgt_row in zip(source_cells, target_cells):
                for src_cell, tgt_cell in zip(src_row, tgt_row):
                    if src_cell.value:
                        # 複製樣式
                        tgt_cell._style = src_cell._style
                        
                        # 複製公式（如果有）
                        if isinstance(src_cell.value, str) and src_cell.value.startswith('='):
                            # 這裡需要調整公式中的相對引用
                            tgt_cell.value = self._adjust_formula_references(
                                src_cell.value, 
                                src_cell.row, 
                                src_cell.column,
                                tgt_cell.row, 
                                tgt_cell.column
                            )
                        
        except Exception as e:
            raise RenderError(f"複製樣式和公式失敗: {str(e)}")
    
    def _adjust_formula_references(
        self, 
        formula: str, 
        src_row: int, 
        src_col: int, 
        tgt_row: int, 
        tgt_col: int
    ) -> str:
        """
        調整公式中的相對引用
        
        Args:
            formula: 原始公式
            src_row: 來源行
            src_col: 來源列
            tgt_row: 目標行
            tgt_col: 目標列
            
        Returns:
            str: 調整後的公式
        """
        import re
        from openpyxl.utils import get_column_letter, column_index_from_string
        
        # 計算行的偏移量
        row_offset = tgt_row - src_row
        
        # 正則表達式匹配儲存格引用（如 A1, B3, C10等）
        cell_ref_pattern = r'([A-Z]+)(\d+)'
        
        def replace_cell_ref(match):
            col_letter = match.group(1)
            row_num = int(match.group(2))
            
            # 調整行號
            new_row = row_num + row_offset
            
            # 保持列不變（因為我們假設公式是相對於同一行的）
            return f"{col_letter}{new_row}"
        
        # 替換公式中的所有儲存格引用
        adjusted_formula = re.sub(cell_ref_pattern, replace_cell_ref, formula)
        
        return adjusted_formula
    
    def _parse_cell_reference(self, cell_ref: str) -> tuple:
        """
        解析儲存格參考
        
        Args:
            cell_ref: 儲存格參考字串
            
        Returns:
            tuple: (row, col) 座標
        """
        from openpyxl.utils import coordinate_to_tuple
        return coordinate_to_tuple(cell_ref)
    
    def _format_cell_reference(self, row: int, col: int) -> str:
        """
        格式化儲存格參考
        
        Args:
            row: 行號
            col: 列號
            
        Returns:
            str: 儲存格參考字串
        """
        from openpyxl.utils import get_column_letter
        return f"{get_column_letter(col)}{row}"
    
    def _copy_cell_style(self, source_cell, target_cell) -> None:
        """
        複製單元格樣式
        
        Args:
            source_cell: 來源單元格
            target_cell: 目標單元格
        """
        try:
            from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
            
            # 複製字體樣式
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    color=source_cell.font.color
                )
            
            # 複製邊框樣式 - 精確複製源儲存格的邊框設定
            if source_cell.border:
                target_cell.border = Border(
                    left=Side(
                        style=source_cell.border.left.style if source_cell.border.left else None,
                        color=source_cell.border.left.color if source_cell.border.left else None
                    ),
                    right=Side(
                        style=source_cell.border.right.style if source_cell.border.right else None,
                        color=source_cell.border.right.color if source_cell.border.right else None
                    ),
                    top=Side(
                        style=source_cell.border.top.style if source_cell.border.top else None,
                        color=source_cell.border.top.color if source_cell.border.top else None
                    ),
                    bottom=Side(
                        style=source_cell.border.bottom.style if source_cell.border.bottom else None,
                        color=source_cell.border.bottom.color if source_cell.border.bottom else None
                    )
                )
            
            # 複製填充樣式
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color
                )
            
            # 複製對齊樣式
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text
                )
            
            # 複製數字格式
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
                
        except Exception as e:
            # 如果樣式複製失敗，不中斷渲染流程，但記錄詳細錯誤資訊
            logger.debug(f"DEBUG: 嘗試設定儲存格樣式時發生錯誤: {e}")
            # 嘗試安全地複製樣式，過濾掉問題的顏色值
            try:
                self._safe_copy_cell_style(source_cell, target_cell)
            except Exception as safe_e:
                logger.debug(f"DEBUG: 安全樣式複製也失敗: {safe_e}")

    def _safe_copy_cell_style(self, source_cell, target_cell) -> None:
        """
        安全地複製單元格樣式，過濾掉問題的顏色值
        
        Args:
            source_cell: 來源單元格
            target_cell: 目標單元格
        """
        try:
            from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
            from openpyxl.styles.colors import Color
            
            # 安全地複製字體樣式
            if source_cell.font:
                font_color = None
                if source_cell.font.color:
                    try:
                        # 驗證顏色值是否有效
                        if hasattr(source_cell.font.color, 'rgb') and source_cell.font.color.rgb:
                            color_value = source_cell.font.color.rgb
                            if color_value and len(color_value) == 8 and all(c in '0123456789ABCDEFabcdef' for c in color_value):
                                font_color = source_cell.font.color
                    except:
                        pass  # 忽略顏色錯誤，使用預設
                
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    color=font_color
                )
            
            # 安全地複製邊框樣式
            if source_cell.border:
                def safe_border_color(border_side):
                    if border_side and border_side.color:
                        try:
                            if hasattr(border_side.color, 'rgb') and border_side.color.rgb:
                                color_value = border_side.color.rgb
                                if color_value and len(color_value) == 8 and all(c in '0123456789ABCDEFabcdef' for c in color_value):
                                    return border_side.color
                        except:
                            pass
                    return None
                
                target_cell.border = Border(
                    left=Side(
                        style=source_cell.border.left.style if source_cell.border.left else None,
                        color=safe_border_color(source_cell.border.left)
                    ),
                    right=Side(
                        style=source_cell.border.right.style if source_cell.border.right else None,
                        color=safe_border_color(source_cell.border.right)
                    ),
                    top=Side(
                        style=source_cell.border.top.style if source_cell.border.top else None,
                        color=safe_border_color(source_cell.border.top)
                    ),
                    bottom=Side(
                        style=source_cell.border.bottom.style if source_cell.border.bottom else None,
                        color=safe_border_color(source_cell.border.bottom)
                    )
                )
            
            # 安全地複製填充樣式
            if source_cell.fill and hasattr(source_cell.fill, 'fill_type'):
                safe_start_color = None
                safe_end_color = None
                
                try:
                    if source_cell.fill.start_color and hasattr(source_cell.fill.start_color, 'rgb'):
                        color_value = source_cell.fill.start_color.rgb
                        if color_value and len(color_value) == 8 and all(c in '0123456789ABCDEFabcdef' for c in color_value):
                            safe_start_color = source_cell.fill.start_color
                except:
                    pass
                
                try:
                    if source_cell.fill.end_color and hasattr(source_cell.fill.end_color, 'rgb'):
                        color_value = source_cell.fill.end_color.rgb
                        if color_value and len(color_value) == 8 and all(c in '0123456789ABCDEFabcdef' for c in color_value):
                            safe_end_color = source_cell.fill.end_color
                except:
                    pass
                
                if safe_start_color or safe_end_color:
                    target_cell.fill = PatternFill(
                        fill_type=source_cell.fill.fill_type,
                        start_color=safe_start_color,
                        end_color=safe_end_color
                    )
            
            # 複製對齊樣式（這通常不會有問題）
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text
                )
            
            # 複製數字格式（這通常不會有問題）
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
                
        except Exception as e:
            logger.debug(f"DEBUG: 安全樣式複製失敗: {e}")
            # 最後的回退方案：只複製最基本的樣式
            try:
                if source_cell.font:
                    target_cell.font = Font(
                        name=source_cell.font.name or 'Calibri',
                        size=source_cell.font.size or 11,
                        bold=source_cell.font.bold or False,
                        italic=source_cell.font.italic or False
                    )
                if source_cell.alignment:
                    target_cell.alignment = Alignment(
                        horizontal=source_cell.alignment.horizontal,
                        vertical=source_cell.alignment.vertical,
                        wrap_text=source_cell.alignment.wrap_text
                    )
            except Exception as final_e:
                logger.debug(f"DEBUG: 最終樣式複製回退也失敗: {final_e}")

    def _find_merged_range_for_cell(self, worksheet: Worksheet, row: int, col: int):
        """
        查找指定儲存格所在的合併範圍
        
        Args:
            worksheet: 工作表
            row: 行號
            col: 列號
            
        Returns:
            合併範圍對象，如果不在合併範圍內則返回None
        """
        for merge_range in worksheet.merged_cells.ranges:
            if (merge_range.min_row <= row <= merge_range.max_row and 
                merge_range.min_col <= col <= merge_range.max_col):
                return merge_range
        return None

    def _detect_template_merge_pattern(self, worksheet: Worksheet, template_row: int, start_col: int, num_cols: int) -> dict:
        """
        偵測模板的合併模式，返回每個欄位應該使用的偏移量
        
        Args:
            worksheet: 工作表
            template_row: 模板行號
            start_col: 開始列號
            num_cols: DataFrame的列數
            
        Returns:
            dict: 包含merge_info的字典，記錄哪些欄位需要合併及跨度
        """
        merge_info = {}
        
        # 檢查模板行上方的標題行是否有合併模式
        for row_offset in [-2, -1, 0]:
            check_row = template_row + row_offset
            if check_row <= 0:
                continue
                
            # 檢查每個可能的合併範圍
            checked_ranges = set()
            for col_idx in range(num_cols + 5):  # 檢查更多列以確保完整性
                actual_col = start_col + col_idx
                merge_range = self._find_merged_range_for_cell(worksheet, check_row, actual_col)
                
                if merge_range and merge_range not in checked_ranges:
                    checked_ranges.add(merge_range)
                    # 如果這是合併範圍的起始列
                    if merge_range.min_col == actual_col:
                        # 記錄合併信息
                        col_offset = actual_col - start_col
                        span = merge_range.max_col - merge_range.min_col + 1
                        if col_offset >= 0 and col_offset < num_cols:
                            merge_info[col_offset] = span
                            logger.debug(f"DEBUG: 偵測到合併模式 - 列{col_offset}跨度{span}列")
        
        return merge_info

    def _is_cell_merged_and_not_top_left(self, worksheet: Worksheet, row: int, col: int) -> bool:
        """
        檢查指定儲存格是否是合併儲存格的非主要部分（即不是合併範圍的左上角）
        
        Args:
            worksheet: 工作表
            row: 行號
            col: 列號
            
        Returns:
            bool: 如果是合併儲存格的非主要部分則返回True
        """
        try:
            cell = worksheet.cell(row=row, column=col)
            
            # 如果儲存格是MergedCell類型，表示它是合併範圍的一部分但不是主要儲存格
            from openpyxl.cell.cell import MergedCell
            if isinstance(cell, MergedCell):
                return True
                
            # 檢查是否是合併範圍的主要儲存格（左上角）
            for merge_range in worksheet.merged_cells.ranges:
                if (merge_range.min_row == row and merge_range.min_col == col):
                    # 這是合併範圍的主要儲存格（左上角），可以使用
                    return False
                elif (merge_range.min_row <= row <= merge_range.max_row and 
                      merge_range.min_col <= col <= merge_range.max_col):
                    # 這是合併範圍的非主要部分，不能使用
                    return True
            
            return False
        except:
            return False

    def _find_next_available_column(self, worksheet: Worksheet, row: int, start_col: int, max_attempts: int = 20) -> int:
        """
        從指定列開始，找到下一個可用的（未被合併占用的）列
        
        Args:
            worksheet: 工作表
            row: 行號
            start_col: 開始搜索的列號
            max_attempts: 最大嘗試次數
            
        Returns:
            int: 可用的列號
        """
        for offset in range(max_attempts):
            test_col = start_col + offset
            if not self._is_cell_merged_and_not_top_left(worksheet, row, test_col):
                return test_col
        
        # 如果找不到可用列，返回原始列號（容錯處理）
        return start_col

    def _copy_merged_cells_for_data_row(self, worksheet: Worksheet, template_row: int, target_row: int, start_col: int, num_cols: int) -> None:
        """
        為數據行複製模板行的合併儲存格設定
        
        Args:
            worksheet: 工作表
            template_row: 模板行號
            target_row: 目標行號
            start_col: 開始列號
            num_cols: 列數
        """
        try:
            # 通用邏輯：查找模板行中的所有合併範圍
            template_merges = []
            for col_idx in range(num_cols + 5):  # 檢查更多列以包含可能的計算欄位
                actual_col = start_col + col_idx
                merge_range = self._find_merged_range_for_cell(worksheet, template_row, actual_col)
                if merge_range and merge_range not in template_merges:
                    template_merges.append(merge_range)
            
            # 為目標行創建相應的合併範圍
            for merge_range in template_merges:
                # 計算相對於模板行的偏移
                row_offset = target_row - template_row
                
                # 創建新的合併範圍
                new_min_row = merge_range.min_row + row_offset
                new_max_row = merge_range.max_row + row_offset
                new_min_col = merge_range.min_col
                new_max_col = merge_range.max_col
                
                # 確保新範圍在合理範圍內
                if new_min_row > 0 and new_max_row > 0:
                    new_range = f"{self._get_column_letter(new_min_col)}{new_min_row}:{self._get_column_letter(new_max_col)}{new_max_row}"
                    
                    # 檢查是否已經存在重疊的合併範圍
                    existing_merge = self._find_merged_range_for_cell(worksheet, new_min_row, new_min_col)
                    if not existing_merge:
                        worksheet.merge_cells(new_range)
                        
                        # 合併後重新設定正確的邊框樣式
                        for row in range(new_min_row, new_max_row + 1):
                            for col in range(new_min_col, new_max_col + 1):
                                template_cell = worksheet.cell(row=template_row + (row - new_min_row), column=col)
                                target_cell = worksheet.cell(row=row, column=col)
                                # 重新複製邊框樣式
                                self._copy_cell_style(template_cell, target_cell)
                        
        except Exception as e:
            # 如果合併失敗，不中斷渲染流程
            logger.debug(f"DEBUG: 合併儲存格複製失敗: {e}")
            pass

    def _get_column_letter(self, col_num: int) -> str:
        """
        將列號轉換為列字母
        
        Args:
            col_num: 列號（1開始）
            
        Returns:
            列字母（如A, B, C...）
        """
        from openpyxl.utils import get_column_letter
        return get_column_letter(col_num)
    
    def _is_range_merged(self, worksheet: Worksheet, range_str: str) -> bool:
        """
        檢查指定範圍是否已經合併
        
        Args:
            worksheet: 工作表
            range_str: 範圍字符串（如"A1:B1"）
            
        Returns:
            bool: 如果範圍已合併返回True
        """
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        
        for merge_range in worksheet.merged_cells.ranges:
            if (merge_range.min_row == min_row and merge_range.max_row == max_row and
                merge_range.min_col == min_col and merge_range.max_col == max_col):
                return True
        return False
    
    def _update_table_object_range(self, worksheet, obj_info, dataframe, tag, start_row, start_col, total_rows):
        """
        更新表格物件的範圍，確保 table.ref 和 autoFilter.ref 保持一致
        
        Args:
            worksheet: Excel工作表
            obj_info: 物件資訊
            dataframe: DataFrame數據
            tag: 標籤物件
            start_row: 起始行
            start_col: 起始列
            total_rows: 總行數
        """
        try:
            # 找到對應的表格物件
            table_name = obj_info.display_name
            logger.debug(f"DEBUG_TABLE_OBJ: 更新表格物件 {table_name} 的範圍")

            if table_name not in worksheet.tables:
                logger.debug(f"DEBUG_TABLE_OBJ: 直接匹配失敗，嘗試位置匹配: {table_name}")
                # 嘗試根據位置匹配表格
                matched_table = self._find_table_by_position(worksheet, start_row, start_col)
                if matched_table:
                    table_name = matched_table
                    logger.debug(f"DEBUG_TABLE_OBJ: 位置匹配成功: {table_name}")
                else:
                    logger.debug(f"DEBUG_TABLE_OBJ: 警告 - 找不到表格物件: {table_name}")
                    return

            table = worksheet.tables[table_name]

            # *** 新增：計算動態起始位置（gap區域計算）***
            original_start_row = start_row
            actual_start_row = self._calculate_dynamic_table_start_row(
                worksheet, table_name, original_start_row, dataframe
            )

            logger.debug(f"DEBUG_TABLE_OBJ: 表格 {table_name} 原始起始行: {original_start_row}, 動態調整後起始行: {actual_start_row}")

            # 如果位置有變化，需要使用調整後的起始行
            if actual_start_row != original_start_row:
                start_row = actual_start_row
                logger.debug(f"DEBUG_TABLE_OBJ: 使用動態調整後的起始行: {start_row}")
            
            # 取得表格原始範圍
            from openpyxl.utils import range_boundaries
            orig_min_col, orig_min_row, orig_max_col, orig_max_row = range_boundaries(table.ref)
            original_table_cols = orig_max_col - orig_min_col + 1
            dataframe_cols = len(dataframe.columns)
            
            # 使用DataFrame欄位數和原始表格欄位數中的較大值
            actual_cols = max(dataframe_cols, original_table_cols)
            
            # 計算新的範圍
            new_end_row = start_row + total_rows - 1
            new_end_col = start_col + actual_cols - 1
            start_cell = self._format_cell_reference(start_row, start_col)
            end_cell = self._format_cell_reference(new_end_row, new_end_col)
            new_range = f"{start_cell}:{end_cell}"

            # 檢查是否有noheader條件
            skip_header = tag.has_condition and tag.condition == "noheader"
            include_header = not skip_header

            logger.debug(f"DEBUG_TABLE_OBJ: 表格 {table_name} - skip_header={skip_header}, include_header={include_header}")
            logger.debug(f"DEBUG_TABLE_OBJ: tag.has_condition={tag.has_condition}, tag.condition='{tag.condition if tag.has_condition else 'None'}'")

            # 使用統一方法更新表格範圍，確保 table.ref 和 autoFilter.ref 同步
            self._update_table_range_sync(table, new_range, table_name, dataframe, include_header)
            
            # 立即驗證更新是否成功
            if hasattr(table, 'autoFilter') and table.autoFilter:
                current_ref = table.autoFilter.ref
                logger.debug(f"DEBUG_TABLE_OBJ: 驗證 - 當前 autoFilter.ref: {current_ref}")
                if current_ref != new_range:
                    logger.debug(f"DEBUG_TABLE_OBJ: 警告！更新後驗證失敗，期望: {new_range}, 實際: {current_ref}")
            else:
                logger.debug(f"DEBUG_TABLE_OBJ: 表格 {table_name} 沒有 autoFilter")
            
        except Exception as e:
            logger.debug(f"DEBUG_TABLE_OBJ: 更新表格物件範圍失敗: {str(e)}")
            import traceback
            logger.debug("例外堆疊", exc_info=True)

    def _calculate_dynamic_table_start_row(
        self,
        worksheet: Worksheet,
        current_table_name: str,
        original_start_row: int,
        dataframe: pd.DataFrame
    ) -> int:
        """
        計算表格物件的動態起始位置（考慮gap區域）

        當同一sheet中有多個表格物件時，需要根據以下規則動態調整：
        1. 找到上一個順序渲染的表格物件
        2. 計算gap區域：上一個表格的max row到當前表格min row的差距
        3. 根據上一個表格的實際渲染結果重新計算當前表格的起始位置

        Args:
            worksheet: 工作表
            current_table_name: 當前表格名稱
            original_start_row: 原始起始行
            dataframe: 當前表格的數據

        Returns:
            int: 動態調整後的起始行
        """
        try:
            logger.debug(f"DEBUG_GAP: 開始計算表格 {current_table_name} 的動態起始位置")

            # 獲取當前表格的列位置以判斷是否在同一個垂直列
            from openpyxl.utils import range_boundaries
            current_table = worksheet.tables[current_table_name]
            current_min_col, current_min_row, current_max_col, current_max_row = range_boundaries(current_table.ref)

            logger.debug(f"DEBUG_GAP: 當前表格 {current_table_name} 列範圍: {current_min_col}-{current_max_col}")

            # 收集同一垂直列的表格（只考慮垂直方向推移）
            same_column_tables = []

            for table_name, table in worksheet.tables.items():
                if hasattr(table, 'ref') and table.ref:
                    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

                    # 檢查是否在同一垂直列（列範圍有重疊）
                    has_column_overlap = not (max_col < current_min_col or min_col > current_max_col)

                    if has_column_overlap:
                        table_info = {
                            'name': table_name,
                            'min_row': min_row,
                            'max_row': max_row,
                            'min_col': min_col,
                            'max_col': max_col,
                            'table': table
                        }
                        same_column_tables.append(table_info)
                        logger.debug(f"DEBUG_GAP: 同列表格 {table_name} - 行範圍:{min_row}-{max_row}, 列範圍:{min_col}-{max_col}")

                    # 保存原始位置信息
                    if not hasattr(self, '_original_table_positions'):
                        self._original_table_positions = {}
                    if table_name not in self._original_table_positions:
                        self._original_table_positions[table_name] = {
                            'original_min_row': min_row,
                            'original_max_row': max_row
                        }

            # 按起始行排序（只考慮同一垂直列的表格）
            same_column_tables.sort(key=lambda x: x['min_row'])
            logger.debug(f"DEBUG_GAP: 在同一垂直列發現 {len(same_column_tables)} 個表格物件")

            # 找到當前表格在同列表格中的位置
            current_table_index = -1
            for i, table_info in enumerate(same_column_tables):
                if table_info['name'] == current_table_name:
                    current_table_index = i
                    break

            if current_table_index == -1:
                logger.debug(f"DEBUG_GAP: 找不到當前表格 {current_table_name} 在同列表格中，使用原始位置")
                return original_start_row

            # 如果是同列中的第一個表格，直接使用原始位置
            if current_table_index == 0:
                logger.debug(f"DEBUG_GAP: 當前表格是同列中的第一個表格，使用原始位置 {original_start_row}")
                return original_start_row

            # 計算gap區域 - 只考慮同列中的上一個表格
            previous_table_info = same_column_tables[current_table_index - 1]
            current_table_info = same_column_tables[current_table_index]

            # 獲取上一個表格的原始位置信息
            previous_table_name = previous_table_info['name']
            if previous_table_name in self._original_table_positions:
                original_previous_max = self._original_table_positions[previous_table_name]['original_max_row']
            else:
                original_previous_max = previous_table_info['max_row']

            # 獲取當前表格的原始位置信息
            if current_table_name in self._original_table_positions:
                original_current_min = self._original_table_positions[current_table_name]['original_min_row']
            else:
                original_current_min = current_table_info['min_row']

            # 計算原始gap大小（基於模板中的設計）
            original_gap_size = original_current_min - original_previous_max

            # 上一個表格的當前實際max row（可能已經被渲染擴展）
            previous_table_actual_max_row = previous_table_info['max_row']

            # 新的起始位置 = 上一個表格的實際結束位置 + 原始gap區域大小
            new_start_row = previous_table_actual_max_row + original_gap_size

            logger.debug(f"DEBUG_GAP: 同列上一個表格 {previous_table_name}:")
            logger.debug(f"DEBUG_GAP:   原始max_row: {original_previous_max}")
            logger.debug(f"DEBUG_GAP:   實際max_row: {previous_table_actual_max_row}")
            logger.debug(f"DEBUG_GAP: 當前表格 {current_table_name}:")
            logger.debug(f"DEBUG_GAP:   原始min_row: {original_current_min}")
            logger.debug(f"DEBUG_GAP:   原始gap大小: {original_gap_size}")
            logger.debug(f"DEBUG_GAP: 動態調整後起始行: {original_start_row} -> {new_start_row}")

            return max(new_start_row, original_start_row)  # 確保不會往上移動

        except Exception as e:
            logger.debug(f"DEBUG_GAP: 計算動態起始位置失敗: {str(e)}")
            import traceback
            logger.debug("例外堆疊", exc_info=True)
            return original_start_row

    def _update_images_after_table_shift(
        self,
        worksheet: Worksheet,
        table_name: str,
        row_offset: int,
        original_end_row: int,
        new_end_row: int
    ) -> None:
        """
        在表格位置變化後更新圖片物件位置

        Args:
            worksheet: 工作表
            table_name: 表格名稱
            row_offset: 行偏移量
            original_end_row: 原始結束行
            new_end_row: 新結束行
        """
        try:
            logger.debug(f"DEBUG_IMG_SHIFT: 開始更新圖片位置，偏移量: {row_offset}")

            # 檢查是否有圖片需要更新
            if not hasattr(worksheet, '_images') or not worksheet._images:
                logger.debug(f"DEBUG_IMG_SHIFT: 工作表沒有圖片物件")
                return

            for image in worksheet._images:
                if hasattr(image, 'anchor'):
                    anchor = image.anchor

                    # 處理TwoCellAnchor類型
                    if hasattr(anchor, '_from') and hasattr(anchor, 'to'):
                        # 檢查圖片是否在受影響的區域
                        from_row = getattr(anchor._from, 'row', 0)
                        to_row = getattr(anchor.to, 'row', 0)

                        # 如果圖片位置在表格變化範圍內，需要調整
                        if from_row > original_end_row:
                            # 圖片在原始表格下方，需要推移
                            self._adjust_image_anchor_position(anchor, row_offset)
                            logger.debug(f"DEBUG_IMG_SHIFT: 調整了位置在表格下方的圖片")

                    # 處理OneCellAnchor類型
                    elif hasattr(anchor, '_from'):
                        from_row = getattr(anchor._from, 'row', 0)
                        if from_row > original_end_row:
                            # 只調整_from位置
                            anchor._from.row = from_row + row_offset
                            logger.debug(f"DEBUG_IMG_SHIFT: 調整了OneCellAnchor圖片位置")

        except Exception as e:
            logger.debug(f"DEBUG_IMG_SHIFT: 更新圖片位置失敗: {str(e)}")

    def _adjust_image_anchor_position(self, anchor, row_offset: int) -> None:
        """
        調整圖片錨點位置

        Args:
            anchor: 圖片錨點
            row_offset: 行偏移量
        """
        try:
            # 保存原始偏移值
            original_from_rowOff = getattr(anchor._from, 'rowOff', 0)
            original_from_colOff = getattr(anchor._from, 'colOff', 0)
            original_to_rowOff = getattr(anchor.to, 'rowOff', 0)
            original_to_colOff = getattr(anchor.to, 'colOff', 0)

            # 更新_from位置
            anchor._from.row = anchor._from.row + row_offset

            # 更新to位置
            anchor.to.row = anchor.to.row + row_offset

            # 恢復原始偏移值（重要：防止圖片變形）
            if hasattr(anchor._from, 'rowOff'):
                anchor._from.rowOff = original_from_rowOff
            if hasattr(anchor._from, 'colOff'):
                anchor._from.colOff = original_from_colOff
            if hasattr(anchor.to, 'rowOff'):
                anchor.to.rowOff = original_to_rowOff
            if hasattr(anchor.to, 'colOff'):
                anchor.to.colOff = original_to_colOff

            logger.debug(f"DEBUG_IMG_ADJUST: 圖片位置調整完成，偏移: {row_offset} 行")

        except Exception as e:
            logger.debug(f"DEBUG_IMG_ADJUST: 調整圖片錨點位置失敗: {str(e)}")

    def _preserve_table_style(self, table, debug_name: str = "") -> dict:
        """
        保留表格的原始樣式設定

        Args:
            table: Excel表格物件
            debug_name: 用於DEBUG訊息的物件名稱

        Returns:
            dict: 保存的樣式信息
        """
        style_info = {}

        try:
            # 保存表格樣式名稱
            if hasattr(table, 'tableStyleInfo') and table.tableStyleInfo:
                style_info['table_style_name'] = table.tableStyleInfo.name if hasattr(table.tableStyleInfo, 'name') else None
                style_info['show_first_column'] = table.tableStyleInfo.showFirstColumn if hasattr(table.tableStyleInfo, 'showFirstColumn') else False
                style_info['show_last_column'] = table.tableStyleInfo.showLastColumn if hasattr(table.tableStyleInfo, 'showLastColumn') else False
                style_info['show_row_stripes'] = table.tableStyleInfo.showRowStripes if hasattr(table.tableStyleInfo, 'showRowStripes') else True
                style_info['show_column_stripes'] = table.tableStyleInfo.showColumnStripes if hasattr(table.tableStyleInfo, 'showColumnStripes') else False

                logger.debug(f"DEBUG_STYLE: 保留表格 {debug_name} 樣式: {style_info['table_style_name']}")
                logger.debug(f"  - showRowStripes: {style_info['show_row_stripes']}")
                logger.debug(f"  - showFirstColumn: {style_info['show_first_column']}")
                logger.debug(f"  - showLastColumn: {style_info['show_last_column']}")
                logger.debug(f"  - showColumnStripes: {style_info['show_column_stripes']}")
            else:
                logger.debug(f"DEBUG_STYLE: 表格 {debug_name} 沒有 tableStyleInfo，使用預設樣式")
                style_info['table_style_name'] = None
                style_info['show_first_column'] = False
                style_info['show_last_column'] = False
                style_info['show_row_stripes'] = True
                style_info['show_column_stripes'] = False

        except Exception as e:
            logger.debug(f"DEBUG_STYLE: 保留表格樣式失敗: {str(e)}")
            # 設定預設值
            style_info = {
                'table_style_name': None,
                'show_first_column': False,
                'show_last_column': False,
                'show_row_stripes': True,
                'show_column_stripes': False
            }

        return style_info

    def _restore_table_style(self, table, style_info: dict, debug_name: str = ""):
        """
        恢復表格的樣式設定

        Args:
            table: Excel表格物件
            style_info: 保存的樣式信息
            debug_name: 用於DEBUG訊息的物件名稱
        """
        try:
            from openpyxl.worksheet.table import TableStyleInfo

            # 確保有樣式信息可恢復
            if not style_info:
                logger.debug(f"DEBUG_STYLE: 表格 {debug_name} 沒有樣式信息可恢復")
                return

            # 創建或更新 tableStyleInfo
            if not hasattr(table, 'tableStyleInfo') or not table.tableStyleInfo:
                table.tableStyleInfo = TableStyleInfo()
                logger.debug(f"DEBUG_STYLE: 為表格 {debug_name} 創建新的 tableStyleInfo")

            # 恢復樣式屬性
            if style_info.get('table_style_name'):
                table.tableStyleInfo.name = style_info['table_style_name']
                logger.debug(f"DEBUG_STYLE: 恢復表格 {debug_name} 樣式名: {style_info['table_style_name']}")

            table.tableStyleInfo.showFirstColumn = style_info.get('show_first_column', False)
            table.tableStyleInfo.showLastColumn = style_info.get('show_last_column', False)
            table.tableStyleInfo.showRowStripes = style_info.get('show_row_stripes', True)
            table.tableStyleInfo.showColumnStripes = style_info.get('show_column_stripes', False)

            logger.debug(f"DEBUG_STYLE: 恢復表格 {debug_name} 樣式設定完成")
            logger.debug(f"  - showRowStripes: {table.tableStyleInfo.showRowStripes}")
            logger.debug(f"  - showFirstColumn: {table.tableStyleInfo.showFirstColumn}")
            logger.debug(f"  - showLastColumn: {table.tableStyleInfo.showLastColumn}")
            logger.debug(f"  - showColumnStripes: {table.tableStyleInfo.showColumnStripes}")

        except Exception as e:
            logger.debug(f"DEBUG_STYLE: 保留表格樣式設定失敗: {str(e)}")
            import traceback
            logger.debug("例外堆疊", exc_info=True)

    def _find_table_by_position(self, worksheet, target_row, target_col):
        """
        根據位置查找表格物件

        Args:
            worksheet: 工作表
            target_row: 目標行
            target_col: 目標列

        Returns:
            str: 匹配的表格名稱，如果沒有匹配則返回None
        """
        try:
            from openpyxl.utils import range_boundaries

            logger.debug(f"DEBUG_TABLE_MATCH: 尋找位置 ({target_row}, {target_col}) 對應的表格")

            best_match = None
            best_distance = float('inf')

            for table_name in worksheet.tables:
                table = worksheet.tables[table_name]
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)

                logger.debug(f"DEBUG_TABLE_MATCH: 檢查表格 {table_name} - 範圍: 行{min_row}-{max_row}, 列{min_col}-{max_col}")

                # 計算距離用於除錯
                row_distance = abs(target_row - min_row)
                col_distance = abs(target_col - min_col)
                total_distance = row_distance + col_distance
                logger.debug(f"DEBUG_TABLE_MATCH: 表格 {table_name} 距離 - 行距:{row_distance}, 列距:{col_distance}, 總距:{total_distance}")

                # 首先檢查是否目標位置在表格範圍內（優先級最高）
                if (min_row <= target_row <= max_row and
                    min_col <= target_col <= max_col):
                    logger.debug(f"DEBUG_TABLE_MATCH: 目標位置在表格範圍內: {table_name}")
                    return table_name

                # 檢查是否在容錯範圍內
                row_tolerance = 8  # 增加容錯範圍到8行
                col_tolerance = 3  # 增加容錯範圍到3列

                if (row_distance <= row_tolerance and col_distance <= col_tolerance):
                    # 同列匹配給予獎勵（減少距離）
                    adjusted_distance = total_distance
                    if target_col == min_col:
                        adjusted_distance = total_distance * 0.7  # 同列匹配30%獎勵
                        logger.debug(f"DEBUG_TABLE_MATCH: 同列獎勵 {table_name} - 調整距離: {adjusted_distance}")

                    # 檢查是否為最佳匹配
                    if adjusted_distance < best_distance:
                        best_match = table_name
                        best_distance = adjusted_distance
                        logger.debug(f"DEBUG_TABLE_MATCH: 更新最佳匹配: {table_name} (距離: {adjusted_distance})")

            if best_match:
                logger.debug(f"DEBUG_TABLE_MATCH: 找到最佳匹配表格: {best_match}")
                return best_match

            logger.debug(f"DEBUG_TABLE_MATCH: 未找到匹配的表格")
            return None

        except Exception as e:
            logger.debug(f"DEBUG_TABLE_MATCH: 位置匹配發生錯誤: {str(e)}")
            return None

    def fix_table_ranges_post_render(self, worksheet):
        """
        後渲染表格範圍修復 - 檢測並修復所有可能的表格範圍問題

        Args:
            worksheet: 工作表對象
        """
        try:
            logger.debug(f"DEBUG_TABLE_FIX: 開始後渲染表格範圍修復")

            from openpyxl.utils import range_boundaries

            for table_name in worksheet.tables:
                table = worksheet.tables[table_name]
                current_ref = table.ref
                min_col, min_row, max_col, max_row = range_boundaries(current_ref)

                logger.debug(f"DEBUG_TABLE_FIX: 檢查表格 {table_name} - 當前範圍: {current_ref}")

                # 檢查表格範圍內是否有實際數據
                has_data = False
                data_max_row = min_row
                data_max_col = min_col

                # 掃描更大的範圍來找到實際數據邊界
                scan_rows = min(min_row + 20, worksheet.max_row + 1)  # 掃描20行範圍
                scan_cols = min(min_col + 10, worksheet.max_column + 1)  # 掃描10列範圍

                for row in range(min_row, scan_rows):
                    for col in range(min_col, scan_cols):
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value is not None and str(cell_value).strip():
                            has_data = True
                            data_max_row = max(data_max_row, row)
                            data_max_col = max(data_max_col, col)

                if has_data and (data_max_row > max_row or data_max_col > max_col):
                    # 發現數據超出表格範圍，需要修復
                    new_max_row = max(data_max_row, max_row)
                    new_max_col = max(data_max_col, max_col)

                    new_range = f"{self._format_cell_reference(min_row, min_col)}:{self._format_cell_reference(new_max_row, new_max_col)}"

                    logger.debug(f"DEBUG_TABLE_FIX: 發現範圍需要修復: {table_name}")
                    logger.debug(f"DEBUG_TABLE_FIX: 原範圍: {current_ref} -> 新範圍: {new_range}")

                    # 推斷是否有表頭
                    header_row_value = worksheet.cell(row=min_row, column=min_col).value
                    has_header = header_row_value and isinstance(header_row_value, str) and not header_row_value.isdigit()

                    logger.debug(f"DEBUG_TABLE_FIX: 推斷表頭狀態: {has_header}")

                    # 使用現有的同步方法更新範圍
                    self._update_table_range_sync(table, new_range, f"[修復]{table_name}", None, has_header)

                    logger.debug(f"DEBUG_TABLE_FIX: 表格 {table_name} 範圍修復完成")
                else:
                    logger.debug(f"DEBUG_TABLE_FIX: 表格 {table_name} 範圍正常，無需修復")

            logger.debug(f"DEBUG_TABLE_FIX: 後渲染表格範圍修復完成")

        except Exception as e:
            logger.debug(f"DEBUG_TABLE_FIX: 後渲染修復失敗: {str(e)}")
            import traceback
            logger.debug("例外堆疊", exc_info=True)

