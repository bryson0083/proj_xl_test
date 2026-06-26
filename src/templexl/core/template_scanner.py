"""
模板掃描器 - 統一管理模板標籤掃描與註冊表建立功能

此模組集中了以下功能：
1. 模板標籤掃描（第一步）
2. 標籤註冊表建立（第二步）
3. 物件發現與註冊
4. 工作表級別的掃描協調

目的：讓模板掃描與註冊表建立機制內聚在一個模組中，
     便於其他模組調用，也方便維護和測試。
"""
import logging

logger = logging.getLogger(__name__)

import uuid
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from ..models.base import BlockType, ObjectType, CellPosition, DataShape, RangePosition, TagType
from ..models.container import Container
from ..models.objects import Block, ObjectInfo, TableObject, ImageObject
from ..models.tag import Tag
from .parser import TemplateParser


class TemplateScanner:
    """
    模板掃描器類別
    
    統一管理模板標籤掃描與註冊表建立功能，
    確保這些核心功能內聚在一個模組中
    """
    
    def __init__(self):
        """初始化模板掃描器"""
        self.parser = TemplateParser()
    
    def scan_and_register_template(self, workbook: Workbook) -> List[Container]:
        """
        第一步：掃描模板標籤
        第二步：建立標籤註冊表
        
        這是模板掃描的主要入口點，集中了掃描和註冊的完整流程
        
        Args:
            workbook: Excel工作簿物件
            
        Returns:
            List[Container]: 容器物件清單，每個工作表對應一個Container
        """
        # 第一步：掃描所有模板標籤
        all_tags = self._scan_template_tags(workbook)
        
        # 第二步：建立標籤註冊表
        containers = self._build_tag_registry(workbook, all_tags)
        
        return containers
    
    def _scan_template_tags(self, workbook: Workbook) -> List[Tag]:
        """
        第一步：掃描模板中的所有標籤
        
        遍歷所有工作表，掃描工作表下所有模板標籤，
        掃描過程忽略 `{{` 與 `}}` 之間的空白
        
        Args:
            workbook: Excel工作簿物件
            
        Returns:
            List[Tag]: 掃描到的所有標籤清單
        """
        # 使用 TemplateParser 進行標籤掃描
        return self.parser.parse_template(workbook)
    
    def _build_tag_registry(self, workbook: Workbook, tags: List[Tag]) -> List[Container]:
        """
        第二步：建立標籤註冊表
        
        將掃描結果進行標籤註冊表建立，
        每個工作表建立一個Container作為註冊表
        
        Args:
            workbook: Excel工作簿物件
            tags: 掃描到的標籤清單
            
        Returns:
            List[Container]: 容器物件清單
        """
        containers = []
        
        # 按工作表分組標籤
        sheet_tags = self._group_tags_by_sheet(workbook, tags)
        
        # 為每個工作表建立Container註冊表
        for sheet_name in workbook.sheetnames:
            container = self._create_container_registry(
                workbook, sheet_name, sheet_tags.get(sheet_name, [])
            )
            containers.append(container)
            
        return containers
    
    def _group_tags_by_sheet(self, workbook: Workbook, tags: List[Tag]) -> Dict[str, List[Tag]]:
        """
        按工作表分組標籤
        
        Args:
            workbook: Excel工作簿
            tags: 標籤清單
            
        Returns:
            Dict[str, List[Tag]]: 按工作表分組的標籤
        """
        sheet_tags = {}
        
        # 初始化每個工作表的標籤清單
        for sheet_name in workbook.sheetnames:
            sheet_tags[sheet_name] = []
        
        # 根據標籤的sheet_name資訊分組
        for tag in tags:
            if tag.sheet_name in sheet_tags:
                sheet_tags[tag.sheet_name].append(tag)
            
        return sheet_tags
    
    def _create_container_registry(self, workbook: Workbook, sheet_name: str, tags: List[Tag]) -> Container:
        """
        為單一工作表建立Container註冊表
        
        Container作為該工作表的物件註冊表，包含：
        1. 標籤物件清單
        2. 表格物件清單  
        3. 圖片物件清單
        4. 其他物件清單
        
        Args:
            workbook: Excel工作簿
            sheet_name: 工作表名稱
            tags: 該工作表的標籤清單
            
        Returns:
            Container: 容器註冊表物件
        """
        container_id = f"container_{sheet_name}_{uuid.uuid4().hex[:8]}"
        worksheet = workbook[sheet_name]
        
        # 建立物件註冊清單
        objects = []
        
        # 1. 從標籤建立物件
        tag_objects = []
        for tag in tags:
            obj_info = self._create_object_from_tag(tag, sheet_name)
            tag_objects.append(obj_info)
        
        # 2. 掃描並註冊表格物件
        table_objects = self._scan_table_objects(worksheet, sheet_name)
        
        # 3. 檢查並綁定標籤與表格物件
        bound_objects, unbound_tags, unbound_tables = self._bind_tags_to_tables(
            tag_objects, table_objects
        )
        
        # 加入綁定後的物件
        objects.extend(bound_objects)
        # 加入未綁定的標籤物件
        objects.extend(unbound_tags)
        # 加入未綁定的表格物件
        objects.extend(unbound_tables)
        
        # 4. 掃描並註冊圖片物件
        image_objects = self._scan_image_objects(worksheet, sheet_name)
        objects.extend(image_objects)
        
        # 5. 掃描並註冊其他物件（文字內容等）
        text_objects = self._scan_text_content(worksheet, sheet_name, tags)
        objects.extend(text_objects)
        
        # 建立Container註冊表
        container = Container(
            container_id=container_id,
            sheet_name=sheet_name,
            blocks=[],  # 區塊分類將在後續步驟完成
            objects=objects
        )
        
        return container
    
    def _create_object_from_tag(self, tag: Tag, sheet_name: str) -> ObjectInfo:
        """
        從標籤建立物件資訊
        
        Args:
            tag: 標籤物件
            sheet_name: 工作表名稱
            
        Returns:
            ObjectInfo: 物件資訊
        """
        obj_id = f"tag_{tag.tag_name}_{uuid.uuid4().hex[:8]}"
        
        # 根據標籤類型決定物件類型
        obj_type = ObjectType.TABLE if tag.tag_type == TagType.TABLE else ObjectType.SIMPLE
        
        return ObjectInfo(
            obj_id=obj_id,
            display_name=tag.tag_name,
            obj_type=obj_type,
            block_id="",  # 區塊ID將在後續步驟設定
            sheet_name=sheet_name,
            is_multi_rows=tag.tag_type == TagType.TABLE,
            having_header=not (tag.has_condition and tag.condition == "noheader"),
            cell_position=tag.cell_position,
            data_shape=DataShape(rows=1, cols=1),
            obj_name=tag.tag_name  # 預設使用標籤名稱作為obj_name
        )
    
    def _scan_table_objects(self, worksheet, sheet_name: str) -> List[ObjectInfo]:
        """
        掃描工作表中的表格物件
        
        Args:
            worksheet: 工作表物件
            sheet_name: 工作表名稱
            
        Returns:
            List[ObjectInfo]: 表格物件清單
        """
        objects = []
        
        # 掃描Excel Table物件
        if hasattr(worksheet, 'tables') and worksheet.tables:
            # 正確的迭代方式：直接迭代表格名稱，然後獲取表格物件
            for table_name in worksheet.tables:
                table = worksheet.tables[table_name]
                obj_info = self._create_table_object_info(table, table_name, sheet_name)
                objects.append(obj_info)
        
        return objects
    
    def _create_table_object_info(self, table: Table, table_name: str, sheet_name: str) -> ObjectInfo:
        """
        建立表格物件資訊
        
        Args:
            table: Excel表格物件
            table_name: 表格名稱
            sheet_name: 工作表名稱
            
        Returns:
            ObjectInfo: 表格物件資訊
        """
        obj_id = f"table_{table_name}_{uuid.uuid4().hex[:8]}"
        
        # 解析表格範圍
        range_parts = table.ref.split(':')
        start_cell = range_parts[0]
        end_cell = range_parts[1] if len(range_parts) > 1 else start_cell
        
        # 轉換為行列座標
        from openpyxl.utils import cell
        start_row, start_col = cell.coordinate_to_tuple(start_cell)
        end_row, end_col = cell.coordinate_to_tuple(end_cell)
        
        # 建立TableObject
        table_obj = TableObject(
            table_id=table_name,
            table_ref=table.ref,
            header_row_count=1 if table.headerRowCount else 0,
            totals_row_count=table.totalsRowCount or 0,
            table_columns=[col.name for col in table.tableColumns] if table.tableColumns else [],
            auto_filter_ref=table.autoFilter.ref if table.autoFilter else "",
            original_range={
                "start_row": start_row,
                "start_col": start_col,
                "end_row": end_row,
                "end_col": end_col
            }
        )
        
        return ObjectInfo(
            obj_id=obj_id,
            display_name=table_name,
            obj_type=ObjectType.TABLE_OBJ,
            block_id="",  # 區塊ID將在後續步驟設定
            sheet_name=sheet_name,
            is_multi_rows=True,
            having_header=table.headerRowCount > 0 if table.headerRowCount else False,
            cell_position=CellPosition(row=start_row, col=start_col),
            data_shape=DataShape(
                rows=end_row - start_row + 1,
                cols=end_col - start_col + 1
            ),
            obj_name=table_name  # 表格物件的obj_name就是其自身名稱
        )
    
    def _scan_image_objects(self, worksheet, sheet_name: str) -> List[ObjectInfo]:
        """
        掃描工作表中的圖片物件
        
        Args:
            worksheet: 工作表物件
            sheet_name: 工作表名稱
            
        Returns:
            List[ObjectInfo]: 圖片物件清單
        """
        objects = []
        
        # 掃描圖片物件
        if hasattr(worksheet, '_images') and worksheet._images:
            for i, image in enumerate(worksheet._images):
                obj_info = self._create_image_object_info(image, i, sheet_name)
                objects.append(obj_info)
        
        return objects
    
    def _create_image_object_info(self, image, index: int, sheet_name: str) -> ObjectInfo:
        """
        建立圖片物件資訊
        
        Args:
            image: 圖片物件
            index: 圖片索引
            sheet_name: 工作表名稱
            
        Returns:
            ObjectInfo: 圖片物件資訊
        """
        obj_id = f"image_{index}_{uuid.uuid4().hex[:8]}"
        
        # 取得圖片位置資訊
        anchor = image.anchor
        
        # 建立ImageObject（使用正確的RangePosition格式）
        try:
            # 嘗試獲取from位置
            from_pos = None
            for from_attr in ['_from', 'from']:
                if hasattr(anchor, from_attr):
                    from_pos = getattr(anchor, from_attr)
                    break
            
            if not from_pos:
                raise ValueError("無法找到圖片錨點的from位置")
                
            from_row = from_pos.row + 1  # openpyxl uses 0-based indexing
            from_col = from_pos.col + 1
            
            # 嘗試獲取to位置 (TwoCellAnchor)
            to_pos = None
            for to_attr in ['_to', 'to']:
                if hasattr(anchor, to_attr):
                    to_pos = getattr(anchor, to_attr)
                    break
            
            if to_pos:
                # TwoCellAnchor - 有to位置
                to_row = to_pos.row + 1
                to_col = to_pos.col + 1
                
                image_obj = ImageObject(
                    image_id=obj_id,
                    anchor_type="TwoCellAnchor",
                    from_position=RangePosition(row=from_row, col=from_col, row_off=0, col_off=0),
                    to_position=RangePosition(row=to_row, col=to_col, row_off=0, col_off=0)
                )
                
                logger.debug(f"DEBUG: 建立TwoCellAnchor圖片物件 - from:({from_row},{from_col}) to:({to_row},{to_col})")
            else:
                # OneCellAnchor - 沒有to位置
                image_obj = ImageObject(
                    image_id=obj_id,
                    anchor_type="OneCellAnchor",
                    from_position=RangePosition(row=from_row, col=from_col, row_off=0, col_off=0),
                    to_position=None
                )
                
                logger.debug(f"DEBUG: 建立OneCellAnchor圖片物件 - from:({from_row},{from_col})")
                
            cell_pos = CellPosition(row=from_row, col=from_col)
            
        except Exception as e:
            logger.debug(f"DEBUG: 圖片物件建立失敗: {e}")
            # 使用預設位置
            from_row, from_col = 1, 1
            image_obj = ImageObject(
                image_id=obj_id,
                anchor_type="OneCellAnchor",
                from_position=RangePosition(row=from_row, col=from_col, row_off=0, col_off=0),
                to_position=None
            )
            cell_pos = CellPosition(row=from_row, col=from_col)
        
        return ObjectInfo(
            obj_id=obj_id,
            display_name=f"Image_{index}",
            obj_type=ObjectType.IMAGE_OBJ,
            block_id="",  # 區塊ID將在後續步驟設定
            sheet_name=sheet_name,
            is_multi_rows=False,
            having_header=False,
            cell_position=cell_pos,
            data_shape=DataShape(rows=1, cols=1)  # 圖片視為單一物件
        )
    
    def _scan_text_content(self, worksheet, sheet_name: str, tags: List[Tag]) -> List[ObjectInfo]:
        """
        掃描工作表中的非標籤文字內容
        
        Args:
            worksheet: 工作表物件
            sheet_name: 工作表名稱
            tags: 已解析的標籤列表（用於排除標籤位置）
            
        Returns:
            List[ObjectInfo]: 文字物件清單
        """
        objects = []
        
        # 建立標籤位置集合，用於排除
        tag_positions = set()
        for tag in tags:
            if tag.sheet_name == sheet_name:
                tag_positions.add((tag.cell_position.row, tag.cell_position.col))
        
        # 掃描所有有內容的儲存格
        for row in worksheet.iter_rows():
            for cell in row:
                if (cell.value and 
                    isinstance(cell.value, str) and 
                    (cell.row, cell.column) not in tag_positions):
                    
                    # 檢查是否包含未識別的標籤語法
                    if '{{' in cell.value and '}}' in cell.value:
                        continue  # 跳過包含標籤語法的儲存格
                    
                    # 建立文字物件
                    obj_info = self._create_text_object_info(cell, sheet_name)
                    objects.append(obj_info)
        
        return objects
    
    def _create_text_object_info(self, cell, sheet_name: str) -> ObjectInfo:
        """
        建立文字物件資訊
        
        Args:
            cell: 儲存格物件
            sheet_name: 工作表名稱
            
        Returns:
            ObjectInfo: 文字物件資訊
        """
        obj_id = f"text_{cell.row}_{cell.column}_{uuid.uuid4().hex[:8]}"
        
        return ObjectInfo(
            obj_id=obj_id,
            display_name=f"Text_{cell.row}_{cell.column}",
            obj_type=ObjectType.TEXT,
            block_id="",  # 區塊ID將在後續步驟設定
            sheet_name=sheet_name,
            is_multi_rows=False,
            having_header=False,
            cell_position=CellPosition(row=cell.row, col=cell.column),
            data_shape=DataShape(rows=1, cols=1)
        )
    
    def get_container_by_sheet(self, containers: List[Container], sheet_name: str) -> Container:
        """
        根據工作表名稱取得對應的Container註冊表
        
        Args:
            containers: 容器清單
            sheet_name: 工作表名稱
            
        Returns:
            Container: 對應的容器註冊表
            
        Raises:
            ValueError: 找不到對應的工作表
        """
        for container in containers:
            if container.sheet_name == sheet_name:
                return container
                
        raise ValueError(f"找不到工作表 '{sheet_name}' 對應的Container註冊表")
    
    def show_scan_summary(self, containers: List[Container]) -> None:
        """
        顯示掃描結果摘要
        
        Args:
            containers: 容器清單
        """
        logger.debug(f"=== 模板掃描結果摘要 ===")
        logger.debug(f"掃描到 {len(containers)} 個工作表")
        
        total_tags = 0
        total_objects = 0
        
        for container in containers:
            logger.debug(f"\n工作表: {container.sheet_name}")
            logger.debug(f"  容器ID: {container.container_id}")
            logger.debug(f"  物件數量: {len(container.objects)}")
            
            # 統計各類型物件
            simple_tag_count = len([obj for obj in container.objects if obj.obj_type == ObjectType.SIMPLE])
            table_tag_count = len([obj for obj in container.objects if obj.obj_type == ObjectType.TABLE])
            table_count = len([obj for obj in container.objects if obj.obj_type == ObjectType.TABLE_OBJ])
            image_count = len([obj for obj in container.objects if obj.obj_type == ObjectType.IMAGE_OBJ])
            text_count = len([obj for obj in container.objects if obj.obj_type == ObjectType.TEXT])
            
            logger.debug(f"    簡單標籤: {simple_tag_count}")
            logger.debug(f"    表格標籤: {table_tag_count}")
            logger.debug(f"    表格物件: {table_count}")
            logger.debug(f"    圖片物件: {image_count}")
            logger.debug(f"    文字物件: {text_count}")
            
            total_tags += simple_tag_count + table_tag_count
            total_objects += len(container.objects)
        
        logger.debug(f"\n總計:")
        logger.debug(f"  標籤總數: {total_tags}")
        logger.debug(f"  物件總數: {total_objects}")
    
    def _bind_tags_to_tables(self, tag_objects: List[ObjectInfo], 
                            table_objects: List[ObjectInfo]) -> tuple:
        """
        檢查並綁定標籤與表格物件
        
        如果標籤位置在表格物件範圍內，則將它們綁定在一起
        
        Args:
            tag_objects: 標籤物件清單
            table_objects: 表格物件清單
            
        Returns:
            tuple: (綁定物件清單, 未綁定標籤清單, 未綁定表格清單)
        """
        bound_objects = []
        unbound_tags = []
        unbound_tables = []
        bound_table_ids = set()
        
        # 檢查每個標籤是否在任何表格物件範圍內
        for tag_obj in tag_objects:
            bound = False
            
            for table_obj in table_objects:
                if self._is_tag_in_table_range(tag_obj, table_obj):
                    # 建立綁定物件
                    bound_obj = self._create_bound_object(tag_obj, table_obj)
                    bound_objects.append(bound_obj)
                    bound_table_ids.add(table_obj.obj_id)
                    bound = True
                    break
            
            if not bound:
                unbound_tags.append(tag_obj)
        
        # 收集未被綁定的表格物件
        for table_obj in table_objects:
            if table_obj.obj_id not in bound_table_ids:
                unbound_tables.append(table_obj)
        
        return bound_objects, unbound_tags, unbound_tables
    
    def _is_tag_in_table_range(self, tag_obj: ObjectInfo, 
                               table_obj: ObjectInfo) -> bool:
        """
        檢查標籤是否在表格物件範圍內
        
        Args:
            tag_obj: 標籤物件
            table_obj: 表格物件
            
        Returns:
            bool: 是否在範圍內
        """
        tag_row = tag_obj.cell_position.row
        tag_col = tag_obj.cell_position.col
        
        table_start_row = table_obj.cell_position.row
        table_start_col = table_obj.cell_position.col
        table_end_row = table_start_row + table_obj.data_shape.rows - 1
        table_end_col = table_start_col + table_obj.data_shape.cols - 1
        
        return (table_start_row <= tag_row <= table_end_row and
                table_start_col <= tag_col <= table_end_col)
    
    def _create_bound_object(self, tag_obj: ObjectInfo, 
                            table_obj: ObjectInfo) -> ObjectInfo:
        """
        建立綁定後的物件
        
        保留標籤的display_name，使用表格物件的位置和範圍
        
        Args:
            tag_obj: 標籤物件
            table_obj: 表格物件
            
        Returns:
            ObjectInfo: 綁定後的物件
        """
        # 如果標籤有noheader條件，則having_header為False，否則使用表格物件的having_header
        # 標籤物件的having_header已經在_create_object_from_tag中根據noheader條件設定
        having_header = tag_obj.having_header if tag_obj.obj_type == ObjectType.TABLE else table_obj.having_header
        
        return ObjectInfo(
            obj_id=tag_obj.obj_id,  # 保留標籤的ID
            display_name=tag_obj.display_name,  # 保留標籤的顯示名稱
            obj_type=ObjectType.TABLE_OBJ,  # 改為TABLE_OBJ類型
            block_id=table_obj.block_id,
            sheet_name=table_obj.sheet_name,
            is_multi_rows=table_obj.is_multi_rows,
            having_header=having_header,  # 考慮標籤的條件
            cell_position=table_obj.cell_position,  # 使用表格的位置
            data_shape=table_obj.data_shape,  # 使用表格的範圍
            obj_name=table_obj.display_name  # 記錄原始表格物件名稱
        )
